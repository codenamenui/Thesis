#!/usr/bin/env python3
"""
merge_results_comprehensive_final.py
Extended with learning curves, hyperparameters, confusion patterns,
and deduplicated seed‑level test metrics from all trials_log.json files.
"""
import argparse
import json
import os
import re
import warnings
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
import matplotlib.ticker as mticker
from sklearn.metrics import precision_recall_fscore_support, accuracy_score

matplotlib.rcParams['figure.dpi']      = 300
matplotlib.rcParams['savefig.dpi']     = 300
matplotlib.rcParams['font.family']     = 'sans-serif'
matplotlib.rcParams['font.sans-serif'] = ['Arial']

warnings.filterwarnings("ignore")

# ----------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------

SUBCLASSES = ["HR", "AI-R", "HF", "AI-F"]

CONDITIONS = {
    "Condition A": "HR100",
    "Condition B": "HR67",
    "Condition C": "HR50",
}
CONDITION_LABELS = {
    "Condition A": "Condition A: Human-Real 100%",
    "Condition B": "Condition B: Human-Real 67%",
    "Condition C": "Condition C: Human-Real 50%",
}
CONDITION_SHORT = {
    "Condition A": "CondA",
    "Condition B": "CondB",
    "Condition C": "CondC",
}
CONDITION_ORDER = {v: i for i, v in enumerate(CONDITION_SHORT.values())}
ARCH_ORDER = {"Tagalog BERT": 0, "Tagalog DistilBERT": 1}

LABEL_STR_TO_INT = {"real": 0, "fake": 1}

SUBPLOT_ORDER = ["HF", "AI-F", "HR", "AI-R"]
SUBPLOT_TITLES = {
    "HF":   "HF (Human Fake)",
    "AI-F": "AI-F (AI-Generated Fake)",
    "HR":   "HR (Human Real)",
    "AI-R": "AI-R (AI-Enhanced Real)",
}

HF_TICKS    = [0, 33, 50, 67, 100]
HF_TICKLBLS = ["0%", "33%", "50%", "67%", "100%"]
HF_PCT_POS  = {0: 0, 33: 1, 50: 2, 67: 3, 100: 4}
HF_POS_TICKS = list(HF_PCT_POS.values())

ARCH_STYLE = {
    "Tagalog BERT": {
        "color":  "#1f77b4",
        "marker": "o",
        "lw":     2.2,
        "ms":     7,
    },
    "Tagalog DistilBERT": {
        "color":  "#d62728",
        "marker": "s",
        "lw":     2.2,
        "ms":     7,
    },
}

COND_STYLE = {
    "CondA": {
        "label":  "Condition A (Human-Real 100%)",
        "color":  "#1f77b4",
        "marker": "o",
        "lw":     2.2,
        "ms":     7,
    },
    "CondB": {
        "label":  "Condition B (Human-Real 67%)",
        "color":  "#ff7f0e",
        "marker": "s",
        "lw":     2.2,
        "ms":     7,
    },
    "CondC": {
        "label":  "Condition C (Human-Real 50%)",
        "color":  "#2ca02c",
        "marker": "^",
        "lw":     2.2,
        "ms":     7,
    },
}

ARCH_LS = {
    "Tagalog BERT":       "-",
    "Tagalog DistilBERT": "--",
}

FIXED_HPARAMS = {
    "Optimizer":                          "AdamW",
    "Learning rate schedule":             "Linear warmup + linear decay",
    "Warmup ratio":                       "10% of total steps",
    "Weight decay":                       "0.01",
    "Early stopping patience":            "2 epochs",
    "Gradient clipping":                  "Max norm 1.0",
    "Dropout (classification head)":      "0.1",
    "Mixed precision":                    "FP16",
    "Maximum sequence length":            "512 tokens",
}

# Valid real news conditions (used to filter unofficial models)
VALID_REAL_CONDITIONS = [
    "HR100-AIR0",
    "HR67-AIR33",
    "HR50-AIR50",
]

# ----------------------------------------------------------------------
# Helper functions
# ----------------------------------------------------------------------

def extract_architecture(model_key: str) -> str:
    return "Tagalog DistilBERT" if "distilbert" in model_key.lower() else "Tagalog BERT"

def extract_hf_label(model_key: str) -> str:
    key   = model_key.lower()
    hf_m  = re.search(r'hf(\d+)',  key)
    aif_m = re.search(r'aif(\d+)', key)
    hf    = int(hf_m.group(1))  if hf_m  else 0
    aif   = int(aif_m.group(1)) if aif_m else (100 - hf)
    return f"HF{hf}-AIF{aif}"

def extract_hf_pct(model_key_or_hf_label: str) -> int:
    m = re.search(r'hf(\d+)', model_key_or_hf_label, re.IGNORECASE)
    return int(m.group(1)) if m else 0

def format_model_display_label(model_key: str) -> str:
    key  = model_key.lower()
    def _get(pattern):
        m = re.search(pattern, key, re.IGNORECASE)
        return int(m.group(1)) if m else None
    hr  = _get(r'(?<!ai)hr(\d+)')
    air = _get(r'air(\d+)')
    hf  = _get(r'hf(\d+)')
    aif = _get(r'aif(\d+)')
    if hr  is not None and air is None: air = 100 - hr
    elif air is not None and hr is None: hr  = 100 - air
    if hf  is not None and aif is None: aif = 100 - hf
    elif aif is not None and hf is None: hf  = 100 - aif
    hr = hr or 0; air = air or 0; hf = hf or 0; aif = aif or 0
    return f"HR{hr}-AIR{air}-HF{hf}-AIF{aif}"

def is_official_model(model_key: str) -> bool:
    """Return True if the model belongs to the 30-model experimental matrix."""
    label = format_model_display_label(model_key)
    parts = label.split("-")
    if len(parts) != 4:
        return False
    real_part = f"{parts[0]}-{parts[1]}"
    if real_part not in VALID_REAL_CONDITIONS:
        return False
    hf_val = int(parts[2].replace("HF", ""))
    aif_val = int(parts[3].replace("AIF", ""))
    return hf_val + aif_val == 100

def compute_logical_sort_key(model_identifier):
    mid  = model_identifier.lower()
    arch = 1 if "distilbert" in mid else 0
    hf_m = re.search(r'hf(\d+)', mid)
    hf   = int(hf_m.group(1)) if hf_m else 0
    return (arch, -hf)

def standardize_and_sort_thesis_data(df):
    if df.empty:
        return df
    if 'Model Identifier' in df.columns:
        df['_sort_key'] = df['Model Identifier'].apply(compute_logical_sort_key)
        df['Model Identifier'] = df['Model Identifier'].apply(format_model_display_label)
        df = df.sort_values('_sort_key').drop(columns=['_sort_key'])
    return df

def _run_mcnemar_pair(res_1, res_2):
    from statsmodels.stats.contingency_tables import mcnemar
    n = min(len(res_1), len(res_2))
    if n == 0:
        return None
    gt = res_1['true_int'].values[:n]
    p1 = res_1['pred_int'].values[:n]
    p2 = res_2['pred_int'].values[:n]
    b = int(np.sum((p1 == gt) & (p2 != gt)))
    c = int(np.sum((p1 != gt) & (p2 == gt)))
    return mcnemar([[0, b], [c, 0]], exact=(b + c < 25)).pvalue

def _apply_bh(df):
    from statsmodels.stats.multitest import multipletests
    if df.empty:
        return df
    _, p_adj, _, _ = multipletests(df['p-value'], alpha=0.05, method='fdr_bh')
    df = df.copy()
    df['p-adj (BH)']  = p_adj
    df['Significant'] = p_adj < 0.05
    return df

def _format_pvalue(p, threshold=1e-4):
    if pd.isna(p):
        return ""
    if p == 0:
        return "0.0"
    if p < threshold:
        return f"{p:.2e}"
    else:
        return f"{p:.4f}"

def _format_pvalue_columns(df):
    for col in ['p-value', 'p-adj (BH)']:
        if col in df.columns:
            df[col] = df[col].apply(_format_pvalue)
    return df

def _arch_sort_val(arch_str):
    return 0 if "distilbert" not in arch_str.lower() else 1

def run_within_condition_mcnemar(condition_df):
    models = condition_df['model_key'].unique()
    rows   = []
    for i in range(len(models)):
        for j in range(i + 1, len(models)):
            m1, m2 = models[i], models[j]
            if extract_architecture(m1) != extract_architecture(m2):
                continue
            res_1 = condition_df[condition_df['model_key'] == m1].sort_values('sample_index')
            res_2 = condition_df[condition_df['model_key'] == m2].sort_values('sample_index')
            pval = _run_mcnemar_pair(res_1, res_2)
            if pval is None:
                continue
            hf_label_1 = extract_hf_label(m1)
            hf_label_2 = extract_hf_label(m2)
            if extract_hf_pct(hf_label_1) < extract_hf_pct(hf_label_2):
                hf_label_1, hf_label_2 = hf_label_2, hf_label_1
            rows.append({
                "Architecture":  extract_architecture(m1),
                "Comparison":    f"{hf_label_1} vs {hf_label_2}",
                "p-value":       pval,
                "_arch_order":   _arch_sort_val(extract_architecture(m1)),
                "_hf_left":      -extract_hf_pct(hf_label_1),
                "_hf_right":     -extract_hf_pct(hf_label_2),
            })
    out = _apply_bh(pd.DataFrame(rows))
    if out.empty:
        return out
    out = (out.sort_values(["_arch_order", "_hf_left", "_hf_right"])
              .drop(columns=["_arch_order", "_hf_left", "_hf_right"])
              .reset_index(drop=True))
    out = _format_pvalue_columns(out)
    return out

def run_cross_condition_mcnemar(total_df):
    df = total_df.copy()
    def _tag(mk):
        ml = mk.lower()
        for ck, hr in CONDITIONS.items():
            if hr.lower() in ml:
                return ck
        return None
    df['_cond']     = df['model_key'].apply(_tag)
    df['_arch']     = df['model_key'].apply(extract_architecture)
    df['_hf_label'] = df['model_key'].apply(extract_hf_label)
    cond_keys = list(CONDITIONS.keys())
    rows      = []
    for (arch, hf_label), grp in df.groupby(['_arch', '_hf_label']):
        for ci in range(len(cond_keys)):
            for cj in range(ci + 1, len(cond_keys)):
                ck_i, ck_j = cond_keys[ci], cond_keys[cj]
                r_i = grp[grp['_cond'] == ck_i].sort_values('sample_index')
                r_j = grp[grp['_cond'] == ck_j].sort_values('sample_index')
                if r_i.empty or r_j.empty:
                    continue
                pval = _run_mcnemar_pair(r_i, r_j)
                if pval is None:
                    continue
                short_i = CONDITION_SHORT[ck_i]
                short_j = CONDITION_SHORT[ck_j]
                rows.append({
                    "Architecture":  arch,
                    "Comparison":    f"{hf_label} [{short_i}] vs [{short_j}]",
                    "p-value":       pval,
                    "_arch_order":   _arch_sort_val(arch),
                    "_hf_pct":       -extract_hf_pct(hf_label),
                    "_cond_left":    CONDITION_ORDER.get(short_i, 99),
                    "_cond_right":   CONDITION_ORDER.get(short_j, 99),
                })
    out = _apply_bh(pd.DataFrame(rows))
    if out.empty:
        return out
    out = (out.sort_values(["_arch_order", "_hf_pct", "_cond_left", "_cond_right"])
              .drop(columns=["_arch_order", "_hf_pct", "_cond_left", "_cond_right"])
              .reset_index(drop=True))
    out = _format_pvalue_columns(out)
    return out

def run_champion_mcnemar(total_df, top_n=3):
    rows = []
    for cond_key, hr_string in CONDITIONS.items():
        cond_df = total_df[total_df['model_key'].str.contains(hr_string, case=False)]
        for arch in ['Tagalog BERT', 'Tagalog DistilBERT']:
            arch_df = cond_df[cond_df['model_key'].apply(extract_architecture) == arch]
            if arch_df.empty:
                continue
            model_acc = {
                mk: accuracy_score(grp['true_int'], grp['pred_int'])
                for mk, grp in arch_df.groupby('model_key')
            }
            top_models = sorted(model_acc, key=model_acc.get, reverse=True)[:top_n]
            if len(top_models) < 2:
                continue
            for i in range(len(top_models)):
                for j in range(i + 1, len(top_models)):
                    m1, m2 = top_models[i], top_models[j]
                    res_1 = arch_df[arch_df['model_key'] == m1].sort_values('sample_index')
                    res_2 = arch_df[arch_df['model_key'] == m2].sort_values('sample_index')
                    pval = _run_mcnemar_pair(res_1, res_2)
                    if pval is None:
                        continue
                    acc1 = model_acc[m1]
                    acc2 = model_acc[m2]
                    lbl1 = f"{extract_hf_label(m1)} (Acc={acc1:.4f})"
                    lbl2 = f"{extract_hf_label(m2)} (Acc={acc2:.4f})"
                    rows.append({
                        "Condition":    CONDITION_SHORT[cond_key],
                        "Architecture": arch,
                        "Comparison":   f"{lbl1} vs {lbl2}",
                        "p-value":      pval,
                        "_cond_order":  CONDITION_ORDER.get(CONDITION_SHORT[cond_key], 99),
                        "_arch_order":  _arch_sort_val(arch),
                        "_acc_left":    -acc1,
                        "_acc_right":   -acc2,
                    })
    out = _apply_bh(pd.DataFrame(rows))
    if out.empty:
        return out
    out = (out.sort_values(["_cond_order", "_arch_order", "_acc_left", "_acc_right"])
              .drop(columns=["_cond_order", "_arch_order", "_acc_left", "_acc_right"])
              .reset_index(drop=True))
    out = _format_pvalue_columns(out)
    return out

def _build_trend_summary(total_df):
    records = []
    for model_key, grp in total_df.groupby('model_key'):
        arch    = extract_architecture(model_key)
        hf_pct  = extract_hf_pct(model_key)
        cond_key = None
        for ck, hr_str in CONDITIONS.items():
            if hr_str.lower() in model_key.lower():
                cond_key = ck
                break
        if cond_key is None:
            continue
        row = {
            'model_key':  model_key,
            'arch':       arch,
            'hf_pct':     hf_pct,
            'condition':  cond_key,
            'cond_short': CONDITION_SHORT[cond_key],
        }
        for sc in SUBCLASSES:
            sg = grp[grp['subclass'] == sc]
            row[f'{sc}_acc'] = (
                accuracy_score(sg['true_int'], sg['pred_int']) * 100
                if not sg.empty else np.nan
            )
        records.append(row)
    return pd.DataFrame(records).sort_values(['condition', 'arch', 'hf_pct'])

def _style_trend_axes(ax, title):
    ax.set_title(title, fontsize=12, fontweight='bold', pad=6)
    ax.set_xlabel('HF Portion', fontsize=10, labelpad=4)
    ax.set_ylabel('Acc (%)', fontsize=10, labelpad=4)
    ax.set_xticks(HF_POS_TICKS)
    ax.set_xticklabels(HF_TICKLBLS, fontsize=9)
    ax.tick_params(axis='y', labelsize=9)
    ax.grid(True, color='#e0e0e0', linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    ax.spines[['top', 'right']].set_visible(False)
    ax.set_ylim(bottom=0, top=105)

# ----------------------------------------------------------------------
# Chart generation (trend lines, bars)
# ----------------------------------------------------------------------

def generate_per_condition_trend_charts(summary_df, out_dir):
    FIG_SIZE = (11, 8.5)
    for cond_key, cond_label in CONDITION_LABELS.items():
        cond_df = summary_df[summary_df['condition'] == cond_key]
        if cond_df.empty:
            continue
        fig, axes = plt.subplots(2, 2, figsize=FIG_SIZE,
                                 gridspec_kw={'hspace': 0.45, 'wspace': 0.28})
        fig.subplots_adjust(top=0.80)
        ax_flat = axes.flatten()
        for idx, sc in enumerate(SUBPLOT_ORDER):
            ax = ax_flat[idx]
            for arch, style in ARCH_STYLE.items():
                sub = cond_df[cond_df['arch'] == arch].sort_values('hf_pct')
                if sub.empty:
                    continue
                x_vals = sub['hf_pct'].map(HF_PCT_POS)
                ax.plot(x_vals, sub[f'{sc}_acc'],
                        color=style['color'], marker=style['marker'],
                        linewidth=style['lw'], markersize=style['ms'],
                        label=arch, zorder=3)
            _style_trend_axes(ax, SUBPLOT_TITLES[sc])
        handles = [
            mlines.Line2D([], [], color=style['color'], marker=style['marker'],
                          linewidth=style['lw'], markersize=style['ms'], label=arch)
            for arch, style in ARCH_STYLE.items()
        ]
        fig.legend(handles=handles, loc='upper center', bbox_to_anchor=(0.5, 0.94),
                   ncol=2, fontsize=10, framealpha=0.9, edgecolor='#cccccc', handlelength=2.5)
        fig.suptitle(f'Subclass Accuracy Across Training Ratios\n{cond_label}',
                     fontsize=14, fontweight='bold', y=0.885)
        fname = os.path.join(out_dir, f'{cond_key}_RatioTrend.png')
        fig.savefig(fname, bbox_inches='tight', dpi=300)
        plt.close(fig)
        print(f'  Saved: {fname}')

def generate_per_arch_condition_comparison_charts(summary_df, out_dir):
    FIG_SIZE = (11, 8.5)
    for arch in ['Tagalog BERT', 'Tagalog DistilBERT']:
        arch_df = summary_df[summary_df['arch'] == arch]
        if arch_df.empty:
            continue
        fig, axes = plt.subplots(2, 2, figsize=FIG_SIZE,
                                 gridspec_kw={'hspace': 0.45, 'wspace': 0.28})
        fig.subplots_adjust(top=0.80)
        ax_flat = axes.flatten()
        for idx, sc in enumerate(SUBPLOT_ORDER):
            ax = ax_flat[idx]
            for ck in CONDITIONS:
                short = CONDITION_SHORT[ck]
                style = COND_STYLE[short]
                sub = arch_df[arch_df['condition'] == ck].sort_values('hf_pct')
                if sub.empty:
                    continue
                x_vals = sub['hf_pct'].map(HF_PCT_POS)
                ax.plot(x_vals, sub[f'{sc}_acc'],
                        color=style['color'], linestyle='-',
                        marker=style['marker'], linewidth=style['lw'],
                        markersize=style['ms'], alpha=1.0,
                        label=CONDITION_LABELS[ck], zorder=3)
            _style_trend_axes(ax, SUBPLOT_TITLES[sc])
        handles = [
            mlines.Line2D([], [], color=COND_STYLE[CONDITION_SHORT[ck]]['color'],
                          linestyle='-',
                          marker=COND_STYLE[CONDITION_SHORT[ck]]['marker'],
                          linewidth=2.2, markersize=7,
                          label=CONDITION_LABELS[ck])
            for ck in CONDITIONS
        ]
        fig.legend(handles=handles, loc='upper center', bbox_to_anchor=(0.5, 0.94),
                   ncol=3, fontsize=9.5, framealpha=0.9, edgecolor='#cccccc', handlelength=2.5)
        fig.suptitle(f'Subclass Accuracy Across Training Ratios\n{arch}',
                     fontsize=14, fontweight='bold', y=0.885)
        safe_arch = arch.replace(' ', '_')
        fname = os.path.join(out_dir, f'{safe_arch}_ConditionComparison_RatioTrend.png')
        fig.savefig(fname, bbox_inches='tight', dpi=300)
        plt.close(fig)
        print(f'  Saved: {fname}')

def generate_overall_accuracy_bars(overall_df, cond_key, out_dir):
    cond_title = CONDITION_LABELS.get(cond_key, cond_key)
    bar_data = overall_df.iloc[::-1].copy()
    arch_color = {
        "Tagalog BERT":       ARCH_STYLE["Tagalog BERT"]["color"],
        "Tagalog DistilBERT": ARCH_STYLE["Tagalog DistilBERT"]["color"],
    }
    model_order = bar_data["Model Identifier"].drop_duplicates().tolist()
    acc_lookup = {}
    for _, row in bar_data.iterrows():
        acc_lookup[(row["Model Identifier"], row["Architecture"])] = row["Accuracy"]
    n_groups = len(model_order)
    y = np.arange(n_groups)
    bar_height = 0.25
    fig, ax = plt.subplots(figsize=(14, max(6, n_groups * 1.2 + 2)))
    for arch, offset in [("Tagalog BERT", +bar_height/2),
                         ("Tagalog DistilBERT", -bar_height/2)]:
        vals = [acc_lookup.get((model, arch), 0) for model in model_order]
        bars = ax.barh(y + offset, vals, height=bar_height,
                       color=arch_color[arch], edgecolor="#333333", linewidth=0.6,
                       label=arch)
        for bar in bars:
            w = bar.get_width()
            ax.text(w + 0.008, bar.get_y() + bar.get_height()/2,
                    f"{w:.4f}", va="center", ha="left",
                    fontsize=9, fontweight="bold", color="#222222")
    ax.set_yticks(y)
    ax.set_yticklabels(model_order, fontsize=9)
    ax.set_xlabel("Overall Binary Accuracy", fontsize=12, fontweight="bold", labelpad=8)
    ax.set_title(f"Overall Model Accuracy: {cond_title}", fontsize=15, fontweight="bold", pad=16)
    ax.set_xlim(0, 1.15)
    ax.xaxis.set_major_formatter(mticker.FormatStrFormatter("%.2f"))
    ax.legend(loc="lower right", framealpha=0.9)
    ax.spines[["top", "right"]].set_visible(False)
    ax.invert_yaxis()
    plt.tight_layout()
    plt.savefig(os.path.join(out_dir, f"{cond_key}_Accuracy_Bars.png"), bbox_inches="tight")
    plt.close()

def apply_journal_table_styling(writer, sheet_id, df):
    from openpyxl.styles import Font, Alignment, Border, Side
    ws = writer.sheets[sheet_id]
    thin = Side(style='thin', color="000000")
    hdr_border = Border(top=thin, bottom=thin)
    bot_border = Border(bottom=thin)
    n_cols = len(df.columns)
    n_rows = ws.max_row
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, name="Arial", size=11)
        cell.border = hdr_border
        cell.alignment = Alignment(horizontal="center")
    for r in range(2, n_rows + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=11)
            cell.alignment = Alignment(horizontal="center")
            cell.border = bot_border if r == n_rows else Border()

# ----------------------------------------------------------------------
# Learning curve helpers
# ----------------------------------------------------------------------

def extract_losses_from_final_results_data(data: dict) -> Dict[str, Dict[str, np.ndarray]]:
    losses = {}
    for model_key, model_data in data.items():
        seed_results = model_data.get("seed_results", [])
        if not seed_results:
            continue
        all_epochs = []
        for seed_res in seed_results:
            history = seed_res.get("history", [])
            for entry in history:
                epoch = entry.get("epoch")
                train_loss = entry.get("train_loss")
                val_loss = entry.get("val_loss")
                if epoch is not None and train_loss is not None and val_loss is not None:
                    all_epochs.append({
                        "epoch": epoch,
                        "train_loss": train_loss,
                        "val_loss": val_loss
                    })
        if not all_epochs:
            continue
        df = pd.DataFrame(all_epochs)
        avg = df.groupby("epoch").agg(
            train_loss=("train_loss", "mean"),
            val_loss=("val_loss", "mean")
        ).reset_index().sort_values("epoch")
        losses[model_key] = {
            "epoch":       avg["epoch"].values.astype(int),
            "train_loss":  avg["train_loss"].values,
            "val_loss":    avg["val_loss"].values,
        }
    return losses

def merge_multiple_final_results(json_paths):
    all_losses = {}
    for jpath in json_paths:
        with open(jpath, 'r') as f:
            data = json.load(f)
        file_losses = extract_losses_from_final_results_data(data)
        for mk, curves in file_losses.items():
            if mk not in all_losses:
                all_losses[mk] = curves
    return all_losses

def generate_learning_curve_plots_from_dict(losses_dict: Dict[str, Dict], out_dir: str):
    if not losses_dict:
        print("  No losses data provided.")
        return

    cond_models = {ck: {} for ck in CONDITIONS}
    for mkey, curves in losses_dict.items():
        for cond, hr_str in CONDITIONS.items():
            if hr_str.lower() in mkey.lower():
                cond_models[cond][mkey] = curves
                break

    for cond, model_dict in cond_models.items():
        if not model_dict:
            continue
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5.5))
        for mkey, curves in model_dict.items():
            label = extract_hf_label(mkey)
            ax1.plot(curves['epoch'], curves['train_loss'],
                     label=label, alpha=0.8, linewidth=1.8)
        ax1.set_title(f'{cond}: Training Loss', fontsize=13, fontweight='bold')
        ax1.set_xlabel('Epoch')
        ax1.set_ylabel('Loss')
        ax1.grid(True, alpha=0.3)
        ax1.legend(fontsize=8, ncol=2, loc='upper right')

        for mkey, curves in model_dict.items():
            if len(curves['val_loss']) == 0:
                continue
            label = extract_hf_label(mkey)
            ax2.plot(curves['epoch'], curves['val_loss'],
                     label=label, alpha=0.8, linewidth=1.8, marker='o')
        ax2.set_title(f'{cond}: Validation Loss', fontsize=13, fontweight='bold')
        ax2.set_xlabel('Epoch')
        ax2.set_ylabel('Loss')
        ax2.grid(True, alpha=0.3)
        ax2.legend(fontsize=8, ncol=2, loc='upper right')

        fig.suptitle(f'Learning Curves: {CONDITION_LABELS[cond]}',
                     fontsize=15, fontweight='bold', y=1.02)
        plt.tight_layout()
        fname = os.path.join(out_dir, f'{cond}_LearningCurves.png')
        fig.savefig(fname, bbox_inches='tight', dpi=300)
        plt.close(fig)
        print(f'  Saved learning curves: {fname}')

def build_learning_curve_excel_data(losses_dict: Dict[str, Dict]) -> pd.DataFrame:
    rows = []
    for model_key, curves in losses_dict.items():
        cond = None
        for ck, hr_str in CONDITIONS.items():
            if hr_str.lower() in model_key.lower():
                cond = ck
                break
        if cond is None:
            cond = "Unknown"
        display_label = format_model_display_label(model_key)
        epochs = curves['epoch']
        train_loss = curves['train_loss']
        val_loss = curves['val_loss']
        for i in range(len(epochs)):
            rows.append({
                "Condition":        cond,
                "Model Key":        model_key,
                "Display Label":    display_label,
                "Epoch":            int(epochs[i]),
                "Train Loss":       round(float(train_loss[i]), 6),
                "Val Loss":         round(float(val_loss[i]), 6) if i < len(val_loss) else None
            })
    return pd.DataFrame(rows)

# ----------------------------------------------------------------------
# Hyperparameter extraction (from training_meta.json)
# ----------------------------------------------------------------------

def load_hyperparams_from_models_dir(models_dir: str) -> Dict[str, Dict]:
    if not os.path.isdir(models_dir):
        print(f"Warning: models_dir '{models_dir}' does not exist.")
        return {}

    hyperparams = {}
    pattern = re.compile(r'^(bert|distilbert)__', re.IGNORECASE)

    for entry in os.listdir(models_dir):
        full_path = os.path.join(models_dir, entry)
        if not os.path.isdir(full_path):
            continue
        if not pattern.match(entry):
            continue

        meta_path = os.path.join(full_path, "final", "training_meta.json")
        if not os.path.isfile(meta_path):
            print(f"Warning: No training_meta.json found for model '{entry}'")
            continue

        try:
            with open(meta_path, 'r') as f:
                meta = json.load(f)
            hp = meta.get("hparams")
            if isinstance(hp, dict):
                hyperparams[entry] = hp
            else:
                print(f"Warning: 'hparams' missing or not dict in {meta_path}")
        except (json.JSONDecodeError, OSError) as e:
            print(f"Warning: Could not read {meta_path}: {e}")

    return hyperparams

def build_combined_hyperparameter_dataframe(hyper_dict: Dict[str, Dict]) -> pd.DataFrame:
    rows = []
    for model_key, hp in hyper_dict.items():
        cond = None
        for ck, hr_str in CONDITIONS.items():
            if hr_str.lower() in model_key.lower():
                cond = ck
                break
        if cond is None:
            cond = "Unknown"
        architecture = extract_architecture(model_key)
        display_label = format_model_display_label(model_key)

        row = {
            "Condition":      cond,
            "Model Key":      model_key,
            "Display Label":  display_label,
            "Architecture":   architecture,
        }
        for k, v in sorted(hp.items()):
            row[k] = v
        for fixed_key, fixed_val in FIXED_HPARAMS.items():
            row[fixed_key] = fixed_val
        rows.append(row)

    df = pd.DataFrame(rows)
    if not df.empty:
        df['_arch_order'] = df['Architecture'].apply(_arch_sort_val)
        df['_hf_pct']     = df['Model Key'].apply(lambda x: -extract_hf_pct(x))
        df = df.sort_values(['Condition', '_arch_order', '_hf_pct']).drop(
            columns=['_arch_order', '_hf_pct']).reset_index(drop=True)
    return df

# ----------------------------------------------------------------------
# Seed‑level test metrics from ALL trials_log.json under trials_dir
# ----------------------------------------------------------------------

def build_seed_metrics_from_trials_dir(trials_dir: str) -> pd.DataFrame:
    """
    Walk trials_dir recursively, find every trials_log.json,
    extract final_ entries, deduplicate by (Model Key, Seed),
    keep only official models, and return a DataFrame with
    Architecture, Model Identifier, Seed, Accuracy, F1, Precision,
    Recall, and subclass accuracies.
    """
    all_rows = []
    if not os.path.isdir(trials_dir):
        print(f"Error: trials_dir '{trials_dir}' not found.")
        return pd.DataFrame()

    for root, dirs, files in os.walk(trials_dir):
        if "trials_log.json" in files:
            path = os.path.join(root, "trials_log.json")
            try:
                with open(path, "r") as f:
                    data = json.load(f)
            except (json.JSONDecodeError, OSError):
                continue

            for key, entry in data.items():
                if not key.startswith("final_"):
                    continue
                mkey = entry.get("mkey")
                if not mkey:
                    continue
                if not is_official_model(mkey):
                    continue
                seed_results = entry.get("seed_results", [])
                for sr in seed_results:
                    seed = sr.get("seed")
                    test_metrics = sr.get("test_metrics")
                    if seed is None or test_metrics is None:
                        continue
                    row = {
                        "Model Key": mkey,
                        "Seed": seed,
                        "Accuracy": test_metrics.get("accuracy"),
                        "F1": test_metrics.get("f1"),
                        "Precision": test_metrics.get("precision"),
                        "Recall": test_metrics.get("recall"),
                    }
                    subclass_acc = test_metrics.get("subclass_acc", {})
                    for sc in SUBCLASSES:
                        row[f"{sc} Acc"] = subclass_acc.get(sc)
                    all_rows.append(row)

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    # Deduplicate: keep first occurrence of each (Model Key, Seed)
    df = df.drop_duplicates(subset=["Model Key", "Seed"], keep="first")

    # Add Architecture and Model Identifier (Display Label)
    df["Architecture"] = df["Model Key"].apply(extract_architecture)
    df["Model Identifier"] = df["Model Key"].apply(format_model_display_label)

    # Order columns
    ordered_cols = [
        "Architecture",
        "Model Identifier",
        "Seed",
        "Accuracy",
        "F1",
        "Precision",
        "Recall",
        "HR Acc",
        "AI-R Acc",
        "HF Acc",
        "AI-F Acc",
    ]
    df = df[ordered_cols]

    # Sort by architecture (BERT first), then model identifier, then seed
    df["_arch_sort"] = df["Architecture"].apply(lambda x: 0 if "BERT" in x else 1)
    df["_label_sort"] = df["Model Identifier"].apply(lambda x: x)
    df = df.sort_values(["_arch_sort", "_label_sort", "Seed"]).drop(
        columns=["_arch_sort", "_label_sort"]
    ).reset_index(drop=True)

    return df


# ----------------------------------------------------------------------
# Confusion pattern analysis (Tagalog BERT only)
# ----------------------------------------------------------------------

def compute_confusion_pattern_table(total_df, models_of_interest):
    rows = []
    for mkey in models_of_interest:
        grp = total_df[total_df['model_key'] == mkey]
        if grp.empty:
            continue
        row = {
            'Condition':    None,
            'Model Key':    mkey,
            'Architecture': extract_architecture(mkey),
            'HF Label':     extract_hf_label(mkey),
        }
        for sc in SUBCLASSES:
            sg = grp[grp['subclass'] == sc]
            if not sg.empty:
                real_pct = 100.0 * np.mean(sg['pred_int'] == LABEL_STR_TO_INT['real'])
                fake_pct = 100.0 - real_pct
            else:
                real_pct, fake_pct = 0.0, 0.0
            row[f'{sc} Real%'] = round(real_pct, 2)
            row[f'{sc} Fake%'] = round(fake_pct, 2)
        rows.append(row)
    return pd.DataFrame(rows)

def plot_confusion_patterns_one_condition(pattern_df, cond_key, cond_label, out_dir):
    n_models = len(pattern_df)
    if n_models == 0:
        return
    fig, axes = plt.subplots(1, n_models, figsize=(4 * n_models, 4.5), squeeze=False)
    axes = axes.flatten()
    subclass_labels = SUBCLASSES
    real_color = '#2ca02c'
    fake_color = '#d62728'

    for idx, (_, model_row) in enumerate(pattern_df.iterrows()):
        ax = axes[idx]
        hf_label = model_row['HF Label']
        model_name = f"{hf_label}"
        real_vals = [model_row[f'{sc} Real%'] for sc in subclass_labels]
        fake_vals = [model_row[f'{sc} Fake%'] for sc in subclass_labels]

        y_pos = np.arange(len(subclass_labels))
        ax.barh(y_pos, real_vals, color=real_color, label='Predicted Real')
        ax.barh(y_pos, fake_vals, left=real_vals, color=fake_color, label='Predicted Fake')

        ax.set_yticks(y_pos)
        ax.set_yticklabels(subclass_labels)
        ax.set_xlim(0, 100)
        ax.set_xlabel('%')
        ax.set_title(model_name, fontsize=10, fontweight='bold')
        ax.invert_yaxis()

    handles, labels = ax.get_legend_handles_labels()
    fig.legend(handles, labels, loc='lower center', ncol=2, framealpha=0.9, fontsize=10)
    plt.suptitle(f'Confusion Patterns: Tagalog BERT\n{cond_label}',
                 fontsize=13, fontweight='bold', y=1.03)
    plt.tight_layout(rect=[0, 0.1, 1, 0.95])
    fname = os.path.join(out_dir, f'ConfusionPatterns_{cond_key}_BERT.png')
    plt.savefig(fname, bbox_inches='tight', dpi=300)
    plt.close()
    print(f'  Saved: {fname}')

# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input_dir",  required=True,
                        help="Directory containing results folders with master_results.xlsx")
    parser.add_argument("--output_dir", required=True,
                        help="Where to save Excel and charts")
    parser.add_argument("--models_dir", required=True,
                        help="Base directory containing individual model folders")
    parser.add_argument("--trials_dir", required=True,
                        help="Directory containing all HPO trials_log.json files (the accounts folder)")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    main_excel = os.path.join(args.output_dir, "Thesis_Results_Final_Comprehensive.xlsx")

    print("Gathering master results...")
    predict_frames = []

    for root, dirs, files in os.walk(args.input_dir):
        if "master_results.xlsx" in files and os.path.basename(root) == "results":
            path = os.path.join(root, "master_results.xlsx")
            print(f"  Found predictions: {path}")
            predict_frames.append(pd.read_excel(path, sheet_name="Predictions", dtype=str))

    if not predict_frames:
        print("Error: No master_results.xlsx found.")
        return

    total_df = pd.concat(predict_frames, ignore_index=True)
    total_df['true_int'] = total_df['true_label'].map(LABEL_STR_TO_INT)
    total_df['pred_int'] = total_df['pred_label'].map(LABEL_STR_TO_INT)

    main_sheets = {}
    overall_dfs = {}

    for cond_key, hr_string in CONDITIONS.items():
        print(f"\nProcessing {cond_key} ({CONDITION_LABELS[cond_key]})...")
        cond_df = total_df[total_df['model_key'].str.contains(hr_string, case=False)]
        if cond_df.empty:
            continue

        overall_rows, subclass_rows = [], []
        for (m_key, arch_raw), grp in cond_df.groupby(["model_key", "arch"]):
            yt, yp = grp['true_int'].values, grp['pred_int'].values
            prec, rec, f1, _ = precision_recall_fscore_support(
                yt, yp, average='binary', zero_division=0)
            overall_rows.append({
                "Model Identifier": m_key,
                "Architecture":     extract_architecture(m_key),
                "Accuracy":         round(accuracy_score(yt, yp), 4),
                "Precision":        round(prec, 4),
                "Recall":           round(rec,  4),
                "F1-Score":         round(f1,   4),
            })
            sc_row = {"Model Identifier": m_key, "Architecture": extract_architecture(m_key)}
            for sc in SUBCLASSES:
                sg = grp[grp['subclass'] == sc]
                sc_row[f"{sc} Accuracy"] = (
                    round(accuracy_score(sg['true_int'], sg['pred_int']), 4)
                    if not sg.empty else 0.0)
            subclass_rows.append(sc_row)

        overall_df  = standardize_and_sort_thesis_data(pd.DataFrame(overall_rows))
        subclass_df = standardize_and_sort_thesis_data(pd.DataFrame(subclass_rows))
        within_df   = run_within_condition_mcnemar(cond_df)

        main_sheets[f"{cond_key}_Overall"]  = overall_df
        main_sheets[f"{cond_key}_Subclass"] = subclass_df
        main_sheets[f"{cond_key}_Stats"]    = within_df
        overall_dfs[cond_key]               = overall_df

        generate_overall_accuracy_bars(overall_df, cond_key, args.output_dir)

    print("\nGenerating ratio trend charts...")
    trend_summary = _build_trend_summary(total_df)

    print("  [Fig-4 style] Per-condition trend charts (both archs)...")
    generate_per_condition_trend_charts(trend_summary, args.output_dir)

    print("  [Comparison charts] One figure per architecture...")
    generate_per_arch_condition_comparison_charts(trend_summary, args.output_dir)

    print("\nRunning cross-condition McNemar...")
    main_sheets["CrossCondition_Stats"] = run_cross_condition_mcnemar(total_df)

    print("\nRunning champion McNemar...")
    main_sheets["Champion_Comparisons"] = run_champion_mcnemar(total_df, top_n=3)

    # ------------------------------------------------------------------
    # Confusion pattern analysis for Tagalog BERT only
    # ------------------------------------------------------------------
    print("\nComputing confusion-pattern analysis for Tagalog BERT (all ratios)...")
    bert_models = [mk for mk in total_df['model_key'].unique()
                   if extract_architecture(mk) == 'Tagalog BERT']

    if bert_models:
        pattern_df = compute_confusion_pattern_table(total_df, bert_models)
        pattern_df.columns = [c.replace('_', ' ') for c in pattern_df.columns]
        main_sheets["Confusion Patterns BERT"] = pattern_df

        for cond in CONDITIONS.keys():
            cond_df = pattern_df[pattern_df['Condition'] == cond]
            if cond_df.empty:
                continue
            cond_df = cond_df.copy()
            cond_df['_hf_pct'] = cond_df['HF Label'].apply(extract_hf_pct)
            cond_df = cond_df.sort_values('_hf_pct').drop(columns=['_hf_pct'])

            plot_confusion_patterns_one_condition(
                cond_df, cond, CONDITION_LABELS.get(cond, cond), args.output_dir
            )
    else:
        print("  No Tagalog BERT models found.")

    # ------------------------------------------------------------------
    # Learning curves from final_results.json
    # ------------------------------------------------------------------
    print("\nSearching for all final_results.json files...")
    json_paths = []
    for root, dirs, files in os.walk(args.input_dir):
        if "final_results.json" in files:
            json_paths.append(os.path.join(root, "final_results.json"))

    learning_curve_data_df = pd.DataFrame()
    if not json_paths:
        print("No final_results.json files found; skipping learning curves.")
    else:
        print(f"Found {len(json_paths)} file(s). Merging and plotting...")
        all_losses = merge_multiple_final_results(json_paths)
        if not all_losses:
            print("No training history extracted from any final_results.json.")
        else:
            print(f"Extracted learning curves for {len(all_losses)} models. Plotting...")
            generate_learning_curve_plots_from_dict(all_losses, args.output_dir)
            learning_curve_data_df = build_learning_curve_excel_data(all_losses)
            learning_curve_data_df.columns = [c.replace('_', ' ') for c in learning_curve_data_df.columns]
            main_sheets["Learning Curves Data"] = learning_curve_data_df

    # ------------------------------------------------------------------
    # Hyperparameters from training_meta.json
    # ------------------------------------------------------------------
    print("\nExtracting hyperparameters from training_meta.json files...")
    hyperparams_dict = load_hyperparams_from_models_dir(args.models_dir)
    hyperparameter_df = pd.DataFrame()
    if hyperparams_dict:
        hyperparameter_df = build_combined_hyperparameter_dataframe(hyperparams_dict)
        hyperparameter_df.columns = [c.replace('_', ' ') for c in hyperparameter_df.columns]
        main_sheets["Hyperparameters"] = hyperparameter_df
        print(f"Combined hyperparameter table built for {len(hyperparams_dict)} models.")
    else:
        print("No hyperparameters found. Check that --models_dir is correct and "
              "every model folder contains final/training_meta.json with an 'hparams' field.")

    # ------------------------------------------------------------------
    # Seed-level test metrics from ALL trials_log.json files (deduplicated)
    # ------------------------------------------------------------------
    print("\nExtracting seed-level test metrics from all trials_log.json files...")
    seed_metrics_df = build_seed_metrics_from_trials_dir(args.trials_dir)
    if not seed_metrics_df.empty:
        main_sheets["Seed Test Metrics"] = seed_metrics_df
        print(f"Extracted seed metrics for {seed_metrics_df['Model Identifier'].nunique()} models "
              f"({len(seed_metrics_df)} seed runs).")
    else:
        print("No seed metrics extracted. Check that the trials_dir contains trials_log.json files.")

    # ------------------------------------------------------------------
    # Write Excel output
    # ------------------------------------------------------------------
    print(f"\nWriting {main_excel}...")
    with pd.ExcelWriter(main_excel, engine='openpyxl') as writer:
        for sheet_name, df in main_sheets.items():
            if df.empty:
                continue
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_journal_table_styling(writer, sheet_name, df)

    print("\n" + "=" * 60)
    print("OUTPUT SUMMARY")
    print("=" * 60)
    print(f"\n[Excel]  {main_excel}")
    for sheet in main_sheets:
        print(f"         └─ {sheet}")
    print("\n[Charts] Ratio Trend (per condition):")
    for ck in CONDITIONS:
        print(f"         └─ {ck}_RatioTrend.png")
    print("\n[Charts] Architecture Comparison:")
    for arch in ["Tagalog_BERT", "Tagalog_DistilBERT"]:
        print(f"         └─ {arch}_ConditionComparison_RatioTrend.png")
    print("\n[Charts] Overall Accuracy Bars:")
    for ck in CONDITIONS:
        print(f"         └─ {ck}_Accuracy_Bars.png")
    if not learning_curve_data_df.empty:
        print("\n[Charts] Learning Curves (per condition):")
        for ck in CONDITIONS:
            fname = os.path.join(args.output_dir, f'{ck}_LearningCurves.png')
            if os.path.exists(fname):
                print(f"         └─ {ck}_LearningCurves.png")
    if not hyperparameter_df.empty:
        print("\n[Table] Hyperparameters sheet added (source: training_meta.json).")
    if 'Confusion Patterns BERT' in main_sheets:
        print("\n[Analysis] Confusion Patterns (Tagalog BERT)")
        print("         └─ Sheet: Confusion Patterns BERT")
        for ck in CONDITIONS:
            print(f"         └─ Chart: ConfusionPatterns_{ck}_BERT.png")
    print("\n[COMPLETE] All results, statistics, and graphics generated.")


if __name__ == "__main__":
    main()