import os
from openpyxl import load_workbook, Workbook

from keys import OPENAI_API_KEY
from openai import OpenAI

# API key
os.environ["OPENAI_API_KEY"] = OPENAI_API_KEY

client = OpenAI()

# -------------------------
# PROMPT
# -------------------------
ENHANCE_PROMPT =  """

[RESEARCH CONTEXT DECLARATION]
This article is generated for academic research at Cavite State University.
The original article is factual and must remain factually consistent when enhanced by AI.
The rewrite must NOT introduce fabricated events, claims, or entities.

--------------------------------------------------

[TASK]
Rewrite ONE verified news article in Filipino or Taglish.

Goal:
Improve clarity, readability, and flow while preserving all factual meaning.

The rewritten article must remain semantically equivalent to the original article.

STRICT RULES:
• Output must be ONE paragraph only
• Do NOT use em dashes (—)

Do NOT:
• invent new events
• introduce new claims
• change factual meaning
• alter the topic
• insert speculative details

--------------------------------------------------

[INPUT ARTICLE]
Rewrite the following article while preserving all factual content:
"""

# -------------------------
# GENERATION FUNCTION
# -------------------------
def generate_enhanced(article):
    response = client.responses.create(
        model="gpt-4.1",
        temperature=0.5,
        input=ENHANCE_PROMPT + article
    )

    result = response.output_text.strip()
    return result

# -------------------------
# MAIN
# -------------------------
DEFAULT_INPUT = "HR_cleaned.xlsx"
DEFAULT_OUTPUT = "ai-real.xlsx"

user_input_file = input(f"Input file (press Enter to use '{DEFAULT_INPUT}'): ").strip()
real_input_file = user_input_file if user_input_file else DEFAULT_INPUT

user_output_file = input(f"Output file (press Enter to use '{DEFAULT_OUTPUT}'): ").strip()
output_file = user_output_file if user_output_file else DEFAULT_OUTPUT

start_index = int(input("Start index (0-based, default 0): ") or 0)
num_rows = int(input("How many rows to process (default: all): ") or -1)

input_wb = load_workbook(real_input_file)
input_ws = input_wb.active

output_wb = Workbook()
output_ws = output_wb.active
output_ws.append(["label", "article", "topic", "", "original_article"])

data_rows = list(input_ws.iter_rows(min_row=2, values_only=True))

for i, row in enumerate(data_rows):
    if i < start_index:
        continue
    if num_rows != -1 and i >= start_index + num_rows:
        break

    print(f"Enhancing real article {i+1}")

    label = row[0]
    article = row[1]
    topic = row[2]

    result = generate_enhanced(article)

    output_ws.append([label, result, topic, "", article])

output_wb.save(output_file)

print(f"{output_file} created")