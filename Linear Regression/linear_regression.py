"""
length_classifier.py
--------------------
Length-only (token count) logistic regression baseline, adapted to read
directly from the Excel workbook produced by stratify.py.

Usage:
    python length_classifier.py                          # prompts for path
    python length_classifier.py stratified_dataset.xlsx

Sheet selection
    An interactive numbered menu is shown. Enter space-separated numbers
    (e.g. "1 3 5") or "all" to process every data sheet.

Output — length_classifier_results.xlsx
    Sheet "Test Results"   : train & test performance per sheet
    Sheet "Val Results"    : val performance per sheet  (kept completely separate)
    Sheet "Class Reports"  : full precision / recall / F1 per sheet × split
"""

import sys
import pathlib

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    classification_report,
    precision_recall_fscore_support,
)
from transformers import AutoTokenizer
import transformers


# ── 1. TOKENIZER (loaded once) ────────────────────────────────────────────────

transformers.logging.set_verbosity_error()

tokenizer = AutoTokenizer.from_pretrained(
    "jcblaise/bert-tagalog-base-cased",
    token=False,
)

def get_token_count(text: str) -> int:
    return len(tokenizer.encode(str(text), truncation=False))


# ── 2. WORKBOOK SELECTION ─────────────────────────────────────────────────────

def resolve_workbook() -> str:
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        path = input("Path to stratified Excel workbook: ").strip()
    resolved = pathlib.Path(path)
    if not resolved.exists():
        raise FileNotFoundError(f"File not found: {resolved.resolve()}")
    return str(resolved)


# ── 3. INTERACTIVE SHEET PICKER ───────────────────────────────────────────────

def pick_sheets(xl: pd.ExcelFile) -> list[str]:
    data_sheets = [s for s in xl.sheet_names if s.lower() != "summary"]

    print("\nAvailable sheets:")
    for i, name in enumerate(data_sheets, 1):
        print(f"  {i:>3}.  {name}")

    print("\nEnter sheet number(s) separated by spaces, or 'all': ", end="")
    raw = input().strip().lower()

    if raw == "all":
        return data_sheets

    chosen = []
    for token in raw.split():
        try:
            idx = int(token) - 1
            if not (0 <= idx < len(data_sheets)):
                raise ValueError
            chosen.append(data_sheets[idx])
        except ValueError:
            print(f"  [WARN] '{token}' is not a valid sheet number — skipped.")

    if not chosen:
        raise ValueError("No valid sheets selected.")

    return chosen


# ── 4. TOKEN-COUNT STATS HELPERS ─────────────────────────────────────────────

def collect_token_stats(series: pd.Series) -> dict:
    return {
        "n":      int(len(series)),
        "mean":   float(series.mean()),
        "median": float(series.median()),
        "std":    float(series.std()),
        "min":    int(series.min()),
        "max":    int(series.max()),
        "p25":    float(series.quantile(0.25)),
        "p75":    float(series.quantile(0.75)),
    }

def print_token_stats(label: str, series: pd.Series) -> None:
    print(f"    {label:<8}  "
          f"n={len(series):<6}  "
          f"mean={series.mean():.1f}  "
          f"median={series.median():.1f}  "
          f"std={series.std():.1f}  "
          f"min={series.min()}  "
          f"max={series.max()}")


# ── 5. PER-SHEET CLASSIFIER ───────────────────────────────────────────────────

def run_sheet(sheet_name: str, df: pd.DataFrame) -> dict:
    """
    Fits on train only.
    Evaluates train + test together (main pipeline).
    Evaluates val separately — stored but printed only after all sheets finish.
    """
    print(f"\n{'═' * 64}")
    print(f"  Sheet : {sheet_name}")
    print(f"{'═' * 64}")

    required = {"article", "label", "split"}
    missing  = required - set(df.columns.str.lower())
    if missing:
        raise ValueError(f"Sheet '{sheet_name}' is missing columns: {missing}")

    df.columns = df.columns.str.lower()

    train_df = df[df["split"] == "train"].copy().reset_index(drop=True)
    val_df   = df[df["split"] == "val"].copy().reset_index(drop=True)
    test_df  = df[df["split"] == "test"].copy().reset_index(drop=True)

    print(f"\n  Split sizes  →  train: {len(train_df)} | "
          f"val: {len(val_df)} | test: {len(test_df)}")

    # ── tokenize ──────────────────────────────────────────────────────────────
    print("\n  Counting tokens …")
    for label, part in [("train", train_df), ("val", val_df), ("test", test_df)]:
        part["token_count"] = part["article"].apply(get_token_count)

    print("\n  Token-count statistics:")
    token_stats: dict[str, dict] = {}
    for label, part in [("train", train_df), ("val", val_df), ("test", test_df)]:
        print_token_stats(label, part["token_count"])
        # overall stats for this split
        token_stats[label] = {"_overall": collect_token_stats(part["token_count"])}
        # per news_type breakdown (if available)
        if "news_type" in part.columns:
            for nt, grp in part.groupby("news_type"):
                token_stats[label][nt] = collect_token_stats(grp["token_count"])

    # ── fit (train only) ──────────────────────────────────────────────────────
    clf = LogisticRegression(random_state=42, max_iter=1000)
    clf.fit(train_df[["token_count"]], train_df["label"])
    print(f"\n  Model  →  coefficient: {clf.coef_[0][0]:.6f}  "
          f"intercept: {clf.intercept_[0]:.6f}")

    # ── helper: evaluate one split ────────────────────────────────────────────
    def evaluate(part: pd.DataFrame) -> dict:
        y_true = part["label"]
        y_pred = clf.predict(part[["token_count"]])
        acc    = accuracy_score(y_true, y_pred)
        prec, rec, f1, sup = precision_recall_fscore_support(
            y_true, y_pred, labels=[0, 1], zero_division=0
        )
        nt_rows = []
        if "news_type" in part.columns:
            for nt, grp in part.groupby("news_type"):
                nt_pred = clf.predict(grp[["token_count"]])
                nt_rows.append({
                    "news_type": nt,
                    "n":         len(grp),
                    "accuracy":  accuracy_score(grp["label"], nt_pred),
                })
        return {
            "n":          len(part),
            "accuracy":   acc,
            "fake_prec":  float(prec[0]), "fake_rec": float(rec[0]),
            "fake_f1":    float(f1[0]),   "fake_sup": int(sup[0]),
            "real_prec":  float(prec[1]), "real_rec": float(rec[1]),
            "real_f1":    float(f1[1]),   "real_sup": int(sup[1]),
            "report_str": classification_report(
                              y_true, y_pred,
                              target_names=["Fake (0)", "Real (1)"],
                              zero_division=0,
                          ),
            "nt_rows":    nt_rows,
        }

    # ── TRAIN & TEST printed now (main pipeline) ──────────────────────────────
    train_m = evaluate(train_df)
    test_m  = evaluate(test_df)
    val_m   = evaluate(val_df)   # computed now, printed later

    print("\n  ── TRAIN & TEST (main pipeline) ──────────────────────────────────")
    for label, m in [("TRAIN", train_m), ("TEST  ← main result", test_m)]:
        print(f"\n  [{label}]  accuracy: {m['accuracy']:.4f} ({m['accuracy']*100:.2f}%)")
        print(m["report_str"])
        if m["nt_rows"]:
            print("  Per news_type:")
            for row in m["nt_rows"]:
                print(f"    {row['news_type']:<8}  n={row['n']:<5}  acc={row['accuracy']:.4f}")

    return {
        "sheet":       sheet_name,
        "train":       train_m,
        "test":        test_m,
        "val":         val_m,
        "coef":        float(clf.coef_[0][0]),
        "intercept":   float(clf.intercept_[0]),
        "token_stats": token_stats,
    }


# ── 6. CONSOLE SUMMARIES ──────────────────────────────────────────────────────

def print_test_summary(records: list[dict]) -> None:
    print(f"\n\n{'═' * 72}")
    print("  TEST RESULTS SUMMARY")
    print(f"{'═' * 72}")
    print(f"{'Sheet':<32} {'Train n':>8} {'Train Acc':>10} {'Test n':>7} {'Test Acc':>10}")
    print("-" * 72)
    for r in records:
        print(
            f"{r['sheet']:<32} "
            f"{r['train']['n']:>8} "
            f"{r['train']['accuracy']:>10.4f} "
            f"{r['test']['n']:>7} "
            f"{r['test']['accuracy']:>10.4f}"
        )
    print("-" * 72)
    accs = [r["test"]["accuracy"] for r in records]
    print(f"\n  Test accuracy  →  mean: {sum(accs)/len(accs):.4f}  "
          f"max: {max(accs):.4f}  min: {min(accs):.4f}")
    mean_acc = sum(accs) / len(accs)
    print()
    if mean_acc < 0.60:
        print("  INTERPRETATION: Length is NOT meaningfully predictive.")
        print("  Cite Horne & Adali (2017) and Salvetti et al. (2016).")
    elif mean_acc < 0.70:
        print("  INTERPRETATION: Length has MODERATE predictive power.")
        print("  Acknowledge in limitations section.")
    else:
        print("  INTERPRETATION: Length is CONSISTENTLY predictive.")
        print("  Report transparently. Cite Horne & Adali (2017) and Salvetti et al. (2016).")


def print_val_summary(records: list[dict]) -> None:
    sep = "─" * 64
    print(f"\n\n{'═' * 64}")
    print("  VAL RESULTS  (diagnostic only — not part of the main pipeline)")
    print(f"{'═' * 64}")
    print("  Val is reported here solely as a split-integrity check.")
    print("  It was never seen by the model during fitting or selection.")
    print(sep)
    for r in records:
        v = r["val"]
        print(f"\n  Sheet : {r['sheet']}")
        print(f"  Val accuracy : {v['accuracy']:.4f} ({v['accuracy']*100:.2f}%)")
        print(v["report_str"])
        if v["nt_rows"]:
            print("  Per news_type:")
            for row in v["nt_rows"]:
                print(f"    {row['news_type']:<8}  n={row['n']:<5}  acc={row['accuracy']:.4f}")
    print(sep)
    accs = [r["val"]["accuracy"] for r in records]
    print(f"\n  Val accuracy  →  mean: {sum(accs)/len(accs):.4f}  "
          f"max: {max(accs):.4f}  min: {min(accs):.4f}")


# ── 7. EXCEL OUTPUT ───────────────────────────────────────────────────────────

DARK_BLUE   = "1F4E79"
ALT_ROW     = "F2F7FB"
GREEN_FILL  = "E2EFDA"
RED_FILL    = "FCE4D6"
YELLOW_FILL = "FFF2CC"
BLUE_FILL   = "BDD7EE"
GREY_FILL   = "F2F2F2"
SECTION_FILL= "D6E4F0"

def _hdr(ws, row: int, col: int, value) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", start_color=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _cell(ws, row: int, col: int, value, bold=False,
          fill_hex=None, align="center", fmt=None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fill_hex:
        c.fill = PatternFill("solid", start_color=fill_hex)
    if fmt:
        c.number_format = fmt

def _pct(ws, row: int, col: int, value, fill_hex=None) -> None:
    _cell(ws, row, col, value, fill_hex=fill_hex, fmt="0.00%")

def _set_widths(ws, widths: list[float]) -> None:
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ── Sheet: Test Results ───────────────────────────────────────────────────────

def write_test_sheet(ws, records: list[dict]) -> None:
    ws.title = "Test Results"
    headers  = [
        "Sheet", "Train n", "Train Acc",
        "Test n", "Test Acc",
        "Fake Prec", "Fake Rec", "Fake F1",
        "Real Prec", "Real Rec", "Real F1",
        "Coefficient", "Intercept",
    ]
    _set_widths(ws, [32, 10, 11, 10, 11, 11, 11, 11, 11, 11, 11, 13, 13])
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 28

    for ri, r in enumerate(records, 2):
        fill = ALT_ROW if ri % 2 == 0 else None
        tr, te = r["train"], r["test"]
        _cell(ws, ri, 1,  r["sheet"],         bold=True, fill_hex=fill, align="left")
        _cell(ws, ri, 2,  tr["n"],             fill_hex=fill)
        _pct( ws, ri, 3,  tr["accuracy"],      fill_hex=GREEN_FILL)
        _cell(ws, ri, 4,  te["n"],             fill_hex=fill)
        _pct( ws, ri, 5,  te["accuracy"],      fill_hex=GREEN_FILL)
        _pct( ws, ri, 6,  te["fake_prec"],     fill_hex=fill)
        _pct( ws, ri, 7,  te["fake_rec"],      fill_hex=fill)
        _pct( ws, ri, 8,  te["fake_f1"],       fill_hex=fill)
        _pct( ws, ri, 9,  te["real_prec"],     fill_hex=fill)
        _pct( ws, ri, 10, te["real_rec"],      fill_hex=fill)
        _pct( ws, ri, 11, te["real_f1"],       fill_hex=fill)
        _cell(ws, ri, 12, round(r["coef"], 6), fill_hex=fill, fmt="0.000000")
        _cell(ws, ri, 13, round(r["intercept"], 6), fill_hex=fill, fmt="0.000000")

    # Interpretation block
    accs     = [r["test"]["accuracy"] for r in records]
    mean_acc = sum(accs) / len(accs)
    last     = len(records) + 3

    interp = [
        f"Mean test accuracy: {mean_acc:.4f} ({mean_acc*100:.2f}%)   "
        f"Max: {max(accs):.4f}   Min: {min(accs):.4f}",
    ]
    if mean_acc < 0.60:
        interp += ["INTERPRETATION: Length is NOT meaningfully predictive.",
                   "Cite Horne & Adali (2017) and Salvetti et al. (2016)."]
    elif mean_acc < 0.70:
        interp += ["INTERPRETATION: Length has MODERATE predictive power.",
                   "Acknowledge in limitations section."]
    else:
        interp += ["INTERPRETATION: Length is CONSISTENTLY predictive.",
                   "Report transparently. Cite Horne & Adali (2017) and Salvetti et al. (2016)."]

    for i, line in enumerate(interp):
        c = ws.cell(row=last + i, column=1, value=line)
        c.font = Font(name="Arial", bold=(i > 0), size=10,
                      color="1F4E79" if i > 0 else "000000")

    ws.freeze_panes = "A2"


# ── Sheet: Val Results ────────────────────────────────────────────────────────

def write_val_sheet(ws, records: list[dict]) -> None:
    ws.title = "Val Results"

    # Disclaimer banner — row 1
    disc = ws.cell(row=1, column=1,
                   value="DIAGNOSTIC ONLY — Val set was never used to fit or tune the model. "
                         "Compare Val Acc vs Test Acc to check split integrity. "
                         "Do not treat this as a performance result.")
    disc.font      = Font(name="Arial", bold=True, size=10, color="9C0006")
    disc.fill      = PatternFill("solid", start_color="FFC7CE")
    disc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.row_dimensions[1].height = 40

    headers = [
        "Sheet", "Val n", "Val Acc",
        "Fake Prec", "Fake Rec", "Fake F1",
        "Real Prec", "Real Rec", "Real F1",
    ]
    _set_widths(ws, [32, 10, 11, 11, 11, 11, 11, 11, 11])
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 2, ci, h)
    ws.row_dimensions[2].height = 28

    for ri, r in enumerate(records, 3):
        fill = ALT_ROW if ri % 2 == 0 else None
        v = r["val"]
        _cell(ws, ri, 1, r["sheet"],    bold=True, fill_hex=fill, align="left")
        _cell(ws, ri, 2, v["n"],        fill_hex=fill)
        _pct( ws, ri, 3, v["accuracy"], fill_hex=YELLOW_FILL)
        _pct( ws, ri, 4, v["fake_prec"],fill_hex=fill)
        _pct( ws, ri, 5, v["fake_rec"], fill_hex=fill)
        _pct( ws, ri, 6, v["fake_f1"],  fill_hex=fill)
        _pct( ws, ri, 7, v["real_prec"],fill_hex=fill)
        _pct( ws, ri, 8, v["real_rec"], fill_hex=fill)
        _pct( ws, ri, 9, v["real_f1"],  fill_hex=fill)

    ws.freeze_panes = "A3"


# ── Sheet: Classification Reports ────────────────────────────────────────────

def write_reports_sheet(ws, records: list[dict]) -> None:
    ws.title = "Class Reports"
    headers  = ["Sheet", "Split", "Class", "Precision", "Recall", "F1", "Support"]
    _set_widths(ws, [32, 10, 14, 12, 12, 12, 12])
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 28

    SPLIT_FILLS = {"train": GREEN_FILL, "test": BLUE_FILL, "val": YELLOW_FILL}
    CLASS_FILLS = {"Fake (0)": RED_FILL, "Real (1)": GREEN_FILL, "macro avg": GREY_FILL}

    ri = 2
    for r in records:
        for split_label, m in [("train", r["train"]),
                                ("test",  r["test"]),
                                ("val",   r["val"])]:
            sp_fill = SPLIT_FILLS.get(split_label)
            classes = [
                ("Fake (0)", m["fake_prec"], m["fake_rec"], m["fake_f1"], m["fake_sup"]),
                ("Real (1)", m["real_prec"], m["real_rec"], m["real_f1"], m["real_sup"]),
                ("macro avg",
                 (m["fake_prec"] + m["real_prec"]) / 2,
                 (m["fake_rec"]  + m["real_rec"])  / 2,
                 (m["fake_f1"]   + m["real_f1"])   / 2,
                 m["fake_sup"] + m["real_sup"]),
            ]
            for i, (cls, prec, rec, f1, sup) in enumerate(classes):
                cls_fill = CLASS_FILLS.get(cls)
                _cell(ws, ri, 1, r["sheet"] if i == 0 else "",
                      bold=(i == 0), align="left")
                _cell(ws, ri, 2, split_label if i == 0 else "",
                      fill_hex=sp_fill if i == 0 else None)
                _cell(ws, ri, 3, cls,   fill_hex=cls_fill)
                _pct( ws, ri, 4, prec,  fill_hex=cls_fill)
                _pct( ws, ri, 5, rec,   fill_hex=cls_fill)
                _pct( ws, ri, 6, f1,    fill_hex=cls_fill)
                _cell(ws, ri, 7, sup,   fill_hex=cls_fill)
                ri += 1

    ws.freeze_panes = "A2"


# ── Sheet: Token Stats ────────────────────────────────────────────────────────

def write_token_sheet(ws, records: list[dict]) -> None:
    ws.title = "Token Stats"

    # Columns: Sheet | Split | Group | n | Mean | Median | Std | Min | Max | P25 | P75
    headers = [
        "Sheet", "Split", "News Type",
        "n", "Mean", "Median", "Std Dev", "Min", "Max", "P25", "P75",
    ]
    _set_widths(ws, [32, 10, 14, 8, 10, 10, 10, 8, 8, 10, 10])
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 28

    SPLIT_FILLS = {"train": GREEN_FILL, "test": BLUE_FILL, "val": YELLOW_FILL}
    NT_COLORS   = {
        "HR":   "C6EFCE",
        "AI-R": "FFEB9C",
        "HF":   "FFC7CE",
        "AI-F": "E2AFFF",
        "_overall": SECTION_FILL,
    }
    NUM_FMT = "0.00"

    ri = 2
    for r in records:
        ts = r.get("token_stats", {})
        for split_label in ("train", "test", "val"):
            split_data = ts.get(split_label, {})
            sp_fill    = SPLIT_FILLS.get(split_label)

            # Sort so _overall comes first, then news types alphabetically
            groups = ["_overall"] + sorted(k for k in split_data if k != "_overall")
            first_in_split = True

            for group in groups:
                stats    = split_data.get(group)
                if stats is None:
                    continue
                nt_fill  = NT_COLORS.get(group, ALT_ROW)
                disp_grp = "Overall" if group == "_overall" else group

                _cell(ws, ri, 1, r["sheet"] if first_in_split else "",
                      bold=first_in_split, align="left")
                _cell(ws, ri, 2, split_label if first_in_split else "",
                      fill_hex=sp_fill if first_in_split else None)
                _cell(ws, ri, 3, disp_grp,
                      bold=(group == "_overall"), fill_hex=nt_fill)
                _cell(ws, ri, 4,  stats["n"],      fill_hex=nt_fill)
                _cell(ws, ri, 5,  stats["mean"],   fill_hex=nt_fill, fmt=NUM_FMT)
                _cell(ws, ri, 6,  stats["median"], fill_hex=nt_fill, fmt=NUM_FMT)
                _cell(ws, ri, 7,  stats["std"],    fill_hex=nt_fill, fmt=NUM_FMT)
                _cell(ws, ri, 8,  stats["min"],    fill_hex=nt_fill)
                _cell(ws, ri, 9,  stats["max"],    fill_hex=nt_fill)
                _cell(ws, ri, 10, stats["p25"],    fill_hex=nt_fill, fmt=NUM_FMT)
                _cell(ws, ri, 11, stats["p75"],    fill_hex=nt_fill, fmt=NUM_FMT)
                ri += 1
                first_in_split = False

    ws.freeze_panes = "A2"


def save_excel(records: list[dict], out_path: pathlib.Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    write_test_sheet(wb.create_sheet("Test Results"), records)
    write_val_sheet(wb.create_sheet("Val Results"), records)
    write_reports_sheet(wb.create_sheet("Class Reports"), records)
    write_token_sheet(wb.create_sheet("Token Stats"), records)
    wb.save(str(out_path))
    print(f"\n  Results saved → {out_path}")
    print("    • 'Test Results'  — train & test accuracy + per-class metrics")
    print("    • 'Val Results'   — val performance (separate / diagnostic only)")
    print("    • 'Class Reports' — precision / recall / F1 for all splits")
    print("    • 'Token Stats'   — token count distributions per split & news type")


# ── 8. MAIN ───────────────────────────────────────────────────────────────────

def main() -> None:
    wb_path  = resolve_workbook()
    xl       = pd.ExcelFile(wb_path)
    selected = pick_sheets(xl)

    print(f"\n  Processing {len(selected)} sheet(s): {selected}")

    all_results = []
    for sheet_name in selected:
        df     = xl.parse(sheet_name)
        result = run_sheet(sheet_name, df)
        all_results.append(result)

    # Test summary first — val summary completely separate, at the very end
    print_test_summary(all_results)
    print_val_summary(all_results)

    out_path = pathlib.Path("length_classifier_results.xlsx")
    save_excel(all_results, out_path)


if __name__ == "__main__":
    main()