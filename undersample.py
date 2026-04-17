"""
undersample.py
--------------
Produces a single Excel workbook with one sheet containing balanced HR and HF data.

Undersampling logic:
  For each topic, find the minimum row count between HR and HF sheets.
  Sample exactly that many rows from EACH sheet for that topic.
  This ensures:
    - Topics are balanced within each news type (HR and HF have equal counts per topic)
    - HR and HF are balanced against each other

Split logic (applied after undersampling):
  - test  : 15% of total_samples, 50:50 HR:HF — FIXED
  - val   : 15% of total_samples, 50:50 HR:HF — FIXED
  - train : 70% of total_samples, 50:50 HR:HF

Usage:
    python undersample.py config_undersample.json
"""

import json
import sys
import math
import pathlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ── config loading ────────────────────────────────────────────────────────────

def load_config(path: str) -> dict:
    with open(path) as f:
        cfg = json.load(f)
    if not cfg.get("input_file"):
        raise ValueError("'input_file' is required in the config.")
    return cfg


# ── data loading ──────────────────────────────────────────────────────────────

def load_sheets(input_file: str, sheet_map: dict) -> dict[str, pd.DataFrame]:
    """Return {news_type: DataFrame} with normalised columns and a stable _idx."""
    xl = pd.read_excel(input_file, sheet_name=None)
    frames = {}
    for news_type, sheet_name in sheet_map.items():
        if sheet_name not in xl:
            raise KeyError(
                f"Sheet '{sheet_name}' (news type '{news_type}') not found. "
                f"Available: {list(xl.keys())}"
            )
        df = xl[sheet_name].copy()
        df.columns = df.columns.str.lower()

        missing = {"label", "article", "topic"} - set(df.columns)
        if missing:
            raise ValueError(f"Sheet '{sheet_name}' missing columns: {missing}")

        df["topic"]     = df["topic"].str.strip().str.lower()
        df["news_type"] = news_type
        df["_idx"]      = range(len(df))
        frames[news_type] = df.reset_index(drop=True)

    return frames


# ── undersampling ─────────────────────────────────────────────────────────────

def compute_topic_caps(frames: dict[str, pd.DataFrame]) -> dict[str, int]:
    """
    For each topic, find the minimum row count across all sheets.
    This becomes the cap — how many rows we draw from EACH sheet.
    """
    topic_counts: dict[str, dict[str, int]] = {}

    for news_type, df in frames.items():
        for topic, group in df.groupby("topic"):
            topic_counts.setdefault(topic, {})[news_type] = len(group)

    caps = {}
    for topic, counts in topic_counts.items():
        min_count = min(counts.values())
        caps[topic] = min_count

        # informational breakdown
        count_str = "  ".join(f"{nt}={n}" for nt, n in sorted(counts.items()))
        print(f"  topic '{topic}': {count_str}  → cap={min_count}")

    return caps


def undersample_sheet(
    df: pd.DataFrame,
    caps: dict[str, int],
    rng: np.random.Generator,
    exclude_idx: set | None = None,
    label: str = "",
) -> pd.DataFrame:
    """
    Sample up to caps[topic] rows per topic from df.
    Rows in exclude_idx are not eligible.
    """
    pool = df if not exclude_idx else df[~df["_idx"].isin(exclude_idx)]
    parts = []

    for topic, cap in caps.items():
        if cap == 0:
            continue
        topic_pool = pool[pool["topic"] == topic]
        if topic_pool.empty:
            print(f"  [WARN]{label} topic '{topic}' not found in pool — skipped.")
            continue
        n_draw = min(cap, len(topic_pool))
        if n_draw < cap:
            print(
                f"  [WARN]{label} topic '{topic}': "
                f"wanted {cap}, only {n_draw} available after exclusions."
            )
        parts.append(
            topic_pool.sample(n=n_draw, random_state=int(rng.integers(1 << 31)))
        )

    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=df.columns)


# ── Excel writing ─────────────────────────────────────────────────────────────

COL_ORDER  = ["label", "article", "topic", "news_type", "split"]
COL_WIDTHS = {"label": 8, "article": 80, "topic": 28, "news_type": 12, "split": 10}

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT   = Font(name="Arial", size=10)

SPLIT_COLORS = {
    "train": "E2EFDA",   # soft green
    "val":   "FFF2CC",   # soft yellow
    "test":  "FCE4D6",   # soft orange
}


def write_sheet(ws, df: pd.DataFrame) -> None:
    cols = [c for c in COL_ORDER if c in df.columns]
    df   = df[cols]

    # header row
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # data rows
    for ri, row in enumerate(df.itertuples(index=False), 2):
        split_val = getattr(row, "split", "train")
        row_fill  = PatternFill("solid", start_color=SPLIT_COLORS.get(split_val, "FFFFFF"))
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = DATA_FONT
            cell.fill      = row_fill
            cell.alignment = Alignment(
                wrap_text=(cols[ci - 1] == "article"), vertical="top"
            )

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = \
            COL_WIDTHS.get(col, 15)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── split allocation ──────────────────────────────────────────────────────────

def split_counts(n: int) -> tuple[int, int, int]:
    """Return (n_train, n_val, n_test) from total n using 70/15/15."""
    n_test  = round(n * 0.15)
    n_val   = round(n * 0.15)
    n_train = n - n_test - n_val
    return n_train, n_val, n_test


# ── main ──────────────────────────────────────────────────────────────────────

def main(config_path: str) -> None:
    cfg         = load_config(config_path)
    seed        = cfg.get("seed", 42)
    input_file  = cfg["input_file"]
    output_file = cfg.get("output_file", "undersampled_dataset.xlsx")
    sheet_map   = cfg.get("sheet_names", {"HR": "HR", "HF": "HF"})

    print(f"\n📂  Input : {input_file}")
    print(f"📄  Output: {output_file}")
    print(f"🌱  Seed  : {seed}\n")

    frames = load_sheets(input_file, sheet_map)

    # ── compute per-topic caps via undersampling ───────────────────────────────
    print("Computing per-topic row caps (min across HR and HF) …")
    caps = compute_topic_caps(frames)
    total_per_sheet = sum(caps.values())
    total_rows      = total_per_sheet * len(frames)   # same count from each sheet
    n_train, n_val, n_test = split_counts(total_rows)

    print(f"\n  Rows per sheet after undersampling : {total_per_sheet}")
    print(f"  Sheets                             : {list(frames.keys())}")
    print(f"  Total rows                         : {total_rows}")
    print(f"  Split  → train={n_train}  val={n_val}  test={n_test}\n")

    # Per-sheet split sizes (equal across news types)
    tr_per, vl_per, te_per = split_counts(total_per_sheet)

    # ── fixed test + val (sampled once, reused) ───────────────────────────────
    print("Sampling fixed test / val sets …")
    rng_fixed = np.random.default_rng(seed)
    used: dict[str, set] = {nt: set() for nt in frames}

    # For test / val we use the same per-topic caps scaled to the split size.
    # We derive per-topic counts proportional to the cap weights.
    cap_total = sum(caps.values())
    def scale_caps(caps: dict, n: int) -> dict[str, int]:
        """Scale topic caps to sum to n using largest-remainder."""
        raw    = {t: c / cap_total * n for t, c in caps.items()}
        floors = {t: int(v) for t, v in raw.items()}
        deficit = n - sum(floors.values())
        for t in sorted(raw, key=lambda t: raw[t] - floors[t], reverse=True)[:deficit]:
            floors[t] += 1
        return floors

    test_caps  = scale_caps(caps, te_per)
    val_caps   = scale_caps(caps, vl_per)

    test_parts, val_parts = [], []

    for news_type in frames:
        # test
        part = undersample_sheet(
            frames[news_type], test_caps, rng_fixed,
            exclude_idx=used[news_type], label=f" [{news_type}/test]"
        )
        part["split"] = "test"
        used[news_type].update(part["_idx"].tolist())
        test_parts.append(part)

        # val
        part = undersample_sheet(
            frames[news_type], val_caps, rng_fixed,
            exclude_idx=used[news_type], label=f" [{news_type}/val]"
        )
        part["split"] = "val"
        used[news_type].update(part["_idx"].tolist())
        val_parts.append(part)

    test_df = pd.concat(test_parts, ignore_index=True)
    val_df  = pd.concat(val_parts,  ignore_index=True)
    print(f"  test rows : {len(test_df)}")
    print(f"  val rows  : {len(val_df)}\n")

    # ── training set ──────────────────────────────────────────────────────────
    print("Sampling training set …")
    train_caps  = scale_caps(caps, tr_per)
    rng_train   = np.random.default_rng(seed + 999)
    train_parts = []

    for news_type in frames:
        part = undersample_sheet(
            frames[news_type], train_caps, rng_train,
            exclude_idx=used[news_type], label=f" [{news_type}/train]"
        )
        part["split"] = "train"
        train_parts.append(part)

    train_df = pd.concat(train_parts, ignore_index=True)
    train_df = train_df.sample(frac=1, random_state=seed).reset_index(drop=True)
    print(f"  train rows: {len(train_df)}\n")

    # ── combine + write ───────────────────────────────────────────────────────
    combined = pd.concat([train_df, val_df, test_df], ignore_index=True)
    combined.drop(columns=["_idx"], inplace=True)

    print("Split summary:")
    print(combined["split"].value_counts().to_string())
    print("\nNews-type distribution:")
    print(combined["news_type"].value_counts().to_string())
    print("\nTopic distribution (train only):")
    print(
        train_df.drop(columns=["_idx"], errors="ignore")["topic"]
        .value_counts()
        .to_string()
    )

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="HR-HF-Undersampled")
    write_sheet(ws, combined)

    out_path = pathlib.Path(output_file)
    wb.save(str(out_path))
    print(f"\n✅  Saved → {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python undersample.py config_undersample.json")
        sys.exit(1)
    main(sys.argv[1])
