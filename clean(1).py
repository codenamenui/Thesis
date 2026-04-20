"""
clean.py
--------
Cleans the 'article' column of an Excel or CSV file containing Filipino news data.

Cleaning pipeline (applied in order):
  1. HTML entity decoding
  2. Unicode space normalization
  3. Collapse consecutive spaces / tabs
  4. Collapse 3+ consecutive newlines into 2
  5. Strip leading / trailing whitespace per line
  6. Source attribution removal
       "Source: 1 | 2 | 3"  /  "Source: CNN Philippines"
       "Source 1 Source 2 Source 3"  (no-colon variant)
  7. Watch video prompts
       "Watch [Full] video [here / on this link / :]"
       "You may watch the video on this link"
  8. PANOORIN prompts
       "Panoorin ang [video/ulat/buong/...] X."  — end-of-string only
       "PANOORIN: [ALL CAPS HEADLINE]"           — all-caps section headers
  9. BASAHIN / NARITO prompts
       "Basahin ang [reaksyon/komento/tugon/opinyon/...] ..."
       "BASAHIN ANG BUONG POST NI X:"            — all-caps variant (ends with colon)
       "Narito ang buong X post [ng netizen]:"   — end-of-string only
 10. Social / click prompts
       "comments and share this video"
       "I-click ang larawan sa ibaba ..."
 11. Final whitespace cleanup

Usage:
    python clean.py <input.xlsx|csv> <output.xlsx>
    python clean.py <input.xlsx|csv>               # writes <input>_cleaned.xlsx
"""

import re
import sys
import html
import pathlib
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── regex rules ───────────────────────────────────────────────────────────────

# 6a. "Source: 1 | 2" / "Source: CNN Philippines" (colon variant — always at end)
_RE_SOURCE_COLON = re.compile(
    r"\s*Sources?\s*:.*$",
    re.IGNORECASE | re.DOTALL,
)

# 6b. "Source 1 Source 2 Source 3" (no-colon numbered variant — always at end)
_RE_SOURCE_NUMBERED = re.compile(
    r"(?:\s+Source\s+\d+){2,}\s*$",
    re.IGNORECASE,
)

# 7. "Watch [Full] video here / on this link / :" / "You may watch the video on this link"
_RE_WATCH = re.compile(
    r"\s*(?:You\s+may\s+)?[Ww]atch\s+(?:Full\s+)?(?:the\s+)?video"
    r"(?:\s+(?:here|on\s+this\s+link)|\s*[:.])?\.?\s*",
    re.IGNORECASE,
)

# 8a. "Panoorin [niyo po] ang ..." — only at end of string.
#     End-anchor is the safety net: mid-article Panoorin sentences always have
#     real content after them and are therefore never at end-of-string.
#     Leading \s+ (not \s*) allows the preceding sentence-period to stay.
#     Trailing punctuation is optional to handle malformed/missing periods.
_RE_PANOORIN_END = re.compile(
    r"\s+Panoorin\s+(?:niyo\s+po\s+)?ang\b[^.!?]{0,250}[.!?]?\s*$",
    re.IGNORECASE,
)

# 8b. "PANOORIN: ALL CAPS HEADLINE" — section header format (all-caps after colon)
#     Only fires when text after colon is all-uppercase (e.g. "PANOORIN: 2 BABAE NATAGPUAN PATAY")
#     Does NOT fire for "PANOORIN: Kung dati..." (mixed case = real content).
_RE_PANOORIN_CAPS = re.compile(
    r"\s*PANOORIN\s*:\s*[A-Z0-9][A-Z0-9 ,.\'\-!\/\"#%&\(\)]+",
    re.MULTILINE,
)

# 9a. "Basahin ang [reaksyon/komento/tugon/opinyon/...] ..." (lowercase, ends with period)
#     Limit raised to 200 chars to catch longer sentences (previous 80-char limit caused misses).
_RE_BASAHIN = re.compile(
    r"Basahin\s+ang\s+"
    r"(?:(?:ilang|mga|naging|buong|kanilang|ilan|ilan\s+sa\s+mga)\s+)*"
    r"(?:sa\s+mga\s+)?"
    r"(?:mga\s+)?"
    r"(?:reaksyon|komento|tugon|opinyon|buong\s+kwento|kanilang|mga)"
    r"[^.]{0,200}\.",
    re.IGNORECASE,
)

# 9b. "BASAHIN ANG [BUONG/NAUNANG/HULING/SUSUNOD] POST NI X:" — all-caps, end-of-string
_RE_BASAHIN_CAPS = re.compile(
    r"\s*BASAHIN\s+ANG\s+(?:BUONG|NAUNANG|HULING|SUSUNOD|ORIHINAL)\s+POST\b[^:]*:?\s*$",
    re.IGNORECASE,
)

# 9c. "Narito ang buong X post [ng netizen]:" — only at end of string
#     Row 92 has content after this phrase so the end-anchor keeps that safe.
_RE_NARITO_END = re.compile(
    r"\s*Narito\s+ang\s+buong\b[^:]{0,80}:\s*$",
    re.IGNORECASE,
)

# 10a. "comments and share this video"
_RE_SHARE = re.compile(
    r"\s*comments\s+and\s+share\s+this\s+video\s*\.?\s*",
    re.IGNORECASE,
)

# 10b. "I-click ang larawan sa ibaba upang makita ..."
_RE_ICLICK = re.compile(
    r"\s*I-click\s+ang\s+larawan\s+sa\s+ibaba\b[^.]*\.?\s*",
    re.IGNORECASE,
)

# Whitespace helpers
_RE_MULTI_SP = re.compile(r"[ \t]+")
_RE_MULTI_NL = re.compile(r"\n{3,}")


# ── cleaning pipeline ─────────────────────────────────────────────────────────

def clean_article(text: str) -> str:
    if not isinstance(text, str):
        return text

    # 1. HTML entity decoding
    text = html.unescape(text)

    # 2. Unicode space normalization
    text = (text
            .replace("\u00a0", " ")
            .replace("\u200b", "")
            .replace("\u2009", " ")
            .replace("\u202f", " "))

    # 3. Collapse consecutive spaces / tabs
    text = _RE_MULTI_SP.sub(" ", text)

    # 4. Collapse 3+ newlines → 2
    text = _RE_MULTI_NL.sub("\n\n", text)

    # 5. Strip trailing whitespace per line
    text = "\n".join(line.strip() for line in text.splitlines())

    # 6. Source attribution
    text = _RE_SOURCE_COLON.sub("", text)
    text = _RE_SOURCE_NUMBERED.sub("", text)

    # 7. Watch video prompts
    text = _RE_WATCH.sub(" ", text)

    # 8. Basahin / Narito prompts  (run BEFORE Panoorin_end so that any Panoorin
    #    sentence newly exposed at end after Basahin removal gets caught below)
    #    BASAHIN_CAPS loops: removing one all-caps BASAHIN can expose another.
    text = _RE_BASAHIN.sub("", text)
    prev = None
    while prev != text:
        prev = text
        text = _RE_BASAHIN_CAPS.sub("", text)
    text = _RE_NARITO_END.sub("", text)

    # 9. Panoorin prompts
    # Loop PANOORIN_END: multiple consecutive Panoorin sentences at end require
    # repeated passes (each pass strips one from the tail).
    prev = None
    while prev != text:
        prev = text
        text = _RE_PANOORIN_END.sub("", text)
    text = _RE_PANOORIN_CAPS.sub("", text)

    # 10. Social / click prompts
    text = _RE_SHARE.sub("", text)
    text = _RE_ICLICK.sub("", text)

    # 11. Final cleanup: collapse any double-spaces left by removals, strip
    text = _RE_MULTI_SP.sub(" ", text).strip()

    return text


# ── stats ─────────────────────────────────────────────────────────────────────

def diff_stats(original: pd.Series, cleaned: pd.Series) -> dict:
    changed    = (original != cleaned).sum()
    total      = len(original)
    orig_chars = original.str.len().sum()
    cln_chars  = cleaned.str.len().sum()
    return {
        "total_rows":        total,
        "rows_changed":      changed,
        "pct_changed":       changed / total if total else 0,
        "chars_before":      orig_chars,
        "chars_after":       cln_chars,
        "chars_removed":     orig_chars - cln_chars,
        "pct_chars_removed": (orig_chars - cln_chars) / orig_chars if orig_chars else 0,
    }


# ── Excel writing ─────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(name="Arial", size=10)
SPLIT_COLORS = {"train": "E2EFDA", "val": "FFF2CC", "test": "FCE4D6"}
COL_ORDER    = ["label", "article", "topic", "news_type", "split"]
COL_WIDTHS   = {"label": 8, "article": 80, "topic": 28, "news_type": 12, "split": 10}


def write_sheet(ws, df: pd.DataFrame) -> None:
    cols = [c for c in COL_ORDER if c in df.columns]
    cols += [c for c in df.columns if c not in cols]
    df = df[cols]

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    has_split = "split" in cols
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = (PatternFill("solid",
                    start_color=SPLIT_COLORS.get(getattr(row, "split", ""), "FFFFFF"))
                if has_split
                else PatternFill("solid",
                    start_color="EEF3FB" if ri % 2 == 0 else "FFFFFF"))
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.alignment = Alignment(
                wrap_text=(cols[ci - 1] == "article"), vertical="top"
            )

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── main ──────────────────────────────────────────────────────────────────────

def main(input_path: str, output_path: str) -> None:
    print(f"\n📂  Input : {input_path}")
    print(f"📄  Output: {output_path}\n")

    ext = pathlib.Path(input_path).suffix.lower()
    raw_sheets = ({"Sheet1": pd.read_csv(input_path)} if ext == ".csv"
                  else pd.read_excel(input_path, sheet_name=None))

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in raw_sheets.items():
        df = df.copy()
        df.columns = df.columns.str.lower()

        if "article" not in df.columns:
            print(f"['{sheet_name}']  no 'article' column — copied unchanged.")
        else:
            original      = df["article"].astype(str)
            df["article"] = original.apply(clean_article)
            s = diff_stats(original, df["article"])
            print(f"['{sheet_name}']  {s['total_rows']} rows")
            print(f"  Rows changed      : {s['rows_changed']}  ({s['pct_changed']:.1%})")
            print(f"  Characters removed: {s['chars_removed']:,}  ({s['pct_chars_removed']:.1%})\n")

        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, df)

    wb.save(output_path)
    print(f"✅  Saved → {output_path}")


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print("Usage: python clean.py <input.xlsx|csv> [output.xlsx]")
        sys.exit(1)

    inp = sys.argv[1]
    out = (sys.argv[2] if len(sys.argv) == 3
           else str(pathlib.Path(inp).with_stem(pathlib.Path(inp).stem + "_cleaned")
                   .with_suffix(".xlsx")))
    main(inp, out)
