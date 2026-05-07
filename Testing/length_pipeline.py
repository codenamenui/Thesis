#!/usr/bin/env python3
"""
length_pipeline.py
------------------
Length‑bias verification pipeline (truncation only).

    1. TOKENIZE  — tokenize raw text using jcblaise/bert-tagalog-base-cased
    2. TRUNCATE  — cap all train articles to a global token threshold
                    (auto‑computed from train set, or supplied manually via --cap)
    3. CLASSIFY  — logistic regression on token_count
                    → overall test accuracy  +  per‑subclass test accuracy

                     BEFORE (pre‑truncation)   &   AFTER (post‑truncation)

Produces per‑subclass token statistics and an Excel report.

Usage:
    python length_pipeline.py stratified_dataset.xlsx --cap 271
"""

import argparse
import re
import pathlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score

# ── Constants ─────────────────────────────────────────────────────────────────
TAGALOG_BERT_MODEL   = "jcblaise/bert-tagalog-base-cased"
TEXT_COL_CANDIDATES  = ["text", "content", "article", "body", "sentence", "news"]
LABEL_MAP            = {"HR": 1, "AI-R": 1, "HF": 0, "AI-F": 0}
SUBCLASSES           = ["HR", "AI-R", "HF", "AI-F"]
SEED                 = 42
rng                  = np.random.default_rng(SEED)

# Colours for Excel
HDR_FILL    = "1F4E79"
GREEN_FILL  = "C6EFCE"
BLUE_FILL   = "BDD7EE"
YELLOW_FILL = "FFEB9C"
RED_FILL    = "FFC7CE"
GREY_FILL   = "D9D9D9"
ALT_ROW     = "F2F2F2"
NT_FILLS    = {"HR": "C6EFCE", "AI-R": "FFEB9C", "HF": "FFC7CE", "AI-F": "E2AFFF"}

# ── Tokenizer ─────────────────────────────────────────────────────────────────
_tokenizer = None

def get_tokenizer():
    global _tokenizer
    if _tokenizer is None:
        from transformers import AutoTokenizer
        print(f"  Loading tokenizer: {TAGALOG_BERT_MODEL} …")
        _tokenizer = AutoTokenizer.from_pretrained(TAGALOG_BERT_MODEL)
        print("  Tokenizer ready.")
    return _tokenizer


def tokenize_texts(texts: list[str], batch_size: int = 128) -> list[list[int]]:
    tok     = get_tokenizer()
    all_ids = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i : i + batch_size]
        enc   = tok(batch, add_special_tokens=True, truncation=False,
                    padding=False, return_attention_mask=False)
        all_ids.extend(enc["input_ids"])
        if i > 0 and i % 2048 == 0:
            print(f"    tokenized {i}/{len(texts)} rows …")
    return all_ids


def decode_ids(ids: list[int]) -> str:
    return get_tokenizer().decode(ids, skip_special_tokens=True)


def find_text_col(cols: list[str]) -> str | None:
    lower = [c.lower() for c in cols]
    for cand in TEXT_COL_CANDIDATES:
        if cand in lower:
            return cols[lower.index(cand)]
    skip = {"label", "split", "news_type", "id", "index", "token_count",
            "original_article", "topic", "_ids", "_text_col"}
    others = [c for c in cols if c.lower() not in skip]
    return others[0] if others else None


# ── Step 1 — Tokenize ─────────────────────────────────────────────────────────
def step_tokenize(sheets: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    print("\n  [1/3] TOKENIZE")
    out = {}
    for sheet_name, df in sheets.items():
        text_col = find_text_col(list(df.columns))
        if text_col is None:
            print(f"        [SKIP] '{sheet_name}' — no text column found.")
            continue
        print(f"        '{sheet_name}' — {len(df)} rows …")
        texts        = df[text_col].fillna("").astype(str).tolist()
        all_ids      = tokenize_texts(texts)
        df           = df.copy()
        df["_ids"]   = all_ids
        df["token_count"] = [len(ids) for ids in all_ids]
        df["_text_col"]   = text_col
        out[sheet_name]   = df
    return out


# ── Step 2 — Truncate ─────────────────────────────────────────────────────────
def compute_global_threshold(sheets: dict[str, pd.DataFrame]) -> tuple[int, str]:
    category_tokens: dict[str, list] = {}
    for df in sheets.values():
        train_df = df[df["split"] == "train"]
        for nt in SUBCLASSES:
            tokens = train_df[train_df["news_type"] == nt]["token_count"].values
            if len(tokens) == 0:
                continue
            category_tokens.setdefault(nt, [])
            category_tokens[nt].extend(tokens.tolist())
    if not category_tokens:
        raise ValueError("No train rows found — cannot compute threshold.")
    means       = {nt: float(np.mean(vals)) for nt, vals in category_tokens.items()}
    shortest_nt = min(means, key=means.get)
    threshold   = int(means[shortest_nt])
    description = (f"mean of {shortest_nt} (train only) across all sheets "
                   f"({', '.join(f'{k}={v:.0f}' for k, v in sorted(means.items()))})")
    return threshold, description


def truncate_ids(ids: list[int], cap: int) -> list[int]:
    if len(ids) <= cap:
        return ids
    return ids[:cap - 1] + [ids[-1]]   # preserve [CLS] … [SEP]


def step_truncate(sheets: dict[str, pd.DataFrame],
                  manual_cap: int | None = None) -> tuple[dict[str, pd.DataFrame], int, str]:
    if manual_cap is not None:
        cap      = manual_cap
        cap_desc = f"fixed cap = {cap} (HF mean, full corpus)"
    else:
        cap, cap_desc = compute_global_threshold(sheets)

    print(f"\n  [2/3] TRUNCATE  →  global cap = {cap} tokens (train only)")
    print(f"        ({cap_desc})")

    out = {}
    for sheet_name, df in sheets.items():
        df       = df.copy()
        text_col = df["_text_col"].iloc[0]
        is_train = df["split"] == "train"

        train_trunc = df.loc[is_train, "_ids"].apply(lambda ids: truncate_ids(ids, cap))
        df.loc[is_train, text_col]      = train_trunc.apply(decode_ids)
        df.loc[is_train, "token_count"] = train_trunc.apply(len)
        df.loc[is_train, "_ids"]        = train_trunc

        pct = (df.loc[is_train, "token_count"] == cap).mean()
        n_test = (~is_train).sum()
        print(f"        '{sheet_name}': {pct:.1%} of train rows truncated  |  {n_test} test rows untouched")
        out[sheet_name] = df
    return out, cap, cap_desc


# ── Step 3 — Classify (overall + per subclass) ────────────────────────────────
def _classify_one_sheet(df: pd.DataFrame, sheet_name: str):
    """Returns a dict with overall test accuracy and per‑subclass accuracies."""
    train = df[df["split"] == "train"]
    test  = df[df["split"] == "test"]
    if len(train) == 0 or len(test) == 0 or train["label"].nunique() < 2:
        return None

    clf = LogisticRegression(random_state=SEED, max_iter=1000)
    clf.fit(train[["token_count"]], train["label"])
    pred_test = clf.predict(test[["token_count"]])

    # overall
    overall_acc = accuracy_score(test["label"], pred_test)

    # per subclass
    subclass_acc = {}
    for nt in SUBCLASSES:
        mask = test["news_type"] == nt
        if mask.sum() == 0:
            subclass_acc[nt] = None
        else:
            subclass_acc[nt] = accuracy_score(test.loc[mask, "label"],
                                              pred_test[mask])

    return {
        "sheet":        sheet_name,
        "overall_acc":  overall_acc,
        "HR_acc":       subclass_acc.get("HR"),
        "AI-R_acc":     subclass_acc.get("AI-R"),
        "HF_acc":       subclass_acc.get("HF"),
        "AI-F_acc":     subclass_acc.get("AI-F"),
    }


def step_classify(pre_sheets: dict[str, pd.DataFrame],
                  post_sheets: dict[str, pd.DataFrame]) -> tuple[list[dict], list[dict]]:
    print("\n  [3/3] CLASSIFY (overall + per subclass)")
    before, after = [], []
    for sheet_name in pre_sheets:
        res_before = _classify_one_sheet(pre_sheets[sheet_name], sheet_name)
        res_after  = _classify_one_sheet(post_sheets[sheet_name], sheet_name)
        if res_before:
            before.append(res_before)
        if res_after:
            after.append(res_after)
        if res_before and res_after:
            print(f"        {sheet_name}: overall Before={res_before['overall_acc']:.4f}  →  After={res_after['overall_acc']:.4f}")
    return before, after


# ── Subclass token statistics helpers ─────────────────────────────────────────
def _compute_subclass_stats(sheets: dict[str, pd.DataFrame]) -> dict:
    all_data = pd.concat(sheets.values(), ignore_index=True)
    stats = {}
    for nt in SUBCLASSES:
        tc = all_data[all_data["news_type"] == nt]["token_count"]
        if len(tc) == 0:
            continue
        stats[nt] = {
            "count": int(len(tc)),
            "mean": float(tc.mean()),
            "median": float(tc.median()),
            "std": float(tc.std()),
            "min": int(tc.min()),
            "max": int(tc.max()),
        }
    return stats


# ── Sheet sorting key ─────────────────────────────────────────────────────────
def _parse_proportions(sheet_name: str):
    m = re.match(r'HR(\d+)-AIR(\d+)-HF(\d+)-AIF(\d+)', sheet_name)
    if m:
        return tuple(int(m.group(i)) for i in range(1, 5))
    return (0,0,0,0)

def sheet_sort_key(name):
    hr, air, hf, aif = _parse_proportions(name)
    real_order = {100: 0, 67: 1, 50: 2}
    fake_order = {100: 0, 67: 1, 50: 2, 33: 3, 0: 4}
    return (real_order.get(hr, 99), fake_order.get(hf, 99), name)


# ── Excel helpers ─────────────────────────────────────────────────────────────
def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=HDR_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")

def _cell(ws, row, col, value, fill_hex=None, bold=False, fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fill_hex: c.fill = PatternFill("solid", start_color=fill_hex)
    if fmt:       c.number_format = fmt

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_overall_accuracy_sheet(ws, before: list[dict], after: list[dict],
                                 cap: int, cap_desc: str):
    ws.title = "Overall Accuracy"
    headers = ["Sheet", "Before Overall Acc", "After Overall Acc", "Δ Overall Acc"]
    _widths(ws, [42, 18, 17, 12])
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)

    b_map = {r["sheet"]: r for r in before}
    a_map = {r["sheet"]: r for r in after}
    ri = 2
    for sheet_name in sorted(b_map, key=sheet_sort_key):
        b = b_map[sheet_name]
        a = a_map.get(sheet_name)
        if a is None: continue
        delta = a["overall_acc"] - b["overall_acc"]
        d_fill = GREEN_FILL if delta < -0.05 else (RED_FILL if delta > 0.05 else GREY_FILL)
        _cell(ws, ri, 1, sheet_name, bold=True, align="left")
        _cell(ws, ri, 2, b["overall_acc"], fill_hex=YELLOW_FILL, fmt="0.00%")
        _cell(ws, ri, 3, a["overall_acc"], fill_hex=BLUE_FILL,   fmt="0.00%")
        _cell(ws, ri, 4, delta,            fill_hex=d_fill, fmt="+0.00%;-0.00%;0.00%")
        ri += 1

    b_accs = [b_map[s]["overall_acc"] for s in b_map if s in a_map]
    a_accs = [a_map[s]["overall_acc"] for s in b_map if s in a_map]
    ri += 1
    _cell(ws, ri, 1, "MEAN", bold=True)
    _cell(ws, ri, 2, np.mean(b_accs), bold=True, fill_hex=YELLOW_FILL, fmt="0.00%")
    _cell(ws, ri, 3, np.mean(a_accs), bold=True, fill_hex=BLUE_FILL,   fmt="0.00%")
    delta_mean = np.mean(a_accs) - np.mean(b_accs)
    _cell(ws, ri, 4, delta_mean, bold=True,
          fill_hex=GREEN_FILL if delta_mean < 0 else RED_FILL, fmt="+0.00%;-0.00%;0.00%")

    ri += 2
    ws.cell(row=ri, column=1, value=f"Global truncation cap: {cap} tokens  |  {cap_desc}").font = Font(name="Arial", italic=True, size=9, color="595959")
    ri += 1
    ws.cell(row=ri, column=1, value="Δ Acc < 0 (green) = length became less predictive after truncation.").font = Font(name="Arial", italic=True, size=9, color="375623")
    ws.freeze_panes = "A2"


def write_subclass_accuracy_sheet(ws, before: list[dict], after: list[dict],
                                  cap: int, cap_desc: str):
    ws.title = "Subclass Accuracy"
    # Columns: Sheet | HR Before | HR After | AI‑R Before | AI‑R After | HF Before | HF After | AI‑F Before | AI‑F After
    headers = ["Sheet"]
    for nt in SUBCLASSES:
        headers.append(f"{nt} Before")
        headers.append(f"{nt} After")
    _widths(ws, [42] + [15]* (len(headers)-1))
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)

    b_map = {r["sheet"]: r for r in before}
    a_map = {r["sheet"]: r for r in after}
    ri = 2
    for sheet_name in sorted(b_map, key=sheet_sort_key):
        b = b_map[sheet_name]
        a = a_map.get(sheet_name)
        if a is None: continue
        _cell(ws, ri, 1, sheet_name, bold=True, align="left")
        col = 2
        for nt in SUBCLASSES:
            nt_b = b.get(f"{nt}_acc")
            nt_a = a.get(f"{nt}_acc")
            _cell(ws, ri, col,   f"{nt_b:.4f}" if nt_b is not None else "—", fill_hex=YELLOW_FILL if nt_b is not None else GREY_FILL, fmt="0.00%")
            _cell(ws, ri, col+1, f"{nt_a:.4f}" if nt_a is not None else "—", fill_hex=BLUE_FILL if nt_a is not None else GREY_FILL, fmt="0.00%")
            col += 2
        ri += 1

    ri += 1
    ws.cell(row=ri, column=1, value=f"Cap: {cap} tokens ({cap_desc})").font = Font(name="Arial", italic=True, size=9, color="595959")
    ws.freeze_panes = "A2"


def write_token_stats_sheet(ws, sheets: dict[str, pd.DataFrame], title: str):
    ws.title = title[:31]
    headers  = ["Sheet", "Split", "News Type", "n", "Mean", "Std", "Min", "Max", "Median"]
    _widths(ws, [42, 10, 12, 8, 10, 10, 8, 8, 10])
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    SPLIT_FILLS = {"train": GREEN_FILL, "test": BLUE_FILL}
    ri = 2
    for sheet_name in sorted(sheets.keys(), key=sheet_sort_key):
        df = sheets[sheet_name]
        for split in ["train", "test"]:
            sdf = df[df["split"] == split]
            first = True
            for nt in SUBCLASSES:
                grp = sdf[sdf["news_type"] == nt]["token_count"]
                if len(grp) == 0: continue
                fill = NT_FILLS.get(nt, ALT_ROW)
                _cell(ws, ri, 1, sheet_name if first else "", bold=first, align="left")
                _cell(ws, ri, 2, split if first else "", fill_hex=SPLIT_FILLS.get(split) if first else None)
                _cell(ws, ri, 3, nt, fill_hex=fill)
                _cell(ws, ri, 4, len(grp), fill_hex=fill)
                _cell(ws, ri, 5, round(grp.mean(), 1), fill_hex=fill, fmt="0.0")
                _cell(ws, ri, 6, round(grp.std(),  1), fill_hex=fill, fmt="0.0")
                _cell(ws, ri, 7, int(grp.min()), fill_hex=fill)
                _cell(ws, ri, 8, int(grp.max()), fill_hex=fill)
                _cell(ws, ri, 9, round(grp.median(),1), fill_hex=fill, fmt="0.0")
                ri += 1
                first = False
    ws.freeze_panes = "A2"


def write_truncation_diag_sheet(ws, sheets_before, sheets_after, cap, cap_desc):
    ws.title = "Truncation Diagnostics"
    headers  = ["Sheet", "News Type", "n", "Orig Median", "Post-Trunc Median",
                "Cap", "% Rows at Cap", "Cap Source"]
    _widths(ws, [42, 12, 8, 14, 18, 8, 16, 50])
    for ci, h in enumerate(headers, 1): _hdr(ws, 1, ci, h)
    ri = 2
    for sheet_name in sorted(sheets_before, key=sheet_sort_key):
        bdf = sheets_before[sheet_name]
        adf = sheets_after.get(sheet_name, pd.DataFrame())
        for nt in SUBCLASSES:
            b_grp = bdf[bdf["news_type"] == nt]["token_count"]
            a_grp = adf[adf["news_type"] == nt]["token_count"] if len(adf) else pd.Series([], dtype=int)
            if len(b_grp) == 0: continue
            pct = (a_grp == cap).mean() if len(a_grp) else float("nan")
            fill = NT_FILLS.get(nt, ALT_ROW)
            p_fill = RED_FILL if pct > 0.5 else (YELLOW_FILL if pct > 0.2 else GREEN_FILL)
            _cell(ws, ri, 1, sheet_name, bold=True, align="left")
            _cell(ws, ri, 2, nt, fill_hex=fill)
            _cell(ws, ri, 3, len(b_grp), fill_hex=fill)
            _cell(ws, ri, 4, round(b_grp.median(), 1), fill_hex=fill, fmt="0.0")
            _cell(ws, ri, 5, round(a_grp.median(), 1) if len(a_grp) else "—", fill_hex=fill, fmt="0.0")
            _cell(ws, ri, 6, cap, fill_hex=YELLOW_FILL, bold=True)
            _cell(ws, ri, 7, pct, fill_hex=p_fill, fmt="0.0%")
            _cell(ws, ri, 8, cap_desc, align="left")
            ri += 1
    ri += 2
    for note in [
        "Cap = mean token count of the category with the lowest train mean across all sheets.",
        "Green = <20% rows truncated  |  Yellow = 20–50%  |  Red = >50%"
    ]:
        ws.cell(row=ri, column=1, value=note).font = Font(name="Arial", italic=True, size=9, color="595959")
        ri += 1
    ws.freeze_panes = "A2"


def write_subclass_token_summary_sheet(ws, before_stats, after_stats, cap, cap_desc):
    ws.title = "Subclass Token Summary"
    headers = ["Stage", "News Type", "Count", "Mean", "Median", "Std Dev", "Min", "Max"]
    _widths(ws, [16, 12, 10, 12, 12, 12, 8, 8])
    for ci, h in enumerate(headers, 1): _hdr(ws, 1, ci, h)
    ri = 2
    for stage_label, stats in [("Before Truncation", before_stats), ("After Truncation", after_stats)]:
        for nt in SUBCLASSES:
            if nt not in stats: continue
            s = stats[nt]
            fill = NT_FILLS.get(nt, ALT_ROW)
            _cell(ws, ri, 1, stage_label if nt == "HR" else "", bold=True, align="left")
            _cell(ws, ri, 2, nt, fill_hex=fill)
            _cell(ws, ri, 3, s["count"], fill_hex=fill)
            _cell(ws, ri, 4, f"{s['mean']:.1f}", fill_hex=fill)
            _cell(ws, ri, 5, f"{s['median']:.1f}", fill_hex=fill)
            _cell(ws, ri, 6, f"{s['std']:.1f}", fill_hex=fill)
            _cell(ws, ri, 7, s["min"], fill_hex=fill)
            _cell(ws, ri, 8, s["max"], fill_hex=fill)
            ri += 1
    ri += 1
    ws.cell(row=ri, column=1, value=f"Cap: {cap} tokens ({cap_desc})").font = Font(name="Arial", italic=True, size=9, color="595959")
    ws.freeze_panes = "A2"


# ── Main ──────────────────────────────────────────────────────────────────────
def load_sheets(path: str) -> dict[str, pd.DataFrame]:
    xl   = pd.ExcelFile(path)
    skip = {"summary", "overall accuracy", "subclass accuracy", "truncation diagnostics",
            "token stats (before)", "token stats (after)", "subclass token summary"}
    sheets = {}
    for name in xl.sheet_names:
        if name.lower() in skip: continue
        df = xl.parse(name)
        df.columns = df.columns.str.strip()
        col_map = {c.lower(): c for c in df.columns}
        for key in ("news_type", "split", "label"):
            if key in col_map and col_map[key] != key:
                df.rename(columns={col_map[key]: key}, inplace=True)
        if "news_type" not in df.columns or "split" not in df.columns:
            print(f"  [SKIP] '{name}' — missing news_type or split.")
            continue
        df["news_type"] = df["news_type"].astype(str).str.upper().str.strip()
        df["split"]     = df["split"].astype(str).str.lower().str.strip()
        df = df[df["split"].isin({"train", "test", "val"})].copy()
        if len(df) == 0: continue
        sheets[name] = df
    return sheets


def print_detailed_summary(before, after, cap, before_subclass, after_subclass):
    print(f"\n{'═' * 72}")
    print(f"  RESULTS  (global cap = {cap} tokens)")
    print(f"{'═' * 72}")
    b_map = {r["sheet"]: r for r in before}
    a_map = {r["sheet"]: r for r in after}
    # Overall
    print("\n  OVERALL ACCURACY:")
    print(f"  {'Sheet':<40} {'Before':>10} {'After':>10} {'Δ':>8}")
    print(f"  {'─' * 70}")
    for sheet in sorted(b_map, key=sheet_sort_key):
        if sheet not in a_map: continue
        b = b_map[sheet]["overall_acc"]
        a = a_map[sheet]["overall_acc"]
        arrow = "▼" if (a-b) < -0.01 else ("▲" if (a-b) > 0.01 else "~")
        print(f"  {sheet:<40} {b:>10.4f} {a:>10.4f} {a-b:>+8.4f} {arrow}")
    b_accs = [b_map[s]["overall_acc"] for s in b_map if s in a_map]
    a_accs = [a_map[s]["overall_acc"] for s in b_map if s in a_map]
    print(f"  {'─' * 70}")
    print(f"  {'MEAN':<40} {np.mean(b_accs):>10.4f} {np.mean(a_accs):>10.4f} "
          f"{np.mean(a_accs)-np.mean(b_accs):>+8.4f}")

    # Per subclass
    print("\n  SUBCLASS ACCURACY (HR / AI‑R / HF / AI‑F):")
    print(f"  {'Sheet':<40} {'HR Before':>10} {'HR After':>10}  |  {'AI-R Before':>10} {'AI-R After':>10}  |  {'HF Before':>10} {'HF After':>10}  |  {'AI-F Before':>10} {'AI-F After':>10}")
    print(f"  {'─' * 130}")
    for sheet in sorted(b_map, key=sheet_sort_key):
        if sheet not in a_map: continue
        b = b_map[sheet]; a = a_map[sheet]
        def _fmt(v): return f"{v:.4f}" if v is not None else "  —"
        print(f"  {sheet:<40} {_fmt(b.get('HR_acc')):>10} {_fmt(a.get('HR_acc')):>10}  |  "
              f"{_fmt(b.get('AI-R_acc')):>10} {_fmt(a.get('AI-R_acc')):>10}  |  "
              f"{_fmt(b.get('HF_acc')):>10} {_fmt(a.get('HF_acc')):>10}  |  "
              f"{_fmt(b.get('AI-F_acc')):>10} {_fmt(a.get('AI-F_acc')):>10}")

    # Token stats (unchanged)
    print("\n  Per‑Subclass Mean Token Counts:")
    for label, stats in [("Before", before_subclass), ("After", after_subclass)]:
        print(f"  {label}:")
        for nt in SUBCLASSES:
            if nt in stats:
                print(f"    {nt:<6}  mean = {stats[nt]['mean']:.1f}  "
                      f"(median = {stats[nt]['median']:.1f}, std = {stats[nt]['std']:.1f})")
    print()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input_file", nargs="?", default=None,
                        help="Path to stratified_dataset.xlsx")
    parser.add_argument("--cap", type=int, default=None,
                        help="Override the global truncation cap (tokens).")
    args = parser.parse_args()

    path = args.input_file or input("Path to stratified_dataset.xlsx: ").strip()
    print(f"\n  Loading '{path}' …")
    raw_sheets = load_sheets(path)
    sorted_names = sorted(raw_sheets.keys(), key=sheet_sort_key)
    sheets = {k: raw_sheets[k] for k in sorted_names}
    print(f"  Found {len(sheets)} valid sheet(s).")

    # 1. Tokenize
    tok_sheets = step_tokenize(sheets)
    pre_clf = {name: df[["token_count", "label", "split", "news_type"]].copy()
               for name, df in tok_sheets.items()}

    # 2. Truncate
    trunc_sheets, cap, cap_desc = step_truncate(tok_sheets, manual_cap=args.cap)
    post_clf = {name: df[["token_count", "label", "split", "news_type"]].copy()
                for name, df in trunc_sheets.items()}

    # Subclass token statistics
    before_subclass = _compute_subclass_stats(pre_clf)
    after_subclass  = _compute_subclass_stats(post_clf)

    # 3. Classify (overall + per subclass)
    before_results, after_results = step_classify(pre_clf, post_clf)

    # Console output
    print_detailed_summary(before_results, after_results, cap,
                           before_subclass, after_subclass)

    # Excel output
    wb = Workbook()
    wb.remove(wb.active)

    write_overall_accuracy_sheet(wb.create_sheet("Overall Accuracy"),
                                 before_results, after_results, cap, cap_desc)
    write_subclass_accuracy_sheet(wb.create_sheet("Subclass Accuracy"),
                                  before_results, after_results, cap, cap_desc)
    write_token_stats_sheet(wb.create_sheet("Token Stats (Before)"), pre_clf, "Token Stats (Before)")
    write_token_stats_sheet(wb.create_sheet("Token Stats (After)"),  post_clf, "Token Stats (After)")
    write_truncation_diag_sheet(wb.create_sheet("Truncation Diagnostics"),
                                pre_clf, post_clf, cap, cap_desc)
    write_subclass_token_summary_sheet(wb.create_sheet("Subclass Token Summary"),
                                       before_subclass, after_subclass, cap, cap_desc)

    out_path = pathlib.Path("length_pipeline_results.xlsx")
    wb.save(str(out_path))
    print(f"  ✓ Saved → {out_path}")
    if args.cap:
        print(f"  Cap override: {args.cap} tokens")


if __name__ == "__main__":
    main()