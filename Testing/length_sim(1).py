"""
length_sim.py
-------------
Generates a simulated dataset where AI-F and AI-R token counts are drawn from
the same distribution as their human counterparts (HF and HR respectively),
making article length a weaker predictor of fake vs real.

Strategy
--------
  AI-F  →  sampled from HF distribution  (same mean, std, min, max per split)
  AI-R  →  sampled from HR distribution  (same mean, std, min, max per split)

This maximally reduces the length signal: the classifier cannot distinguish
AI from human by length alone because the distributions fully overlap.

Input options
-------------
  Option A — original row-level dataset (one row per article, raw text):
    The script will tokenize text using jcblaise/bert-tagalog-base-cased.
    Required columns: text (or content/article/body), label, split, news_type
    python length_sim.py stratified_dataset.xlsx

  Option B — length_classifier_results.xlsx (Token Stats sheet):
    Uses pre-computed mean/std/min/max/n stats to synthesise rows.
    python length_sim.py length_classifier_results.xlsx

The script auto-detects which format it receives.
Output: simulated_lengths.xlsx with columns token_count, label, split, news_type.

Dependencies
------------
    pip install transformers torch openpyxl pandas numpy scikit-learn
"""

import sys
import pathlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Tagalog BERT tokenizer (lazy-loaded only when needed for raw text input)
TAGALOG_BERT_MODEL = "jcblaise/bert-tagalog-base-cased"
_tokenizer = None

def get_tokenizer():
    global _tokenizer
    if _tokenizer is None:
        try:
            from transformers import AutoTokenizer
            print(f"  Loading tokenizer: {TAGALOG_BERT_MODEL} …")
            _tokenizer = AutoTokenizer.from_pretrained(TAGALOG_BERT_MODEL)
            print(f"  Tokenizer ready.")
        except ImportError:
            raise ImportError("Install transformers: pip install transformers torch")
    return _tokenizer

def count_tokens(texts: list[str]) -> list[int]:
    """Tokenize a list of texts and return token counts (excluding special tokens)."""
    tok = get_tokenizer()
    counts = []
    batch_size = 128
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        enc = tok(batch, add_special_tokens=True, truncation=False,
                  padding=False, return_attention_mask=False)
        counts.extend(len(ids) for ids in enc["input_ids"])
        if i % 1024 == 0 and i > 0:
            print(f"    Tokenized {i}/{len(texts)} rows …")
    return counts

TEXT_COLUMN_CANDIDATES = ["text", "content", "article", "body", "sentence", "news"]

SEED = 42
rng  = np.random.default_rng(SEED)

# ── Colour palette ────────────────────────────────────────────────────────────
HDR_FILL    = "1F4E79"
GREEN_FILL  = "C6EFCE"
BLUE_FILL   = "BDD7EE"
YELLOW_FILL = "FFEB9C"
RED_FILL    = "FFC7CE"
ALT_ROW     = "F2F2F2"

NT_FILLS = {
    "HR":   "C6EFCE",
    "AI-R": "FFEB9C",
    "HF":   "FFC7CE",
    "AI-F": "E2AFFF",
}
SPLIT_FILLS = {
    "train": GREEN_FILL,
    "val":   YELLOW_FILL,
    "test":  BLUE_FILL,
}

LABEL_MAP = {"HR": 1, "AI-R": 1, "HF": 0, "AI-F": 0}


# ── Sampling helpers ──────────────────────────────────────────────────────────

def _clamp(arr: np.ndarray, lo: int, hi: int) -> np.ndarray:
    """Clip to observed min/max and ensure positive integers."""
    return np.clip(np.round(arr).astype(int), max(lo, 1), hi)


def _sample_like(ref_tokens: np.ndarray, n: int) -> np.ndarray:
    """
    Draw n token counts from a distribution matching ref_tokens.
    Uses a truncated normal parameterised by ref's mean and std,
    clipped to ref's [min, max] range.
    Falls back to bootstrapping when std is 0 or n is very small.
    """
    mean = float(ref_tokens.mean())
    std  = float(ref_tokens.std())
    lo   = int(ref_tokens.min())
    hi   = int(ref_tokens.max())

    if std < 1 or n == 0:
        return rng.choice(ref_tokens, size=n, replace=True)

    samples = rng.normal(loc=mean, scale=std, size=n * 3)
    samples = _clamp(samples, lo, hi)

    # Filter to range, then take exactly n (pad with bootstrap if needed)
    in_range = samples[(samples >= lo) & (samples <= hi)]
    if len(in_range) < n:
        extra = rng.choice(ref_tokens, size=n - len(in_range), replace=True)
        in_range = np.concatenate([in_range, extra])
    return in_range[:n]


# ── Per-sheet simulation ──────────────────────────────────────────────────────

def simulate_sheet(sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.str.lower().str.strip()

    required = {"token_count", "label", "split", "news_type"}
    missing  = required - set(df.columns)
    if missing:
        raise ValueError(f"Sheet '{sheet_name}' missing columns: {missing}")

    df["news_type"] = df["news_type"].str.upper().str.strip()
    df["split"]     = df["split"].str.lower().str.strip()

    target_splits = {"train", "test"}
    df = df[df["split"].isin(target_splits)].copy()

    rows = []
    for split, split_df in df.groupby("split"):
        nt_groups = {nt: g for nt, g in split_df.groupby("news_type")}

        # Reference distributions from human categories
        hf_tokens = split_df.loc[split_df["news_type"] == "HF",  "token_count"].values
        hr_tokens = split_df.loc[split_df["news_type"] == "HR",  "token_count"].values

        for nt, group in nt_groups.items():
            n = len(group)
            original_tokens = group["token_count"].values

            if nt == "AI-F":
                if len(hf_tokens) == 0:
                    print(f"  [WARN] {sheet_name}/{split}: no HF reference; keeping AI-F lengths")
                    new_tokens = original_tokens
                else:
                    new_tokens = _sample_like(hf_tokens, n)

            elif nt == "AI-R":
                if len(hr_tokens) == 0:
                    print(f"  [WARN] {sheet_name}/{split}: no HR reference; keeping AI-R lengths")
                    new_tokens = original_tokens
                else:
                    new_tokens = _sample_like(hr_tokens, n)

            else:
                # HR and HF keep their original lengths unchanged
                new_tokens = original_tokens

            for tok in new_tokens:
                rows.append({
                    "token_count": int(tok),
                    "label":       LABEL_MAP[nt],
                    "split":       split,
                    "news_type":   nt,
                })

    return pd.DataFrame(rows, columns=["token_count", "label", "split", "news_type"])


# ── Excel formatting helpers ──────────────────────────────────────────────────

def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=HDR_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _cell(ws, row, col, value, fill_hex=None, bold=False, fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fill_hex:
        c.fill = PatternFill("solid", start_color=fill_hex)
    if fmt:
        c.number_format = fmt


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Excel writers ─────────────────────────────────────────────────────────────

def write_data_sheet(ws, sheet_name: str, df: pd.DataFrame) -> None:
    ws.title = sheet_name[:31]
    headers  = ["token_count", "label", "split", "news_type"]
    widths   = [14, 8, 10, 12]
    _set_widths(ws, widths)

    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 24

    for ri, row in enumerate(df.itertuples(index=False), 2):
        nt    = row.news_type
        fill  = NT_FILLS.get(nt, ALT_ROW)
        _cell(ws, ri, 1, row.token_count, fill_hex=fill)
        _cell(ws, ri, 2, row.label,       fill_hex=fill)
        _cell(ws, ri, 3, row.split,       fill_hex=SPLIT_FILLS.get(row.split))
        _cell(ws, ri, 4, nt,              fill_hex=fill)

    ws.freeze_panes = "A2"


def write_summary_sheet(ws, summaries: list[dict]) -> None:
    ws.title = "Simulation Summary"
    headers  = [
        "Original Sheet", "Split", "News Type",
        "n", "Orig Mean", "Sim Mean", "Orig Std", "Sim Std",
        "Orig Min", "Sim Min", "Orig Max", "Sim Max",
    ]
    widths = [38, 8, 12, 8, 11, 11, 10, 10, 10, 10, 10, 10]
    _set_widths(ws, widths)

    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 28

    ri = 2
    for entry in summaries:
        nt   = entry["news_type"]
        fill = NT_FILLS.get(nt, ALT_ROW)
        _cell(ws, ri, 1,  entry["sheet"],     bold=True, align="left")
        _cell(ws, ri, 2,  entry["split"],     fill_hex=SPLIT_FILLS.get(entry["split"]))
        _cell(ws, ri, 3,  nt,                 fill_hex=fill)
        _cell(ws, ri, 4,  entry["n"],         fill_hex=fill)
        _cell(ws, ri, 5,  round(entry["orig_mean"], 1), fill_hex=fill, fmt="0.0")
        _cell(ws, ri, 6,  round(entry["sim_mean"],  1), fill_hex=fill, fmt="0.0")
        _cell(ws, ri, 7,  round(entry["orig_std"],  1), fill_hex=fill, fmt="0.0")
        _cell(ws, ri, 8,  round(entry["sim_std"],   1), fill_hex=fill, fmt="0.0")
        _cell(ws, ri, 9,  entry["orig_min"],  fill_hex=fill)
        _cell(ws, ri, 10, entry["sim_min"],   fill_hex=fill)
        _cell(ws, ri, 11, entry["orig_max"],  fill_hex=fill)
        _cell(ws, ri, 12, entry["sim_max"],   fill_hex=fill)
        ri += 1

    # Legend
    note_row = ri + 2
    notes = [
        "AI-F token counts are resampled from the HF (human-fake) distribution per split.",
        "AI-R token counts are resampled from the HR (human-real) distribution per split.",
        "HR and HF token counts are unchanged from the original dataset.",
        "Goal: length becomes a near-chance predictor → MCC ≈ 0 in linear_preview.py.",
    ]
    for i, note in enumerate(notes):
        c = ws.cell(row=note_row + i, column=1, value=note)
        c.font = Font(name="Arial", italic=True, size=9, color="595959")
        ws.merge_cells(
            start_row=note_row + i, start_column=1,
            end_row=note_row + i,   end_column=12,
        )

    ws.freeze_panes = "A2"


# ── Console summary ───────────────────────────────────────────────────────────

def print_summary(sheet_name: str, orig_df: pd.DataFrame, sim_df: pd.DataFrame) -> None:
    print(f"\n  {'─' * 60}")
    print(f"  Sheet: {sheet_name}")
    print(f"  {'NT':<6} {'Split':<6} {'n':>5} | "
          f"{'OrigMean':>9} {'SimMean':>9} | {'OrigStd':>8} {'SimStd':>8}")
    print(f"  {'─' * 60}")

    orig_df = orig_df.copy()
    sim_df  = sim_df.copy()
    orig_df["news_type"] = orig_df["news_type"].str.upper().str.strip()
    sim_df["news_type"]  = sim_df["news_type"].str.upper().str.strip()

    for split in ["train", "test"]:
        for nt in ["HR", "AI-R", "HF", "AI-F"]:
            o = orig_df[(orig_df["split"] == split) & (orig_df["news_type"] == nt)]["token_count"]
            s = sim_df[ (sim_df["split"]  == split) & (sim_df["news_type"]  == nt)]["token_count"]
            if len(o) == 0 and len(s) == 0:
                continue
            om = o.mean() if len(o) else float("nan")
            sm = s.mean() if len(s) else float("nan")
            os_ = o.std()  if len(o) else float("nan")
            ss  = s.std()  if len(s) else float("nan")
            tag = " ← resampled" if nt in ("AI-F", "AI-R") else ""
            print(f"  {nt:<6} {split:<6} {len(s):>5} | "
                  f"{om:>9.1f} {sm:>9.1f} | {os_:>8.1f} {ss:>8.1f}{tag}")


# ── Stats-based simulation (Option B: from Token Stats sheet) ─────────────────

def _sample_from_stats(mean: float, std: float, lo: int, hi: int, n: int) -> np.ndarray:
    """Sample n integers from a truncated normal defined by summary stats."""
    if std < 1 or n == 0:
        return np.full(n, int(mean), dtype=int)
    samples = rng.normal(loc=mean, scale=std, size=n * 4)
    samples = _clamp(samples, lo, hi)
    in_range = samples[(samples >= lo) & (samples <= hi)]
    if len(in_range) < n:
        # pad by resampling within range
        extra = rng.integers(lo, hi + 1, size=n - len(in_range))
        in_range = np.concatenate([in_range, extra])
    return in_range[:n]


def simulate_from_token_stats(token_stats_df: pd.DataFrame) -> tuple[dict[str, pd.DataFrame], list[dict]]:
    """
    Build simulated DataFrames from pre-computed Token Stats.
    Returns (sheet_name -> sim_df, summary_entries).
    """
    df = token_stats_df.copy()

    # Forward-fill the Sheet and Split columns (they're merged-cell style)
    df["Sheet"] = df["Sheet"].ffill()
    df["Split"] = df["Split"].ffill()
    df.columns = df.columns.str.strip()

    # Normalise
    df["News Type"] = df["News Type"].astype(str).str.upper().str.strip()
    df["Split"]     = df["Split"].astype(str).str.lower().str.strip()

    target_splits = {"train", "test"}
    df = df[df["Split"].isin(target_splits)]
    df = df[df["News Type"] != "OVERALL"]

    results: dict[str, pd.DataFrame] = {}
    summaries: list[dict] = []

    for sheet_name, sheet_df in df.groupby("Sheet"):
        rows = []

        for split, split_df in sheet_df.groupby("Split"):
            stats: dict[str, dict] = {}
            for _, row in split_df.iterrows():
                nt = row["News Type"]
                stats[nt] = {
                    "mean": float(row["Mean"]),
                    "std":  float(row["Std Dev"]),
                    "min":  int(row["Min"]),
                    "max":  int(row["Max"]),
                    "n":    int(row["n"]),
                }

            # Determine reference distributions
            hf_stats = stats.get("HF")
            hr_stats = stats.get("HR")

            for nt, st in stats.items():
                n = st["n"]
                orig_mean, orig_std = st["mean"], st["std"]
                orig_min,  orig_max = st["min"],  st["max"]

                if nt == "AI-F" and hf_stats:
                    new_tokens = _sample_from_stats(
                        hf_stats["mean"], hf_stats["std"],
                        hf_stats["min"],  hf_stats["max"], n,
                    )
                elif nt == "AI-R" and hr_stats:
                    new_tokens = _sample_from_stats(
                        hr_stats["mean"], hr_stats["std"],
                        hr_stats["min"],  hr_stats["max"], n,
                    )
                else:
                    # HR and HF: reconstruct from their own stats (unchanged)
                    new_tokens = _sample_from_stats(
                        orig_mean, orig_std, orig_min, orig_max, n,
                    )

                label = LABEL_MAP.get(nt, 0)
                for tok in new_tokens:
                    rows.append({
                        "token_count": int(tok),
                        "label":       label,
                        "split":       split,
                        "news_type":   nt,
                    })

                sim_arr = np.array([r["token_count"] for r in rows
                                    if r["split"] == split and r["news_type"] == nt])
                summaries.append({
                    "sheet":     sheet_name,
                    "split":     split,
                    "news_type": nt,
                    "n":         n,
                    "orig_mean": orig_mean,
                    "sim_mean":  float(new_tokens.mean()),
                    "orig_std":  orig_std,
                    "sim_std":   float(new_tokens.std()),
                    "orig_min":  orig_min,
                    "sim_min":   int(new_tokens.min()),
                    "orig_max":  orig_max,
                    "sim_max":   int(new_tokens.max()),
                })

        results[sheet_name] = pd.DataFrame(rows, columns=["token_count", "label", "split", "news_type"])

    return results, summaries


# ── Input format detection ────────────────────────────────────────────────────

def is_results_file(xl: pd.ExcelFile) -> bool:
    """Return True if this looks like length_classifier_results.xlsx."""
    return "Token Stats" in xl.sheet_names


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    paths = sys.argv[1:]
    if not paths:
        path = input("Path to dataset or length_classifier_results.xlsx: ").strip()
    else:
        path = paths[0]

    xl = pd.ExcelFile(path)
    wb = Workbook()
    wb.remove(wb.active)
    all_summaries: list[dict] = []

    if is_results_file(xl):
        # ── Option B: results file with Token Stats sheet ──────────────────
        print(f"\n  Detected results file — reading Token Stats sheet.")
        token_stats_df = xl.parse("Token Stats")
        sim_map, all_summaries = simulate_from_token_stats(token_stats_df)

        print(f"\n  Simulated {len(sim_map)} sheet(s): {', '.join(sim_map)}")
        for sheet_name, sim_df in sim_map.items():
            sim_sheet_name = (sheet_name + " (sim)")[:31]
            write_data_sheet(wb.create_sheet(sim_sheet_name), sim_sheet_name, sim_df)

            # Console summary
            print(f"\n  {'─' * 64}")
            print(f"  Sheet: {sheet_name}")
            print(f"  {'NT':<6} {'Split':<6} {'n':>5} | "
                  f"{'OrigMean':>9} {'SimMean':>9} | {'OrigStd':>8} {'SimStd':>8}")
            print(f"  {'─' * 64}")
            for entry in [e for e in all_summaries if e["sheet"] == sheet_name]:
                tag = " ← resampled" if entry["news_type"] in ("AI-F", "AI-R") else ""
                print(f"  {entry['news_type']:<6} {entry['split']:<6} {entry['n']:>5} | "
                      f"{entry['orig_mean']:>9.1f} {entry['sim_mean']:>9.1f} | "
                      f"{entry['orig_std']:>8.1f} {entry['sim_std']:>8.1f}{tag}")

    else:
        # ── Option A: raw row-level dataset, tokenize text ────────────────
        sheets = [s for s in xl.sheet_names
                  if s.lower() not in ("summary", "length comparison", "simulation summary")]
        print(f"\nFound {len(sheets)} sheet(s): {', '.join(sheets)}")

        for sheet_name in sheets:
            orig_df = xl.parse(sheet_name)
            orig_df.columns = orig_df.columns.str.lower().str.strip()

            # Locate text column
            text_col = next((c for c in TEXT_COLUMN_CANDIDATES if c in orig_df.columns), None)
            if text_col is None:
                # Try any column with string-heavy data
                str_cols = [c for c in orig_df.columns
                            if orig_df[c].dtype == object and c not in ("label", "split", "news_type")]
                if str_cols:
                    text_col = str_cols[0]
                    print(f"  [INFO] Using '{text_col}' as text column for '{sheet_name}'")

            required_meta = {"label", "split", "news_type"}
            missing = required_meta - set(orig_df.columns)
            if missing or text_col is None:
                print(f"  [SKIP] '{sheet_name}' — missing columns: {missing or {'text column'}}")
                continue

            # Filter to train/test only
            orig_df["split"]     = orig_df["split"].astype(str).str.lower().str.strip()
            orig_df["news_type"] = orig_df["news_type"].astype(str).str.upper().str.strip()
            orig_df = orig_df[orig_df["split"].isin({"train", "test"})].copy()

            # Tokenize if token_count not already present
            if "token_count" not in orig_df.columns:
                print(f"\n  Tokenizing '{sheet_name}' ({len(orig_df)} rows) …")
                texts = orig_df[text_col].fillna("").astype(str).tolist()
                orig_df["token_count"] = count_tokens(texts)
            else:
                print(f"\n  '{sheet_name}' already has token_count — skipping tokenization.")

            print(f"  Simulating '{sheet_name}' …")
            sim_df = simulate_sheet(sheet_name, orig_df)
            print_summary(sheet_name, orig_df, sim_df)

            sim_sheet_name = (sheet_name + " (sim)")[:31]
            write_data_sheet(wb.create_sheet(sim_sheet_name), sim_sheet_name, sim_df)

            for split in ["train", "test"]:
                for nt in ["HR", "AI-R", "HF", "AI-F"]:
                    o = orig_df[(orig_df["split"] == split) & (orig_df["news_type"] == nt)]["token_count"]
                    s = sim_df[ (sim_df["split"]  == split) & (sim_df["news_type"]  == nt)]["token_count"]
                    if len(s) == 0:
                        continue
                    all_summaries.append({
                        "sheet":     sheet_name,
                        "split":     split,
                        "news_type": nt,
                        "n":         len(s),
                        "orig_mean": float(o.mean()) if len(o) else float("nan"),
                        "sim_mean":  float(s.mean()),
                        "orig_std":  float(o.std())  if len(o) else float("nan"),
                        "sim_std":   float(s.std()),
                        "orig_min":  int(o.min())    if len(o) else 0,
                        "sim_min":   int(s.min()),
                        "orig_max":  int(o.max())    if len(o) else 0,
                        "sim_max":   int(s.max()),
                    })

    write_summary_sheet(wb.create_sheet("Simulation Summary"), all_summaries)

    out_path = pathlib.Path("simulated_lengths.xlsx")
    wb.save(str(out_path))
    print(f"\n  ✓ Saved → {out_path}")
    print(f"\n  Next step (before-vs-after comparison):")
    print(f"    python linear_preview.py stratified_dataset.xlsx simulated_lengths.xlsx")
    print(f"\n  Or single-file check:")
    print(f"    python linear_preview.py simulated_lengths.xlsx")


if __name__ == "__main__":
    main()