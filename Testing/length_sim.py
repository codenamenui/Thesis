"""
length_sim.py
-------------
Simulates the expected token-length distributions AFTER applying Fix 1 & Fix 2
to the AI generation scripts, WITHOUT calling any API.

Fix 1 (AI-F):  Article lengths are sampled from the HF (human fake) distribution
               within the same split, so AI-F variance mirrors real human variance.

Fix 2 (AI-R):  Article lengths are drawn within ±10 % of the original HR article
               that was rewritten, so AI-R lengths stay close to their source.

Output
------
  simulated_lengths.xlsx   — same structure as the input, but AI-F and AI-R
                             token_count columns replaced with simulated values.
  A console summary comparing before/after mean ± std per news_type per split.

Usage
-----
    python length_sim.py                        # prompts for path
    python length_sim.py stratified_dataset.xlsx
"""

import sys
import pathlib
import random
import warnings

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from transformers import AutoTokenizer
import transformers

warnings.filterwarnings("ignore")
transformers.logging.set_verbosity_error()

# ── Tokenizer ─────────────────────────────────────────────────────────────────

tokenizer = AutoTokenizer.from_pretrained(
    "jcblaise/bert-tagalog-base-cased",
    token=False,
)

def token_count(text: str) -> int:
    return len(tokenizer.encode(str(text), truncation=False))


# ── Helpers ───────────────────────────────────────────────────────────────────

def sample_from_pool(pool: list[int], n: int, rng: np.random.Generator) -> list[int]:
    """Sample n values with replacement from pool."""
    return list(rng.choice(pool, size=n, replace=True))


def jitter_within_pct(value: int, pct: float, rng: np.random.Generator) -> int:
    """Return a value within ±pct of value (integers only, min 1)."""
    lo = max(1, int(value * (1 - pct)))
    hi = max(lo + 1, int(value * (1 + pct)))
    return int(rng.integers(lo, hi))


# ── Core simulation ───────────────────────────────────────────────────────────

def simulate_sheet(df: pd.DataFrame, seed: int = 42) -> pd.DataFrame:
    """
    Returns a copy of df with token_count replaced for AI-F and AI-R rows.

    Rules
    -----
    AI-F  →  sample from HF token_count pool (same split)
    AI-R  →  jitter the HR mean within ±10 % (same split), to mimic
             'rewrite stays close to original length'

    If a split is missing HF or HR rows, falls back to the other splits' pool.
    """
    rng = np.random.default_rng(seed)
    df = df.copy()
    df.columns = df.columns.str.lower()

    # Tokenize if not already present
    if "token_count" not in df.columns:
        print("  Tokenizing … (this may take a moment)")
        df["token_count"] = df["article"].apply(token_count)

    # Normalise news_type casing
    df["news_type"] = df["news_type"].str.upper().str.strip()

    for split in df["split"].unique():
        mask_split = df["split"] == split

        # ── HF pool for this split (Fix 1 source) ───────────────────────────
        hf_pool = df[mask_split & (df["news_type"] == "HF")]["token_count"].tolist()
        if not hf_pool:
            # fallback: all HF rows
            hf_pool = df[df["news_type"] == "HF"]["token_count"].tolist()

        # ── HR pool mean for Fix 2 ───────────────────────────────────────────
        hr_vals = df[mask_split & (df["news_type"] == "HR")]["token_count"].tolist()
        if not hr_vals:
            hr_vals = df[df["news_type"] == "HR"]["token_count"].tolist()
        hr_mean = int(np.mean(hr_vals)) if hr_vals else 400

        # ── Apply Fix 1: AI-F ────────────────────────────────────────────────
        aif_mask = mask_split & (df["news_type"] == "AI-F")
        n_aif = aif_mask.sum()
        if n_aif > 0 and hf_pool:
            simulated = sample_from_pool(hf_pool, n_aif, rng)
            df.loc[aif_mask, "token_count"] = simulated

        # ── Apply Fix 2: AI-R ────────────────────────────────────────────────
        air_mask = mask_split & (df["news_type"] == "AI-R")
        n_air = air_mask.sum()
        if n_air > 0:
            simulated = [jitter_within_pct(hr_mean, 0.10, rng) for _ in range(n_air)]
            df.loc[air_mask, "token_count"] = simulated

    return df


# ── Console summary ───────────────────────────────────────────────────────────

def print_comparison(before: pd.DataFrame, after: pd.DataFrame, sheet: str) -> None:
    print(f"\n{'═' * 72}")
    print(f"  {sheet}")
    print(f"{'═' * 72}")
    header = f"  {'Split':<8} {'Type':<8} {'Before mean':>12} {'Before std':>11} "
    header += f"{'After mean':>11} {'After std':>10}"
    print(header)
    print("  " + "─" * 68)

    for split in sorted(before["split"].unique()):
        for nt in ["HR", "AI-R", "HF", "AI-F"]:
            b = before[(before["split"] == split) & (before["news_type"] == nt)]["token_count"]
            a = after [(after ["split"] == split) & (after ["news_type"] == nt)]["token_count"]
            if b.empty and a.empty:
                continue
            bm = f"{b.mean():.1f}" if not b.empty else "—"
            bs = f"{b.std():.1f}"  if not b.empty else "—"
            am = f"{a.mean():.1f}" if not a.empty else "—"
            as_ = f"{a.std():.1f}" if not a.empty else "—"
            changed = " ◄" if (not b.empty and not a.empty and abs(b.mean() - a.mean()) > 5) else ""
            print(f"  {split:<8} {nt:<8} {bm:>12} {bs:>11} {am:>11} {as_:>10}{changed}")
    print()


# ── Excel output ──────────────────────────────────────────────────────────────

NT_FILLS = {
    "HR":   "C6EFCE",
    "AI-R": "FFEB9C",
    "HF":   "FFC7CE",
    "AI-F": "E2AFFF",
}
HDR_FILL = "1F4E79"


def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=HDR_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _cell(ws, row, col, value, fill_hex=None, bold=False, fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fill_hex:
        c.fill = PatternFill("solid", start_color=fill_hex)
    if fmt:
        c.number_format = fmt


def write_comparison_sheet(ws, records: list[dict]) -> None:
    ws.title = "Length Comparison"

    headers = [
        "Sheet", "Split", "News Type",
        "Before Mean", "Before Std", "Before Min", "Before Max",
        "After Mean",  "After Std",  "After Min",  "After Max",
        "Δ Mean",
    ]
    col_widths = [36, 10, 12, 13, 12, 12, 12, 12, 12, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 28

    ri = 2
    for rec in records:
        sheet = rec["sheet"]
        for split in ["train", "val", "test"]:
            for nt in ["HR", "AI-R", "HF", "AI-F"]:
                key = (split, nt)
                if key not in rec:
                    continue
                b = rec[key]["before"]
                a = rec[key]["after"]
                fill = NT_FILLS.get(nt, "FFFFFF")
                delta = a["mean"] - b["mean"]
                _cell(ws, ri, 1,  sheet,       bold=True, align="left")
                _cell(ws, ri, 2,  split)
                _cell(ws, ri, 3,  nt,          fill_hex=fill)
                _cell(ws, ri, 4,  round(b["mean"], 1), fill_hex=fill, fmt="0.0")
                _cell(ws, ri, 5,  round(b["std"],  1), fill_hex=fill, fmt="0.0")
                _cell(ws, ri, 6,  b["min"],            fill_hex=fill)
                _cell(ws, ri, 7,  b["max"],            fill_hex=fill)
                _cell(ws, ri, 8,  round(a["mean"], 1), fill_hex=fill, fmt="0.0")
                _cell(ws, ri, 9,  round(a["std"],  1), fill_hex=fill, fmt="0.0")
                _cell(ws, ri, 10, a["min"],            fill_hex=fill)
                _cell(ws, ri, 11, a["max"],            fill_hex=fill)
                delta_fill = "C6EFCE" if abs(delta) < 20 else "FFC7CE"
                _cell(ws, ri, 12, round(delta, 1), fill_hex=delta_fill, fmt="+0.0;-0.0;0.0")
                ri += 1

    ws.freeze_panes = "A2"


def write_data_sheet(ws, df: pd.DataFrame, title: str) -> None:
    ws.title = title
    cols = [c for c in ["article", "label", "topic", "split", "news_type", "token_count"]
            if c in df.columns]
    for ci, h in enumerate(cols, 1):
        _hdr(ws, 1, ci, h)
    ws.column_dimensions["A"].width = 80
    for ci in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    for ri, row in enumerate(df[cols].itertuples(index=False), 2):
        nt = str(getattr(row, "news_type", "")).upper()
        fill = NT_FILLS.get(nt)
        for ci, val in enumerate(row, 1):
            _cell(ws, ri, ci, val, fill_hex=(fill if ci > 1 else None),
                  align=("left" if ci == 1 else "center"))

    ws.freeze_panes = "B2"


# ── Main ──────────────────────────────────────────────────────────────────────

def resolve_workbook() -> str:
    if len(sys.argv) > 1:
        return sys.argv[1]
    return input("Path to stratified Excel workbook: ").strip()


def pick_sheets(xl: pd.ExcelFile) -> list[str]:
    data_sheets = [s for s in xl.sheet_names if s.lower() != "summary"]
    print("\nAvailable sheets:")
    for i, name in enumerate(data_sheets, 1):
        print(f"  {i:>3}.  {name}")
    print("\nEnter sheet number(s) separated by spaces, or 'all': ", end="")
    raw = input().strip().lower()
    if raw == "all":
        return data_sheets
    chosen = []
    for tok in raw.split():
        try:
            idx = int(tok) - 1
            if not (0 <= idx < len(data_sheets)):
                raise ValueError
            chosen.append(data_sheets[idx])
        except ValueError:
            print(f"  [WARN] '{tok}' skipped — invalid number.")
    if not chosen:
        raise ValueError("No valid sheets selected.")
    return chosen


def stats(series: pd.Series) -> dict:
    return {
        "mean": float(series.mean()),
        "std":  float(series.std()),
        "min":  int(series.min()),
        "max":  int(series.max()),
    }


def main() -> None:
    wb_path  = resolve_workbook()
    xl       = pd.ExcelFile(wb_path)
    selected = pick_sheets(xl)

    print(f"\n  Simulating {len(selected)} sheet(s) …")

    out_wb      = Workbook()
    out_wb.remove(out_wb.active)
    comp_ws     = out_wb.create_sheet("Length Comparison")
    comp_records: list[dict] = []

    for sheet_name in selected:
        df_orig = xl.parse(sheet_name)
        df_orig.columns = df_orig.columns.str.lower()

        # Ensure news_type exists
        if "news_type" not in df_orig.columns:
            print(f"  [SKIP] '{sheet_name}' has no news_type column.")
            continue

        df_orig["news_type"] = df_orig["news_type"].str.upper().str.strip()

        # Tokenize original
        if "token_count" not in df_orig.columns:
            print(f"  Tokenizing '{sheet_name}' …")
            df_orig["token_count"] = df_orig["article"].apply(token_count)

        df_sim = simulate_sheet(df_orig)

        # Console comparison
        print_comparison(df_orig, df_sim, sheet_name)

        # Build comparison records
        rec = {"sheet": sheet_name}
        for split in df_orig["split"].unique():
            for nt in ["HR", "AI-R", "HF", "AI-F"]:
                b_ser = df_orig[(df_orig["split"] == split) & (df_orig["news_type"] == nt)]["token_count"]
                a_ser = df_sim [(df_sim ["split"] == split) & (df_sim ["news_type"] == nt)]["token_count"]
                if b_ser.empty:
                    continue
                rec[(str(split).lower(), nt)] = {
                    "before": stats(b_ser),
                    "after":  stats(a_ser),
                }
        comp_records.append(rec)

        # Write data sheet (simulated)
        sim_ws = out_wb.create_sheet(sheet_name[:28] + " (sim)")
        write_data_sheet(sim_ws, df_sim, sim_ws.title)

    write_comparison_sheet(comp_ws, comp_records)

    out_path = pathlib.Path("simulated_lengths.xlsx")
    out_wb.save(str(out_path))
    print(f"\n  Saved → {out_path}")
    print("    • 'Length Comparison'  — before/after stats per sheet × split × news_type")
    print("    • '* (sim)' sheets     — full dataset with simulated token_count values")
    print("\n  Next: run  python linear_preview.py simulated_lengths.xlsx")


if __name__ == "__main__":
    main()
