"""
linear_preview.py
-----------------
A stripped-down length-only logistic regression classifier designed to run
on the output of length_sim.py (simulated_lengths.xlsx).

Differences from the full length_classifier.py
-----------------------------------------------
• Reads token_count directly from the sheet — no re-tokenization needed.
• Reports accuracy, MCC, and per-class F1 for train / val / test.
• Prints a compact side-by-side Before vs After summary when given two files.
• Outputs linear_preview_results.xlsx with the same sheet layout.

Usage
-----
    # Single file (simulated only)
    python linear_preview.py simulated_lengths.xlsx

    # Before-vs-after comparison (original first, simulated second)
    python linear_preview.py stratified_dataset.xlsx simulated_lengths.xlsx
"""

import sys
import pathlib

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    matthews_corrcoef,
    precision_recall_fscore_support,
    classification_report,
)


# ── Colour palette (matches length_classifier_results.xlsx) ──────────────────
HDR_FILL    = "1F4E79"
GREEN_FILL  = "C6EFCE"
BLUE_FILL   = "BDD7EE"
YELLOW_FILL = "FFEB9C"
RED_FILL    = "FFC7CE"
GREY_FILL   = "D9D9D9"
ALT_ROW     = "F2F2F2"
ORANGE_FILL = "FCE4D6"


def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=HDR_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _cell(ws, row, col, value, fill_hex=None, bold=False, fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fill_hex:
        c.fill = PatternFill("solid", start_color=fill_hex)
    if fmt:
        c.number_format = fmt


def _pct(ws, row, col, value, fill_hex=None):
    _cell(ws, row, col, value, fill_hex=fill_hex, fmt="0.00%")


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Classifier ────────────────────────────────────────────────────────────────

def evaluate(clf, part: pd.DataFrame) -> dict:
    y_true = part["label"]
    y_pred = clf.predict(part[["token_count"]])
    acc    = accuracy_score(y_true, y_pred)
    mcc    = matthews_corrcoef(y_true, y_pred)
    prec, rec, f1, sup = precision_recall_fscore_support(
        y_true, y_pred, labels=[0, 1], zero_division=0
    )
    return {
        "n":          len(part),
        "accuracy":   acc,
        "mcc":        mcc,
        "fake_prec":  float(prec[0]), "fake_rec": float(rec[0]),
        "fake_f1":    float(f1[0]),   "fake_sup": int(sup[0]),
        "real_prec":  float(prec[1]), "real_rec": float(rec[1]),
        "real_f1":    float(f1[1]),   "real_sup": int(sup[1]),
        "report_str": classification_report(
                          y_true, y_pred,
                          target_names=["Fake (0)", "Real (1)"],
                          zero_division=0,
                      ),
    }


def run_sheet(sheet_name: str, df: pd.DataFrame) -> dict:
    df = df.copy()
    df.columns = df.columns.str.lower()

    required = {"token_count", "label", "split"}
    missing  = required - set(df.columns)
    if missing:
        raise ValueError(f"Sheet '{sheet_name}' missing: {missing}")

    train_df = df[df["split"] == "train"].reset_index(drop=True)
    val_df   = df[df["split"] == "val"].reset_index(drop=True)
    test_df  = df[df["split"] == "test"].reset_index(drop=True)

    clf = LogisticRegression(random_state=42, max_iter=1000)
    clf.fit(train_df[["token_count"]], train_df["label"])

    train_m = evaluate(clf, train_df)
    val_m   = evaluate(clf, val_df)
    test_m  = evaluate(clf, test_df)

    # Token stats per news_type (if available)
    token_stats: dict[str, dict] = {}
    if "news_type" in df.columns:
        df["news_type"] = df["news_type"].str.upper().str.strip()
        for split_label, part in [("train", train_df), ("val", val_df), ("test", test_df)]:
            token_stats[split_label] = {}
            for nt, grp in part.groupby("news_type"):
                s = grp["token_count"]
                token_stats[split_label][nt] = {
                    "n":      len(s),
                    "mean":   float(s.mean()),
                    "std":    float(s.std()),
                    "min":    int(s.min()),
                    "max":    int(s.max()),
                    "median": float(s.median()),
                }

    return {
        "sheet":       sheet_name,
        "coef":        float(clf.coef_[0][0]),
        "intercept":   float(clf.intercept_[0]),
        "train":       train_m,
        "val":         val_m,
        "test":        test_m,
        "token_stats": token_stats,
    }


# ── Console output ────────────────────────────────────────────────────────────

def print_results(records: list[dict], label: str) -> None:
    print(f"\n{'═' * 80}")
    print(f"  {label}")
    print(f"{'═' * 80}")
    header = f"  {'Sheet':<42} {'Train Acc':>10} {'Test Acc':>10} {'Test MCC':>10}"
    print(header)
    print("  " + "─" * 74)
    accs = []
    mccs = []
    for r in records:
        ta  = r["train"]["accuracy"]
        tea = r["test"]["accuracy"]
        mcc = r["test"]["mcc"]
        accs.append(tea)
        mccs.append(mcc)
        print(f"  {r['sheet']:<42} {ta:>10.4f} {tea:>10.4f} {mcc:>10.4f}")
    print("  " + "─" * 74)
    print(f"  {'MEAN':<42} {'':>10} {np.mean(accs):>10.4f} {np.mean(mccs):>10.4f}")
    print(f"  {'MAX':<42} {'':>10} {np.max(accs):>10.4f} {np.max(mccs):>10.4f}")
    print(f"  {'MIN':<42} {'':>10} {np.min(accs):>10.4f} {np.min(mccs):>10.4f}")


def print_before_after(before: list[dict], after: list[dict]) -> None:
    print(f"\n{'═' * 90}")
    print("  BEFORE vs AFTER LENGTH FIX")
    print(f"{'═' * 90}")
    hdr = (f"  {'Sheet':<38} "
           f"{'B.TestAcc':>10} {'A.TestAcc':>10} "
           f"{'B.MCC':>9} {'A.MCC':>9} "
           f"{'ΔMCC':>7}")
    print(hdr)
    print("  " + "─" * 85)

    b_map = {r["sheet"]: r for r in before}
    a_map = {r["sheet"]: r for r in after}

    for sheet in b_map:
        # Match simulated sheet name (may have ' (sim)' suffix)
        sim_key = next((k for k in a_map if sheet in k), None)
        if sim_key is None:
            continue
        b = b_map[sheet]
        a = a_map[sim_key]
        delta_mcc = a["test"]["mcc"] - b["test"]["mcc"]
        arrow = "▲" if delta_mcc > 0.01 else ("▼" if delta_mcc < -0.01 else "~")
        print(f"  {sheet:<38} "
              f"{b['test']['accuracy']:>10.4f} {a['test']['accuracy']:>10.4f} "
              f"{b['test']['mcc']:>9.4f} {a['test']['mcc']:>9.4f} "
              f"{delta_mcc:>+7.4f} {arrow}")

    print()
    print("  MCC interpretation: 0 = chance, 1 = perfect")
    print("  Goal after fix: MCC should drop toward 0 (length becomes less predictive)")


# ── Excel writers ─────────────────────────────────────────────────────────────

def write_results_sheet(ws, records: list[dict], title: str, tag: str = "") -> None:
    ws.title = (title + (" " + tag if tag else ""))[:31]

    headers = [
        "Sheet", "Train n", "Train Acc", "Train MCC",
        "Test n",  "Test Acc",  "Test MCC",
        "Fake F1", "Real F1",
        "Coefficient", "Intercept",
    ]
    widths = [38, 10, 11, 11, 10, 11, 11, 10, 10, 14, 12]
    _set_widths(ws, widths)

    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 30

    for ri, r in enumerate(records, 2):
        fill = ALT_ROW if ri % 2 == 0 else None
        _cell(ws, ri, 1,  r["sheet"],                 bold=True, align="left")
        _cell(ws, ri, 2,  r["train"]["n"],             fill_hex=fill)
        _pct( ws, ri, 3,  r["train"]["accuracy"],      fill_hex=BLUE_FILL)
        _cell(ws, ri, 4,  round(r["train"]["mcc"],4),  fill_hex=BLUE_FILL, fmt="0.0000")
        _cell(ws, ri, 5,  r["test"]["n"],              fill_hex=fill)
        _pct( ws, ri, 6,  r["test"]["accuracy"],       fill_hex=YELLOW_FILL)
        _cell(ws, ri, 7,  round(r["test"]["mcc"], 4),  fill_hex=YELLOW_FILL, fmt="0.0000")
        _pct( ws, ri, 8,  r["test"]["fake_f1"],        fill_hex=RED_FILL)
        _pct( ws, ri, 9,  r["test"]["real_f1"],        fill_hex=GREEN_FILL)
        _cell(ws, ri, 10, round(r["coef"],      6),    fill_hex=fill, fmt="0.000000")
        _cell(ws, ri, 11, round(r["intercept"], 6),    fill_hex=fill, fmt="0.000000")

    # Summary stats
    last = len(records) + 3
    accs = [r["test"]["accuracy"] for r in records]
    mccs = [r["test"]["mcc"]      for r in records]
    _cell(ws, last,     1, f"Mean test accuracy: {np.mean(accs):.4f}",
          bold=True, align="left")
    _cell(ws, last + 1, 1, f"Mean test MCC:      {np.mean(mccs):.4f}",
          bold=True, align="left")

    # Interpretation banner
    low_mcc  = all(abs(m) < 0.15 for m in mccs)
    low_acc  = all(a < 0.60 for a in accs)
    last += 3
    if low_mcc or low_acc:
        interp = [
            "LENGTH BIAS ASSESSMENT",
            "Length is NOT a reliable predictor after the fix — good!",
            "BERT will need to learn linguistic features, not article length.",
        ]
        color = "375623"
        bg    = "C6EFCE"
    else:
        interp = [
            "LENGTH BIAS ASSESSMENT",
            "Length is STILL predictive — consider stricter length control in generation prompts.",
            "Check the Token Stats sheet for which news_type group is driving the signal.",
        ]
        color = "9C0006"
        bg    = "FFC7CE"

    for i, line in enumerate(interp):
        c = ws.cell(row=last + i, column=1, value=line)
        c.font = Font(name="Arial", bold=True, size=10, color=color)
        c.fill = PatternFill("solid", start_color=bg)

    ws.freeze_panes = "A2"


def write_comparison_sheet(ws, before: list[dict], after: list[dict]) -> None:
    ws.title = "Before vs After"
    headers = [
        "Sheet",
        "Before Test Acc", "After Test Acc",
        "Before MCC",      "After MCC",
        "Δ MCC", "Δ Acc",
    ]
    _set_widths(ws, [38, 17, 16, 14, 13, 10, 10])
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 28

    b_map = {r["sheet"]: r for r in before}
    a_map = {r["sheet"]: r for r in after}

    ri = 2
    for sheet in b_map:
        sim_key = next((k for k in a_map if sheet in k), None)
        if sim_key is None:
            continue
        b = b_map[sheet]
        a = a_map[sim_key]
        d_mcc = a["test"]["mcc"]      - b["test"]["mcc"]
        d_acc = a["test"]["accuracy"] - b["test"]["accuracy"]
        fill  = ALT_ROW if ri % 2 == 0 else None
        mcc_fill = GREEN_FILL if d_mcc < -0.05 else (RED_FILL if d_mcc > 0.05 else GREY_FILL)
        _cell(ws, ri, 1, sheet,                       bold=True, align="left")
        _pct( ws, ri, 2, b["test"]["accuracy"],        fill_hex=fill)
        _pct( ws, ri, 3, a["test"]["accuracy"],        fill_hex=fill)
        _cell(ws, ri, 4, round(b["test"]["mcc"], 4),   fill_hex=fill, fmt="0.0000")
        _cell(ws, ri, 5, round(a["test"]["mcc"], 4),   fill_hex=fill, fmt="0.0000")
        _cell(ws, ri, 6, round(d_mcc, 4), fill_hex=mcc_fill, fmt="+0.0000;-0.0000;0.0000")
        _cell(ws, ri, 7, round(d_acc, 4), fill_hex=fill,     fmt="+0.0000;-0.0000;0.0000")
        ri += 1

    # Legend
    ws.cell(row=ri + 2, column=1,
            value="Δ MCC < 0 (green) = length became LESS predictive after fix — desired outcome").font = \
        Font(name="Arial", italic=True, size=9, color="375623")

    ws.freeze_panes = "A2"


def write_token_stats_sheet(ws, records: list[dict], tag: str = "") -> None:
    ws.title = ("Token Stats " + tag)[:31]
    headers = ["Sheet", "Split", "News Type", "n", "Mean", "Std", "Min", "Max", "Median"]
    _set_widths(ws, [38, 10, 12, 8, 10, 10, 8, 8, 10])
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    ws.row_dimensions[1].height = 28

    NT_FILLS = {"HR": "C6EFCE", "AI-R": "FFEB9C", "HF": "FFC7CE", "AI-F": "E2AFFF"}
    SPLIT_FILLS = {"train": GREEN_FILL, "test": BLUE_FILL, "val": YELLOW_FILL}
    NUM_FMT = "0.00"

    ri = 2
    for r in records:
        ts = r.get("token_stats", {})
        for split in ["train", "val", "test"]:
            if split not in ts:
                continue
            first = True
            for nt in ["HR", "AI-R", "HF", "AI-F"]:
                if nt not in ts[split]:
                    continue
                s = ts[split][nt]
                nt_fill = NT_FILLS.get(nt, ALT_ROW)
                sp_fill = SPLIT_FILLS.get(split)
                _cell(ws, ri, 1, r["sheet"] if first else "", bold=first, align="left")
                _cell(ws, ri, 2, split if first else "", fill_hex=sp_fill if first else None)
                _cell(ws, ri, 3, nt,           fill_hex=nt_fill)
                _cell(ws, ri, 4, s["n"],        fill_hex=nt_fill)
                _cell(ws, ri, 5, s["mean"],     fill_hex=nt_fill, fmt=NUM_FMT)
                _cell(ws, ri, 6, s["std"],      fill_hex=nt_fill, fmt=NUM_FMT)
                _cell(ws, ri, 7, s["min"],      fill_hex=nt_fill)
                _cell(ws, ri, 8, s["max"],      fill_hex=nt_fill)
                _cell(ws, ri, 9, s["median"],   fill_hex=nt_fill, fmt=NUM_FMT)
                ri += 1
                first = False

    ws.freeze_panes = "A2"


# ── Workbook / sheet plumbing ─────────────────────────────────────────────────

def resolve_workbook(prompt: str) -> str:
    return input(prompt).strip()


def pick_sheets(xl: pd.ExcelFile, label: str) -> list[str]:
    data_sheets = [s for s in xl.sheet_names if s.lower() not in ("summary", "length comparison")]
    print(f"\nSheets in '{label}':")
    for i, name in enumerate(data_sheets, 1):
        print(f"  {i:>3}.  {name}")
    print("Enter sheet number(s) or 'all': ", end="")
    raw = input().strip().lower()
    if raw == "all":
        return data_sheets
    chosen = []
    for tok in raw.split():
        try:
            idx = int(tok) - 1
            if not (0 <= idx < len(data_sheets)):
                raise ValueError
            chosen.append(data_sheets[idx])
        except ValueError:
            print(f"  [WARN] '{tok}' skipped.")
    if not chosen:
        raise ValueError("No valid sheets selected.")
    return chosen


def load_and_run(path: str, label: str) -> list[dict]:
    xl       = pd.ExcelFile(path)
    selected = pick_sheets(xl, label)
    results  = []
    for sheet_name in selected:
        df = xl.parse(sheet_name)
        print(f"  Running '{sheet_name}' …")
        results.append(run_sheet(sheet_name, df))
    return results


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    paths = sys.argv[1:]

    if len(paths) == 0:
        path = input("Path to workbook (simulated or original): ").strip()
        paths = [path]

    if len(paths) == 1:
        # Single-file mode
        print(f"\n  Single-file mode: {paths[0]}")
        results = load_and_run(paths[0], paths[0])
        print_results(results, paths[0])

        wb = Workbook()
        wb.remove(wb.active)
        write_results_sheet(wb.create_sheet("Results"), results, "Results")
        write_token_stats_sheet(wb.create_sheet("Token Stats"), results)

    else:
        # Before-vs-after mode
        print(f"\n  Before file : {paths[0]}")
        before = load_and_run(paths[0], "BEFORE")
        print(f"\n  After file  : {paths[1]}")
        after  = load_and_run(paths[1], "AFTER (simulated)")

        print_results(before, "BEFORE (original lengths)")
        print_results(after,  "AFTER  (simulated lengths)")
        print_before_after(before, after)

        wb = Workbook()
        wb.remove(wb.active)
        write_results_sheet(wb.create_sheet("Before"), before, "Before", "Before")
        write_results_sheet(wb.create_sheet("After"),  after,  "After",  "After")
        write_comparison_sheet(wb.create_sheet("Before vs After"), before, after)
        write_token_stats_sheet(wb.create_sheet("Token Stats (Before)"), before, "Before")
        write_token_stats_sheet(wb.create_sheet("Token Stats (After)"),  after,  "After")

    out_path = pathlib.Path("linear_preview_results.xlsx")
    wb.save(str(out_path))
    print(f"\n  Saved → {out_path}")


if __name__ == "__main__":
    main()
