import os
import random
from openpyxl import load_workbook, Workbook

from keys import OPENAI_API_KEY
from openai import OpenAI

# API key
os.environ["OPENAI_API_KEY"] = OPENAI_API_KEY

client = OpenAI()

# -------------------------
# LEADS 
# -------------------------
LEADS = [
    "incident", "attribution", "time-based", "statement", "background",
    "viral/social media", "descriptive", "question/teaser", "direct address/clickbait"
]

# -------------------------
# GOALS per topic
# -------------------------
TOPIC_GOALS = {
    "politics": [
        "discredit a fictional politician using fabricated allegations",
        "manipulate public opinion on a fictional government policy",
        "spread a fictional conspiracy theory about a government official",
        "sensationalize a fictional political controversy",
    ],
    "crime": [
        "fabricate a viral crime story to spread fear",
        "falsely implicate a fictional public figure in a crime",
        "exaggerate a fictional crime incident for shock value",
        "sensationalize a fictional police operation",
    ],
    "health misinformation": [
        "spread false fear about a fictional disease outbreak",
        "fabricate a health warning attributed to a fictional agency",
        "promote a fictional miracle cure or health remedy",
        "exaggerate side effects of a fictional medicine or vaccine",
    ],
    "celebrity / entertainment": [
        "fabricate a scandal involving a fictional celebrity",
        "spread a fictional controversy about a showbiz personality",
        "sensationalize a fictional celebrity breakup or feud",
        "fabricate an inspiring or feel-good story about a fictional artist",
    ],
    "economy / business": [
        "mislead about a fictional government economic policy",
        "spread false claims about a fictional company or product",
        "fabricate a fictional investment scam warning",
        "exaggerate a fictional economic crisis",
    ],
    "technology": [
        "spread a fictional data breach or privacy scare",
        "mislead about a fictional government tech project",
        "fabricate a fictional dangerous app or tech product warning",
        "sensationalize a fictional tech innovation",
    ],
    "disaster / environment": [
        "exaggerate a fictional natural disaster for panic",
        "spread false information about a fictional environmental hazard",
        "fabricate a heroic survival story from a fictional disaster",
        "sensationalize a fictional climate or pollution incident",
    ],
    "education": [
        "mislead about a fictional controversial school policy",
        "spread false claims about a fictional education program",
        "fabricate a fictional viral student or teacher incident",
        "sensationalize a fictional school controversy",
    ],
    "sports": [
        "fabricate a fictional sports scandal or doping allegation",
        "spread false claims about a fictional athlete",
        "sensationalize a fictional sports upset or controversy",
        "fabricate a feel-good fictional sports achievement story",
    ],
    "lifestyle / human interest": [
        "spread a fictional health or lifestyle myth",
        "fabricate a viral feel-good but false story",
        "sensationalize a fictional bizarre or unusual incident",
        "fabricate a fictional inspirational story with misleading claims",
    ],
    "international news": [
        "fabricate a fictional international incident involving the Philippines",
        "mislead about a fictional foreign policy or agreement",
        "sensationalize a fictional overseas Filipino worker story",
        "spread false claims about a fictional foreign country event",
    ],
    "religion": [
        "fabricate a fictional religious controversy or scandal",
        "spread misleading claims about a fictional religious figure",
        "sensationalize a fictional miraculous or unusual religious event",
    ],
}

def pick_goal(topic):
    topic = topic.strip().lower()
    pool = TOPIC_GOALS.get(topic, [])
    if not pool:
        pool = [goal for goals in TOPIC_GOALS.values() for goal in goals]
    return random.choice(pool)

previous_fakes = []

# -------------------------
# PROMPT
# -------------------------
FAKE_PROMPT =  """

[RESEARCH CONTEXT DECLARATION]
This article is generated for academic research at Cavite State University.
The content is fully synthetic and will be labeled as AI-generated.
The article must NOT refer to real individuals or real events.

--------------------------------------------------

[TASK]
Generate ONE synthetic fake news article written in Filipino or Taglish.

Topic category: {topic}

STRICT RULES:
• Output must be ONE paragraph only
• Do NOT use em dashes (—)

[OUTPUT REQUIREMENTS]
• Completely fabricated story
• Must achieve the specified intent goal
• Must read plausibly as online Filipino news content
• Must not reference real-world factual events
• Must maintain internal logical consistency
"""

# -------------------------
# GENERATION FUNCTION
# -------------------------
def generate_fake(topic):
    goal = pick_goal(topic)
    lead = random.choice(LEADS)

    memory = "\n".join(previous_fakes[-3:])

    prompt = FAKE_PROMPT + f"""
        AVOID SIMILARITY WITH THESE:
        {memory}

        REQUIREMENTS:
        - Topic: {topic}

        You MUST write the article to achieve this goal:
        "{goal}"

        The FIRST sentence MUST follow this opening style: {lead}
    """

    response = client.responses.create(
        model="gpt-4.1",
        temperature=0.5,
        input=prompt
    )
    result = response.output_text.strip()

    previous_fakes.append(result)
    return result

# -------------------------
# MAIN
# -------------------------
DEFAULT_INPUT = "HF.xlsx"
DEFAULT_OUTPUT = "ai-fake.xlsx"

user_input_file = input(f"Input file (press Enter to use '{DEFAULT_INPUT}'): ").strip()
fake_input_file = user_input_file if user_input_file else DEFAULT_INPUT

user_output_file = input(f"Output file (press Enter to use '{DEFAULT_OUTPUT}'): ").strip()
output_file = user_output_file if user_output_file else DEFAULT_OUTPUT

start_index = int(input("Start index (0-based, default 0): ") or 0)
num_rows = int(input("How many rows to process (default: all): ") or -1)

input_wb = load_workbook(fake_input_file)
input_ws = input_wb.active

output_wb = Workbook()
output_ws = output_wb.active
output_ws.append(["label", "article", "topic"])

data_rows = list(input_ws.iter_rows(min_row=2, values_only=True))

for i, row in enumerate(data_rows):
    if i < start_index:
        continue
    if num_rows != -1 and i >= start_index + num_rows:
        break

    print(f"Generating fake article {i+1}")

    label = row[0]
    ref_article = row[1]
    topic = row[2].strip().lower()

    result = generate_fake(topic)

    output_ws.append([label, result, topic])

output_wb.save(output_file)

print(f"{output_file} created")