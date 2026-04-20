"""
clean.py
--------
Cleans the 'article' column of an Excel or CSV file containing news data.

Cleaning pipeline — each step is backed by the reviewed literature:

  1. HTML entity decoding
       ↳ Standard first pass before any other parsing.

  2. Script and stylesheet removal
       ↳ Reyes et al.: "cleaned HTML content by removing JavaScript, style
         sheets, and non-article elements" [50†L128-L132].

  3. HTML tag stripping
       ↳ Multiple studies call for removing all HTML markup via regex or
         BeautifulSoup [47†L354-L359] [54†L179-L184].

  4. URL removal
       ↳ Regex-based removal of HTTP/S links is cited explicitly as a
         standard cleaning step [54†L179-L184] [27†L280-L282].

  5. Email and IP address removal
       ↳ Cited alongside URL removal as common regex targets [27†L280-L282].

  6. Emoji removal
       ↳ One study explicitly lists emojis among elements to remove
         ("URLs, HTML tags, emojis, special characters") [18†L668-L670].

  7. Author byline / photo credit removal
       ↳ One pipeline "removed bylines with author names, photographer
         names etc." to avoid data leakage [59†L300-L307].

  8. Unicode whitespace normalization
       ↳ Implied by the universal text-normalization step described across
         all reviewed pipelines.

  9. Consecutive whitespace / newline collapsing
       ↳ Part of the "standardizing the input" consensus [52+].

 10. Short-row filtering (< 10 words)
       ↳ Removing rows with fewer than 10 words as "poorly informative"
         [24†L238-L241].

Steps deliberately EXCLUDED (not supported by the reviewed studies):
  - Stopword removal  — studies recommend keeping stopwords for BERT
                        [19†L682-L687].
  - Stemming          — skipped in BERT pipelines; subword tokenizer handles
                        morphology [19†L682-L687].
  - Dataset-specific Filipino boilerplate (Panoorin, Basahin, Source:, etc.)
                      — no citation in the reviewed literature; out of scope
                        for a general preprocessing pipeline.

Usage:
    python clean.py <input.xlsx|csv> <output.xlsx>
    python clean.py <input.xlsx|csv>               # writes <input>_cleaned.xlsx
"""

import re
import sys
import html
import pathlib
import unicodedata

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── compiled regex patterns ───────────────────────────────────────────────────

# 2. Script and stylesheet blocks (with content between tags)
_RE_SCRIPT = re.compile(r"<script[^>]*>.*?</script>", re.IGNORECASE | re.DOTALL)
_RE_STYLE  = re.compile(r"<style[^>]*>.*?</style>",  re.IGNORECASE | re.DOTALL)

# 3. All remaining HTML / XML tags
#    [47†L354-L359] [54†L179-L184]
_RE_HTML_TAGS = re.compile(r"<[^>]+>")

# 4. URLs  [54†L179-L184] [27†L280-L282]
_RE_URL = re.compile(r"https?://\S+|www\.\S+", re.IGNORECASE)

# 5. Email addresses and bare IP addresses  [27†L280-L282]
_RE_EMAIL = re.compile(r"\S+@\S+\.\S+")
_RE_IP    = re.compile(r"\b\d{1,3}(?:\.\d{1,3}){3}\b")

# 6. Emoji — matches all Unicode emoji / pictograph ranges  [18†L668-L670]
_RE_EMOJI = re.compile(
    "["
    "\U0001F600-\U0001F64F"   # emoticons
    "\U0001F300-\U0001F5FF"   # symbols & pictographs
    "\U0001F680-\U0001F6FF"   # transport & map
    "\U0001F1E0-\U0001F1FF"   # flags
    "\U00002700-\U000027BF"   # dingbats
    "\U0001F900-\U0001F9FF"   # supplemental symbols
    "\U00002600-\U000026FF"   # miscellaneous symbols
    "\U0001FA00-\U0001FA6F"   # chess, etc.
    "\U0001FA70-\U0001FAFF"   # food, objects
    "]+",
    flags=re.UNICODE,
)

# 7. Author bylines and photo credits  [59†L300-L307]
#    Matches common patterns: "By John Doe", "Photo by Jane Smith",
#    "Reporter: Name", "Written by Name"
_RE_BYLINE = re.compile(
    r"(?:(?:Photo|Image|Video|Written|Reported|Story)?\s*[Bb]y\s+[A-Z][a-zA-Z\s\-\.]{2,40})"
    r"|(?:(?:Reporter|Photographer|Author|Writer)\s*:\s*[A-Z][a-zA-Z\s\-\.]{2,40})",
    re.MULTILINE,
)

# 8. Unicode whitespace variants → plain space
_UNICODE_SPACES = str.maketrans({
    "\u00a0": " ",   # non-breaking space
    "\u200b": "",    # zero-width space (drop entirely)
    "\u200c": "",    # zero-width non-joiner
    "\u200d": "",    # zero-width joiner
    "\u2009": " ",   # thin space
    "\u202f": " ",   # narrow no-break space
    "\u3000": " ",   # ideographic space
    "\ufeff": "",    # BOM
})

# 9. Whitespace collapsing helpers
_RE_MULTI_SPACE = re.compile(r"[ \t]+")
_RE_MULTI_NL    = re.compile(r"\n{3,}")


# ── core cleaning function ────────────────────────────────────────────────────

def clean_article(text: str) -> str:
    """
    Apply the study-backed preprocessing pipeline to a single article string.
    Returns the cleaned string, or the original value if it is not a string.
    """
    if not isinstance(text, str):
        return text

    # 1. HTML entity decoding  (e.g. &amp; → &, &nbsp; → space)
    text = html.unescape(text)

    # 2. Remove <script> and <style> blocks (content + tags)
    text = _RE_SCRIPT.sub(" ", text)
    text = _RE_STYLE.sub(" ", text)

    # 3. Strip all remaining HTML tags
    text = _RE_HTML_TAGS.sub(" ", text)

    # 4. Remove URLs
    text = _RE_URL.sub(" ", text)

    # 5. Remove email addresses and IP addresses
    text = _RE_EMAIL.sub(" ", text)
    text = _RE_IP.sub(" ", text)

    # 6. Remove emoji characters
    text = _RE_EMOJI.sub(" ", text)

    # 7. Remove author bylines and photo credits
    text = _RE_BYLINE.sub(" ", text)

    # 8. Normalize Unicode whitespace variants
    text = text.translate(_UNICODE_SPACES)

    # Normalize Unicode to NFC form (combines accented characters properly)
    text = unicodedata.normalize("NFC", text)

    # 9a. Collapse runs of spaces / tabs to a single space
    text = _RE_MULTI_SPACE.sub(" ", text)

    # 9b. Strip trailing space on each line
    text = "\n".join(line.strip() for line in text.splitlines())

    # 9c. Collapse 3+ consecutive blank lines to 2
    text = _RE_MULTI_NL.sub("\n\n", text)

    # Final strip
    text = text.strip()

    return text


# ── short-row filter  [24†L238-L241] ─────────────────────────────────────────

MIN_WORDS = 10

def is_informative(text: str) -> bool:
    """Return True if the cleaned article contains at least MIN_WORDS words."""
    if not isinstance(text, str):
        return False
    return len(text.split()) >= MIN_WORDS


# ── statistics ────────────────────────────────────────────────────────────────

def diff_stats(original: pd.Series, cleaned: pd.Series, dropped: int) -> dict:
    changed    = (original != cleaned).sum()
    total      = len(original)
    orig_chars = original.str.len().sum()
    cln_chars  = cleaned.str.len().sum()
    return {
        "total_rows":        total,
        "rows_changed":      int(changed),
        "pct_changed":       changed / total if total else 0,
        "rows_dropped":      dropped,
        "chars_before":      int(orig_chars),
        "chars_after":       int(cln_chars),
        "chars_removed":     int(orig_chars - cln_chars),
        "pct_chars_removed": (orig_chars - cln_chars) / orig_chars if orig_chars else 0,
    }


# ── Excel writing ─────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(name="Arial", size=10)
SPLIT_COLORS = {"train": "E2EFDA", "val": "FFF2CC", "test": "FCE4D6"}
COL_ORDER    = ["label", "article", "topic", "news_type", "split"]
COL_WIDTHS   = {"label": 8, "article": 80, "topic": 28, "news_type": 12, "split": 10}


def write_sheet(ws, df: pd.DataFrame) -> None:
    cols = [c for c in COL_ORDER if c in df.columns]
    cols += [c for c in df.columns if c not in cols]
    df = df[cols]

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    has_split = "split" in cols
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = (
            PatternFill("solid", start_color=SPLIT_COLORS.get(getattr(row, "split", ""), "FFFFFF"))
            if has_split
            else PatternFill("solid", start_color="EEF3FB" if ri % 2 == 0 else "FFFFFF")
        )
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.alignment = Alignment(
                wrap_text=(cols[ci - 1] == "article"), vertical="top"
            )

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── main ──────────────────────────────────────────────────────────────────────

def main(input_path: str, output_path: str) -> None:
    print(f"\n📂  Input : {input_path}")
    print(f"📄  Output: {output_path}\n")

    ext = pathlib.Path(input_path).suffix.lower()
    raw_sheets = (
        {"Sheet1": pd.read_csv(input_path)}
        if ext == ".csv"
        else pd.read_excel(input_path, sheet_name=None)
    )

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in raw_sheets.items():
        df = df.copy()
        df.columns = df.columns.str.lower()

        if "article" not in df.columns:
            print(f"['{sheet_name}']  no 'article' column — copied unchanged.")
            ws = wb.create_sheet(title=sheet_name)
            write_sheet(ws, df)
            continue

        original      = df["article"].astype(str)
        df["article"] = original.apply(clean_article)

        # Drop rows whose cleaned article is below the minimum word threshold
        before_drop = len(df)
        df = df[df["article"].apply(is_informative)].reset_index(drop=True)
        dropped = before_drop - len(df)

        s = diff_stats(original, df["article"], dropped)
        print(f"['{sheet_name}']  {s['total_rows']} rows")
        print(f"  Rows changed      : {s['rows_changed']}  ({s['pct_changed']:.1%})")
        print(f"  Rows dropped      : {s['rows_dropped']}  (< {MIN_WORDS} words after cleaning)")
        print(f"  Characters removed: {s['chars_removed']:,}  ({s['pct_chars_removed']:.1%})\n")

        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, df)

    wb.save(output_path)
    print(f"✅  Saved → {output_path}")


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print("Usage: python clean.py <input.xlsx|csv> [output.xlsx]")
        sys.exit(1)

    inp = sys.argv[1]
    out = (
        sys.argv[2]
        if len(sys.argv) == 3
        else str(
            pathlib.Path(inp)
            .with_stem(pathlib.Path(inp).stem + "_cleaned")
            .with_suffix(".xlsx")
        )
    )
    main(inp, out)
