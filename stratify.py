"""
stratify.py
-----------
Produces a single Excel workbook where each sheet corresponds to one
news-type-ratio configuration.

Split logic (applied to every sheet):
  - test  : 15% of total_samples, always 25:25:25:25 across news types — FIXED
  - val   : 15% of total_samples, always 25:25:25:25 across news types — FIXED
  - train : 70% of total_samples, at the sheet's configured news-type ratio

Test and val rows are sampled once and reused identically in every sheet.
Training rows are sampled fresh per configuration from the remaining pool.

Usage:
    python stratify.py config.json
"""

import json
import sys
import math
import pathlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ── config loading ────────────────────────────────────────────────────────────

def load_config(path: str) -> dict:
    with open(path) as f:
        cfg = json.load(f)

    total = sum(cfg.get("topic_ratios", {}).values())
    if not math.isclose(total, 1.0, abs_tol=1e-6):
        raise ValueError(
            f"topic_ratios must sum to 1.0 (currently {total:.6f})."
        )
    if not cfg.get("total_samples"):
        raise ValueError("'total_samples' is required in the config.")
    return cfg


# ── data loading ──────────────────────────────────────────────────────────────

def load_sheets(input_file: str, sheet_map: dict) -> dict[str, pd.DataFrame]:
    """Return {news_type: DataFrame} with normalised columns and a stable _idx."""
    xl = pd.read_excel(input_file, sheet_name=None)
    frames = {}
    for news_type, sheet_name in sheet_map.items():
        if sheet_name not in xl:
            raise KeyError(
                f"Sheet '{sheet_name}' (news type '{news_type}') not found. "
                f"Available: {list(xl.keys())}"
            )
        df = xl[sheet_name].copy()
        missing = {"label", "article", "topic"} - set(df.columns.str.lower())
        if missing:
            raise ValueError(f"Sheet '{sheet_name}' missing columns: {missing}")

        df.columns = df.columns.str.lower()
        df["topic"]     = df["topic"].str.strip().str.lower()
        df["news_type"] = news_type
        df["_idx"]      = range(len(df))   # stable row id within news type
        frames[news_type] = df.reset_index(drop=True)
    return frames


# ── ratio → counts ────────────────────────────────────────────────────────────

def ratios_to_counts(ratios: dict[str, float], total: int) -> dict[str, int]:
    """Largest-remainder allocation so counts sum exactly to total."""
    raw    = {k: v * total for k, v in ratios.items()}
    floors = {k: int(v) for k, v in raw.items()}
    deficit = total - sum(floors.values())
    for k in sorted(raw, key=lambda k: raw[k] - floors[k], reverse=True)[:deficit]:
        floors[k] += 1
    return floors


# ── topic-stratified sampler ──────────────────────────────────────────────────

def sample_stratified(
    df: pd.DataFrame,
    n_total: int,
    topic_ratios: dict[str, float],
    rng: np.random.Generator,
    exclude_idx: set | None = None,
) -> pd.DataFrame:
    """
    Sample n_total rows from df, topic-stratified.
    Rows whose _idx is in exclude_idx are not eligible.
    """
    pool = df if not exclude_idx else df[~df["_idx"].isin(exclude_idx)]
    topic_counts = ratios_to_counts(topic_ratios, n_total)

    parts = []
    for topic, count in topic_counts.items():
        if count == 0:
            continue
        topic_pool = pool[pool["topic"] == topic]
        if topic_pool.empty:
            print(f"  [WARN] topic '{topic}' not found — skipped.")
            continue
        n_draw = min(count, len(topic_pool))
        if n_draw < count:
            print(f"  [WARN] topic '{topic}': wanted {count}, only {n_draw} available.")
        parts.append(
            topic_pool.sample(n=n_draw, random_state=int(rng.integers(1 << 31)))
        )

    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=df.columns)


# ── sheet name builder ────────────────────────────────────────────────────────

def sheet_label(type_ratios: dict[str, float]) -> str:
    """e.g. {HR:0.25, HF:0.25, AI-F:0.25, AI-R:0.25} → 'HR25-HF25-AIF25-AIR25'"""
    parts = []
    for k, v in type_ratios.items():
        key = k.replace("-", "")          # 'AI-F' → 'AIF'
        pct = int(round(v * 100))
        parts.append(f"{key}{pct}")
    return "-".join(parts)


# ── Excel writing ─────────────────────────────────────────────────────────────

COL_ORDER   = ["label", "article", "topic", "news_type", "split"]
COL_WIDTHS  = {"label": 8, "article": 80, "topic": 20, "news_type": 12, "split": 10}

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
EVEN_FILL   = PatternFill("solid", start_color="D9E1F2")
DATA_FONT   = Font(name="Arial", size=10)

SPLIT_COLORS = {
    "train": "E2EFDA",   # soft green
    "val":   "FFF2CC",   # soft yellow
    "test":  "FCE4D6",   # soft orange
}


def write_sheet(ws, df: pd.DataFrame) -> None:
    cols = [c for c in COL_ORDER if c in df.columns]
    df   = df[cols]

    # header
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # data
    for ri, row in enumerate(df.itertuples(index=False), 2):
        split_val = getattr(row, "split", "train")
        row_fill  = PatternFill("solid", start_color=SPLIT_COLORS.get(split_val, "FFFFFF"))
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = DATA_FONT
            cell.fill      = row_fill
            cell.alignment = Alignment(
                wrap_text=(cols[ci - 1] == "article"), vertical="top"
            )

    # column widths
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = \
            COL_WIDTHS.get(col, 15)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── main ──────────────────────────────────────────────────────────────────────

EQUAL_RATIO = {"HR": 0.25, "HF": 0.25, "AI-F": 0.25, "AI-R": 0.25}

def main(config_path: str) -> None:
    cfg           = load_config(config_path)
    seed          = cfg.get("seed", 42)
    input_file    = cfg["input_file"]
    output_file   = cfg.get("output_file", "stratified_dataset.xlsx")
    sheet_map     = cfg.get("sheet_names",
                        {"HR": "HR", "HF": "HF", "AI-F": "AI-F", "AI-R": "AI-R"})
    total_samples = cfg["total_samples"]
    topic_ratios  = cfg["topic_ratios"]
    ratio_configs = cfg["news_type_ratios"]

    n_test  = ratios_to_counts({"test": 0.15, "val": 0.15, "train": 0.70}, total_samples)
    n_tv    = n_test["test"]   # test size
    n_vv    = n_test["val"]    # val size
    n_tr    = n_test["train"]  # train size

    print(f"\n📂  Input        : {input_file}")
    print(f"📄  Output       : {output_file}")
    print(f"🌱  Seed         : {seed}")
    print(f"📦  Total samples: {total_samples}  "
          f"(train={n_tr}, val={n_vv}, test={n_tv})")
    print(f"📊  Topic ratios : {topic_ratios}")
    print(f"🔢  Ratio configs: {len(ratio_configs)}")
    print()

    frames = load_sheets(input_file, sheet_map)

    # ── sample fixed test + val (25:25:25:25, topic-stratified) ──────────────
    print("Sampling fixed test / val sets (25:25:25:25) …")
    rng_fixed    = np.random.default_rng(seed)
    tv_counts    = ratios_to_counts(EQUAL_RATIO, n_tv)   # counts per news type for test
    vv_counts    = ratios_to_counts(EQUAL_RATIO, n_vv)   # counts per news type for val

    test_parts, val_parts = [], []
    used: dict[str, set] = {nt: set() for nt in frames}   # _idx sets per news type

    for news_type, count in tv_counts.items():
        part = sample_stratified(frames[news_type], count, topic_ratios, rng_fixed,
                                 exclude_idx=used[news_type])
        part["split"] = "test"
        used[news_type].update(part["_idx"].tolist())
        test_parts.append(part)

    for news_type, count in vv_counts.items():
        part = sample_stratified(frames[news_type], count, topic_ratios, rng_fixed,
                                 exclude_idx=used[news_type])
        part["split"] = "val"
        used[news_type].update(part["_idx"].tolist())
        val_parts.append(part)

    test_df = pd.concat(test_parts, ignore_index=True)
    val_df  = pd.concat(val_parts,  ignore_index=True)
    print(f"  test rows : {len(test_df)}")
    print(f"  val rows  : {len(val_df)}\n")

    # ── build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    for idx, type_ratios in enumerate(ratio_configs, 1):
        ratio_sum = sum(type_ratios.values())
        if not math.isclose(ratio_sum, 1.0, abs_tol=1e-6):
            raise ValueError(
                f"news_type_ratios[{idx-1}] must sum to 1.0 "
                f"(currently {ratio_sum:.6f})."
            )

        name = sheet_label(type_ratios)
        print(f"[{idx}/{len(ratio_configs)}] Sheet '{name}'  train ratio={type_ratios}")

        # sample training rows (exclude already-used test+val rows)
        train_counts = ratios_to_counts(type_ratios, n_tr)
        train_parts  = []
        for news_type, count in train_counts.items():
            rng_train = np.random.default_rng(seed + idx * 1000)
            part = sample_stratified(frames[news_type], count, topic_ratios, rng_train,
                                     exclude_idx=used[news_type])
            part["split"] = "train"
            train_parts.append(part)

        train_df = pd.concat(train_parts, ignore_index=True)

        # combine train + val + test, shuffle training rows only
        train_df = train_df.sample(frac=1, random_state=seed).reset_index(drop=True)
        combined = pd.concat([train_df, val_df, test_df], ignore_index=True)
        combined.drop(columns=["_idx"], inplace=True)

        print(f"  rows — train:{len(train_df)}  val:{len(val_df)}  test:{len(test_df)}  "
              f"total:{len(combined)}")
        print(f"  news-type dist (train):\n"
              f"{train_df['news_type'].value_counts().to_string()}")

        ws = wb.create_sheet(title=name)
        write_sheet(ws, combined)

    out_path = pathlib.Path(output_file)
    wb.save(str(out_path))
    print(f"\n✅  Saved → {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python stratify.py config.json")
        sys.exit(1)
    main(sys.argv[1])
