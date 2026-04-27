"""
combine.py
----------
Combines multiple Excel spreadsheets (each with label, article, topic columns)
into a single output workbook with one sheet per news type.

Scans each configured folder for .xlsx files, reads them, validates columns,
and stacks them. Skips non-data files (e.g. generation scripts saved as xlsx).

Usage:
    python combine.py config_combine.json

Example config (config_combine.json):
    {
        "output_file": "combined.xlsx",
        "sources": [
            {
                "folder": "AI-F",
                "news_type": "HF",
                "exclude": ["AI-F Generation.xlsx"]
            },
            {
                "folder": "AI-R",
                "news_type": "HR",
                "exclude": ["AI-R Generation.xlsx"],
                "extra_cols": ["original_article"]
            }
        ]
    }

Config keys:
    output_file  (optional) default: combined.xlsx
    sources      (required) list of source groups:
        folder       (required) path to folder containing .xlsx files
        news_type    (required) label written into the news_type column (e.g. HF, HR)
        exclude      (optional) list of filenames to skip within the folder
        extra_cols   (optional) list of additional columns to keep (e.g. ["original_article"])
                     Columns that are missing in a file are silently filled with NaN.
"""

import json
import sys
import pathlib
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── config ────────────────────────────────────────────────────────────────────

def load_config(path: str) -> dict:
    with open(path) as f:
        cfg = json.load(f)
    if not cfg.get("sources"):
        raise ValueError("'sources' list is required in the config.")
    for s in cfg["sources"]:
        if not s.get("folder"):
            raise ValueError("Each source entry must have a 'folder' key.")
        if not s.get("news_type"):
            raise ValueError("Each source entry must have a 'news_type' key.")
    return cfg


# ── reading ───────────────────────────────────────────────────────────────────

REQUIRED_COLS = {"label", "article", "topic"}

def read_source(
    folder: str,
    news_type: str,
    exclude: list[str],
    extra_cols: list[str],
) -> pd.DataFrame:
    folder_path = pathlib.Path(folder)
    if not folder_path.exists():
        raise FileNotFoundError(f"Folder not found: {folder_path.resolve()}")

    import re
    def _numeric_key(p):
        m = re.search(r"\d+", p.stem)
        return int(m.group()) if m else 0
    xlsx_files = sorted(folder_path.glob("*.xlsx"), key=_numeric_key)
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx files found in: {folder_path.resolve()}")

    exclude_set = {e.lower() for e in (exclude or [])}
    parts = []

    for xlsx_path in xlsx_files:
        if xlsx_path.name.lower() in exclude_set:
            print(f"  [SKIP] {xlsx_path.name}")
            continue

        df = pd.read_excel(xlsx_path)
        df.columns = df.columns.str.strip().str.lower()

        missing_required = REQUIRED_COLS - set(df.columns)
        if missing_required:
            print(f"  [SKIP] {xlsx_path.name} — missing columns: {missing_required}")
            continue

        # Always include required columns
        cols_to_keep = ["label", "article", "topic"]

        # Include extra columns if requested; fill with NaN if absent in this file
        for col in extra_cols:
            col_lower = col.strip().lower()
            if col_lower in df.columns:
                cols_to_keep.append(col_lower)
            else:
                df[col_lower] = pd.NA
                cols_to_keep.append(col_lower)
                print(f"  [WARN] {xlsx_path.name} — '{col_lower}' not found, filled with NaN")

        df = df[cols_to_keep].copy()
        df["news_type"] = news_type
        parts.append(df)
        print(f"  [OK]   {xlsx_path.name:40s}  {len(df):>5} rows")

    if not parts:
        raise ValueError(f"No valid files loaded from '{folder}' for news_type '{news_type}'.")

    combined = pd.concat(parts, ignore_index=True)
    combined["topic"]   = combined["topic"].str.strip().str.lower()
    combined["article"] = combined["article"].astype(str).str.strip()
    if "original_article" in combined.columns:
        combined["original_article"] = combined["original_article"].astype(str).str.strip()
    return combined


# ── Excel writing ─────────────────────────────────────────────────────────────

BASE_COL_ORDER  = ["label", "article", "original_article", "topic", "news_type"]
COL_WIDTHS = {
    "label":            8,
    "article":         80,
    "original_article": 80,
    "topic":           28,
    "news_type":       12,
}

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT   = Font(name="Arial", size=10)
EVEN_FILL   = PatternFill("solid", start_color="EEF3FB")
ODD_FILL    = PatternFill("solid", start_color="FFFFFF")


def write_sheet(ws, df: pd.DataFrame) -> None:
    # Only include columns that actually exist in this sheet's data,
    # respecting the canonical column order.
    cols = [c for c in BASE_COL_ORDER if c in df.columns]
    df   = df[cols]

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wrap_cols = {"article", "original_article"}

    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.alignment = Alignment(
                wrap_text=(cols[ci - 1] in wrap_cols), vertical="top"
            )

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── summary ───────────────────────────────────────────────────────────────────

def print_summary(frames: dict[str, pd.DataFrame]) -> None:
    print("\n── Summary ──────────────────────────────────────────")
    total = 0
    for news_type, df in frames.items():
        extra = [c for c in df.columns if c not in {*REQUIRED_COLS, "news_type"}]
        extra_note = f"  [extra: {', '.join(extra)}]" if extra else ""
        print(f"\n  {news_type}  ({len(df)} rows){extra_note}")
        for topic, grp in df.groupby("topic"):
            print(f"    {topic:<30} {len(grp):>5} rows")
        total += len(df)
    print(f"\n  TOTAL: {total} rows across {len(frames)} sheet(s)")
    print("─────────────────────────────────────────────────────\n")


# ── main ──────────────────────────────────────────────────────────────────────

def main(config_path: str) -> None:
    cfg         = load_config(config_path)
    output_file = cfg.get("output_file", "combined.xlsx")
    sources     = cfg["sources"]

    print(f"\n📄  Output: {output_file}\n")

    frames: dict[str, pd.DataFrame] = {}

    for source in sources:
        folder     = source["folder"]
        news_type  = source["news_type"]
        exclude    = source.get("exclude", [])
        extra_cols = [c.strip().lower() for c in source.get("extra_cols", [])]

        extra_note = f"  extra_cols={extra_cols}" if extra_cols else ""
        print(f"📂  {folder}  →  news_type='{news_type}'{extra_note}")
        df = read_source(folder, news_type, exclude, extra_cols)
        frames[news_type] = df
        print(f"       Loaded {len(df)} rows\n")

    print_summary(frames)

    wb = Workbook()
    wb.remove(wb.active)

    for news_type, df in frames.items():
        ws = wb.create_sheet(title=news_type)
        write_sheet(ws, df)

    out_path = pathlib.Path(output_file)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    sheets = list(frames.keys())
    print(f"✅  Saved → {out_path}  (sheets: {', '.join(repr(s) for s in sheets)})")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python combine.py config_combine.json")
        sys.exit(1)
    main(sys.argv[1])