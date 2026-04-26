"""
merge_sheets.py
---------------
Pulls specific sheets from multiple Excel workbooks and combines them
into a single output workbook — one sheet per source, preserving data as-is.

Usage:
    python merge_sheets.py config_merge.json

Example config (config_merge.json):
    {
        "output_file": "merged.xlsx",
        "sources": [
            { "file": "human_real.xlsx",    "sheet": "HR" },
            { "file": "human_fake.xlsx",    "sheet": "HF" },
            { "file": "combined.xlsx",      "sheet": "HF",  "rename": "AI-F" },
            { "file": "combined.xlsx",      "sheet": "HR",  "rename": "AI-R" }
        ]
    }

Config keys per source:
    file    (required) path to the source Excel file
    sheet   (required) sheet name to read from that file
    rename  (optional) name to use in the output workbook (defaults to sheet)
"""

import json
import sys
import pathlib
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── config ────────────────────────────────────────────────────────────────────

def load_config(path: str) -> dict:
    with open(path) as f:
        cfg = json.load(f)
    if not cfg.get("sources"):
        raise ValueError("'sources' list is required in the config.")
    for s in cfg["sources"]:
        if not s.get("file"):
            raise ValueError("Each source must have a 'file' key.")
        if not s.get("sheet"):
            raise ValueError("Each source must have a 'sheet' key.")
    return cfg


# ── Excel writing ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT   = Font(name="Arial", size=10)
EVEN_FILL   = PatternFill("solid", start_color="EEF3FB")
ODD_FILL    = PatternFill("solid", start_color="FFFFFF")

COL_WIDTHS  = {"label": 8, "article": 80, "topic": 28, "news_type": 12, "split": 10}


def write_sheet(ws, df: pd.DataFrame) -> None:
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        for ci, val in enumerate(row, 1):
            col_name = df.columns[ci - 1]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.alignment = Alignment(
                wrap_text=(col_name == "article"), vertical="top"
            )

    for ci, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── main ──────────────────────────────────────────────────────────────────────

def main(config_path: str) -> None:
    cfg         = load_config(config_path)
    output_file = cfg.get("output_file", "merged.xlsx")
    sources     = cfg["sources"]

    print(f"\n📄  Output: {output_file}\n")

    wb = Workbook()
    wb.remove(wb.active)

    for source in sources:
        file_path  = pathlib.Path(source["file"])
        sheet_name = source["sheet"]
        out_name   = source.get("rename", sheet_name)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path.resolve()}")

        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df.columns = df.columns.str.strip().str.lower()

        ws = wb.create_sheet(title=out_name)
        write_sheet(ws, df)

        print(f"  ✓  {file_path.name}  [{sheet_name}]  →  sheet '{out_name}'  ({len(df)} rows)")

    out_path = pathlib.Path(output_file)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"\n✅  Saved → {out_path}  ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python merge_sheets.py config_merge.json")
        sys.exit(1)
    main(sys.argv[1])
