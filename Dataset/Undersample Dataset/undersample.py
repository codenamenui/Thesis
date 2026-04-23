"""
undersample.py
--------------
Produces a single Excel workbook with one sheet containing balanced HR and HF data,
plus a Summary sheet with counts and percentages for every breakdown.

Undersampling logic:
  For each topic, find the minimum row count between HR and HF sheets.
  Sample exactly that many rows from EACH sheet for that topic.
  This ensures:
    - Topics are balanced within each news type (HR and HF have equal counts per topic)
    - HR and HF are balanced against each other

Split logic (applied after undersampling):
  - test  : 15% of total_samples, 50:50 HR:HF — FIXED (sampled once, never changes)
  - val   : 15% of total_samples, 50:50 HR:HF — FIXED (sampled once, never changes)
  - train : 70% of total_samples, 50:50 HR:HF

Usage:
    python undersample.py config_undersample.json
"""

import json
import sys
import math
import pathlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── config loading ────────────────────────────────────────────────────────────

def load_config(path: str) -> dict:
    with open(path) as f:
        cfg = json.load(f)
    if not cfg.get("input_file"):
        raise ValueError("'input_file' is required in the config.")
    return cfg


# ── data loading ──────────────────────────────────────────────────────────────

def load_sheets(input_file: str, sheet_map: dict) -> dict[str, pd.DataFrame]:
    """Return {news_type: DataFrame} with normalised columns and a stable _idx."""
    xl = pd.read_excel(input_file, sheet_name=None)
    frames = {}
    for news_type, sheet_name in sheet_map.items():
        if sheet_name not in xl:
            raise KeyError(
                f"Sheet '{sheet_name}' (news type '{news_type}') not found. "
                f"Available: {list(xl.keys())}"
            )
        df = xl[sheet_name].copy()
        df.columns = df.columns.str.lower()

        missing = {"label", "article", "topic"} - set(df.columns)
        if missing:
            raise ValueError(f"Sheet '{sheet_name}' missing columns: {missing}")

        df["topic"]     = df["topic"].str.strip().str.lower()
        df["news_type"] = news_type
        df["_idx"]      = range(len(df))
        frames[news_type] = df.reset_index(drop=True)

    return frames


# ── undersampling ─────────────────────────────────────────────────────────────

def compute_topic_caps(frames: dict[str, pd.DataFrame]) -> dict[str, int]:
    """
    For each topic, find the minimum row count across all sheets.
    This becomes the cap — how many rows we draw from EACH sheet.
    """
    topic_counts: dict[str, dict[str, int]] = {}

    for news_type, df in frames.items():
        for topic, group in df.groupby("topic"):
            topic_counts.setdefault(topic, {})[news_type] = len(group)

    caps = {}
    for topic, counts in topic_counts.items():
        min_count = min(counts.values())
        caps[topic] = min_count

        # informational breakdown
        count_str = "  ".join(f"{nt}={n}" for nt, n in sorted(counts.items()))
        print(f"  topic '{topic}': {count_str}  → cap={min_count}")

    return caps


def undersample_sheet(
    df: pd.DataFrame,
    caps: dict[str, int],
    rng: np.random.Generator,
    exclude_idx: set | None = None,
    label: str = "",
) -> pd.DataFrame:
    """
    Sample up to caps[topic] rows per topic from df.
    Rows in exclude_idx are not eligible.
    """
    pool = df if not exclude_idx else df[~df["_idx"].isin(exclude_idx)]
    parts = []

    for topic, cap in caps.items():
        if cap == 0:
            continue
        topic_pool = pool[pool["topic"] == topic]
        if topic_pool.empty:
            print(f"  [WARN]{label} topic '{topic}' not found in pool — skipped.")
            continue
        n_draw = min(cap, len(topic_pool))
        if n_draw < cap:
            print(
                f"  [WARN]{label} topic '{topic}': "
                f"wanted {cap}, only {n_draw} available after exclusions."
            )
        parts.append(
            topic_pool.sample(n=n_draw, random_state=int(rng.integers(1 << 31)))
        )

    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=df.columns)


# ── Excel styles ──────────────────────────────────────────────────────────────

COL_ORDER  = ["label", "article", "topic", "news_type", "split"]
COL_WIDTHS = {"label": 8, "article": 80, "topic": 28, "news_type": 12, "split": 10}

HEADER_FILL    = PatternFill("solid", start_color="1F4E79")   # dark navy
HEADER_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=11)

SECTION_FILL   = PatternFill("solid", start_color="1F4E79")   # same navy for section titles
SECTION_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=12)

SUBHDR_FILL    = PatternFill("solid", start_color="2E75B6")   # medium blue for table headers
SUBHDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)

TOTAL_FILL     = PatternFill("solid", start_color="D6DCE4")   # light grey for totals
TOTAL_FONT     = Font(name="Arial", bold=True, size=10)

EVEN_FILL      = PatternFill("solid", start_color="EEF3FB")   # very light blue
ODD_FILL       = PatternFill("solid", start_color="FFFFFF")
DATA_FONT      = Font(name="Arial", size=10)

SPLIT_COLORS = {
    "train": "E2EFDA",   # soft green
    "val":   "FFF2CC",   # soft yellow
    "test":  "FCE4D6",   # soft orange
}

THIN_BORDER = Border(
    bottom=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin",    color="B8CCE4"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


# ── data sheet writer ─────────────────────────────────────────────────────────

def write_data_sheet(ws, df: pd.DataFrame) -> None:
    cols = [c for c in COL_ORDER if c in df.columns]
    df   = df[cols]

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    for ri, row in enumerate(df.itertuples(index=False), 2):
        split_val = getattr(row, "split", "train")
        row_fill  = PatternFill("solid", start_color=SPLIT_COLORS.get(split_val, "FFFFFF"))
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = DATA_FONT
            cell.fill      = row_fill
            cell.alignment = Alignment(
                wrap_text=(cols[ci - 1] == "article"), vertical="top"
            )

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── summary sheet writer ──────────────────────────────────────────────────────

def _cell(ws, row, col, value, font=None, fill=None, alignment=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:        c.font        = font
    if fill:        c.fill        = fill
    if alignment:   c.alignment   = alignment
    if number_format: c.number_format = number_format
    return c


def _section_title(ws, row, col, title, n_cols):
    """Write a wide section-title cell spanning n_cols columns."""
    c = _cell(ws, row, col, title, font=SECTION_FONT, fill=SECTION_FILL, alignment=LEFT)
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row,   end_column=col + n_cols - 1
    )
    ws.row_dimensions[row].height = 20
    return row + 1


def _table_headers(ws, row, col, headers):
    for i, h in enumerate(headers):
        _cell(ws, row, col + i, h, font=SUBHDR_FONT, fill=SUBHDR_FILL, alignment=CENTER)
    ws.row_dimensions[row].height = 18
    return row + 1


def _data_row(ws, row, col, values, even=True, bold=False, fill_override=None):
    fill = fill_override or (EVEN_FILL if even else ODD_FILL)
    fnt  = Font(name="Arial", bold=bold, size=10)
    for i, v in enumerate(values):
        align = LEFT if i == 0 else CENTER
        fmt   = "0.0%" if isinstance(v, float) and 0 <= v <= 1 else None
        _cell(ws, row, col + i, v, font=fnt, fill=fill, alignment=align,
              number_format=fmt)
    return row + 1


def write_summary_sheet(
    ws,
    combined: pd.DataFrame,
    frames_raw: dict[str, pd.DataFrame],   # original counts before undersampling
    caps: dict[str, int],
) -> None:
    """
    Writes 5 summary tables to ws:
      1. Dataset Overview
      2. Split Breakdown (train / val / test)
      3. News-Type Breakdown by Split
      4. Topic Breakdown by Split
      5. Pre- vs Post-Undersampling per Topic
    """

    # column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    total = len(combined)
    splits     = ["train", "val", "test"]
    news_types = sorted(combined["news_type"].unique())
    topics     = sorted(combined["topic"].unique())

    ROW = 1  # current write row

    # ── 1. OVERVIEW ───────────────────────────────────────────────────────────
    ROW = _section_title(ws, ROW, 1, "1 · Dataset Overview", 4)
    ROW = _table_headers(ws, ROW, 1, ["Metric", "Value"])

    overview_rows = [
        ("Total rows (final)", total),
        ("  Train rows",       len(combined[combined["split"] == "train"])),
        ("  Val rows",         len(combined[combined["split"] == "val"])),
        ("  Test rows",        len(combined[combined["split"] == "test"])),
        ("News types",         ", ".join(news_types)),
        ("Unique topics",      len(topics)),
        ("Split ratio",        "70 / 15 / 15  (train / val / test)"),
        ("Val & test fixed?",  "YES — sampled once with fixed seed; identical across any re-run"),
    ]
    for i, (k, v) in enumerate(overview_rows):
        ROW = _data_row(ws, ROW, 1, [k, v], even=(i % 2 == 0))

    ROW += 1  # blank separator

    # ── 2. SPLIT BREAKDOWN ────────────────────────────────────────────────────
    ROW = _section_title(ws, ROW, 1, "2 · Split Breakdown", 3)
    ROW = _table_headers(ws, ROW, 1, ["Split", "Count", "% of Total"])

    for i, split in enumerate(splits):
        n = len(combined[combined["split"] == split])
        ROW = _data_row(ws, ROW, 1, [split, n, n / total], even=(i % 2 == 0))

    ROW = _data_row(ws, ROW, 1, ["TOTAL", total, 1.0],
                    bold=True, fill_override=TOTAL_FILL)
    ROW += 1

    # ── 3. NEWS-TYPE BREAKDOWN BY SPLIT ───────────────────────────────────────
    ROW = _section_title(ws, ROW, 1, "3 · News-Type Breakdown by Split", 5)
    ROW = _table_headers(ws, ROW, 1, ["News Type", "Train", "Train %", "Val", "Val %",
                                       "Test", "Test %", "Total", "Total %"])
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12

    grand_total = total
    i = 0
    for nt in news_types:
        row_vals = [nt]
        nt_total = 0
        for split in splits:
            n = len(combined[(combined["news_type"] == nt) & (combined["split"] == split)])
            split_total = len(combined[combined["split"] == split])
            row_vals += [n, n / split_total]
            nt_total += n
        row_vals += [nt_total, nt_total / grand_total]
        ROW = _data_row(ws, ROW, 1, row_vals, even=(i % 2 == 0))
        i += 1

    # totals row
    totals_row = ["TOTAL"]
    for split in splits:
        s_total = len(combined[combined["split"] == split])
        totals_row += [s_total, 1.0]
    totals_row += [grand_total, 1.0]
    ROW = _data_row(ws, ROW, 1, totals_row, bold=True, fill_override=TOTAL_FILL)
    ROW += 1

    # ── 4. TOPIC BREAKDOWN BY SPLIT ───────────────────────────────────────────
    ROW = _section_title(ws, ROW, 1, "4 · Topic Breakdown by Split", 9)
    ROW = _table_headers(ws, ROW, 1, ["Topic", "Train", "Train %", "Val", "Val %",
                                       "Test", "Test %", "Total", "Total %"])

    i = 0
    topic_totals = []
    for topic in topics:
        row_vals = [topic]
        t_total  = 0
        for split in splits:
            n = len(combined[(combined["topic"] == topic) & (combined["split"] == split)])
            split_total = len(combined[combined["split"] == split])
            row_vals += [n, n / split_total]
            t_total  += n
        row_vals += [t_total, t_total / grand_total]
        topic_totals.append(t_total)
        ROW = _data_row(ws, ROW, 1, row_vals, even=(i % 2 == 0))
        i += 1

    totals_row = ["TOTAL"]
    for split in splits:
        s_total = len(combined[combined["split"] == split])
        totals_row += [s_total, 1.0]
    totals_row += [grand_total, 1.0]
    ROW = _data_row(ws, ROW, 1, totals_row, bold=True, fill_override=TOTAL_FILL)
    ROW += 1

    # ── 5. PRE vs POST UNDERSAMPLING ──────────────────────────────────────────
    all_news_types = sorted(frames_raw.keys())
    hdr = ["Topic"]
    for nt in all_news_types:
        hdr += [f"{nt} Before", f"{nt} After", f"{nt} Dropped"]
    hdr += ["Cap (per type)"]

    ROW = _section_title(ws, ROW, 1,
        "5 · Pre- vs Post-Undersampling Topic Counts (per news type)", len(hdr))
    ROW = _table_headers(ws, ROW, 1, hdr)

    # extend column widths for extra cols
    for ci in range(2, len(hdr) + 2):
        col_letter = get_column_letter(ci)
        if ws.column_dimensions[col_letter].width < 14:
            ws.column_dimensions[col_letter].width = 14

    i = 0
    for topic in topics:
        cap = caps.get(topic, 0)
        row_vals = [topic]
        for nt in all_news_types:
            before = len(frames_raw[nt][frames_raw[nt]["topic"] == topic])
            after  = cap   # same cap applied to every news type
            dropped = before - after
            row_vals += [before, after, dropped]
        row_vals.append(cap)
        ROW = _data_row(ws, ROW, 1, row_vals, even=(i % 2 == 0))
        i += 1

    # grand totals row
    grand_vals = ["TOTAL"]
    for nt in all_news_types:
        before_total = sum(
            len(frames_raw[nt][frames_raw[nt]["topic"] == t]) for t in topics
        )
        after_total  = sum(caps.get(t, 0) for t in topics)
        grand_vals  += [before_total, after_total, before_total - after_total]
    grand_vals.append(sum(caps.get(t, 0) for t in topics))
    ROW = _data_row(ws, ROW, 1, grand_vals, bold=True, fill_override=TOTAL_FILL)

    ws.freeze_panes = "B1"


# ── split allocation ──────────────────────────────────────────────────────────

def split_counts(n: int) -> tuple[int, int, int]:
    """Return (n_train, n_val, n_test) from total n using 70/15/15."""
    n_test  = round(n * 0.15)
    n_val   = round(n * 0.15)
    n_train = n - n_test - n_val
    return n_train, n_val, n_test


# ── main ──────────────────────────────────────────────────────────────────────

def main(config_path: str) -> None:
    cfg         = load_config(config_path)
    seed        = cfg.get("seed", 42)
    input_file  = cfg["input_file"]
    output_file = cfg.get("output_file", "undersampled_dataset.xlsx")
    sheet_map   = cfg.get("sheet_names", {"HR": "HR", "HF": "HF"})

    print(f"\n📂  Input : {input_file}")
    print(f"📄  Output: {output_file}")
    print(f"🌱  Seed  : {seed}\n")

    frames = load_sheets(input_file, sheet_map)

    # ── compute per-topic caps via undersampling ───────────────────────────────
    print("Computing per-topic row caps (min across HR and HF) …")
    caps = compute_topic_caps(frames)
    total_per_sheet = sum(caps.values())
    total_rows      = total_per_sheet * len(frames)   # same count from each sheet
    n_train, n_val, n_test = split_counts(total_rows)

    print(f"\n  Rows per sheet after undersampling : {total_per_sheet}")
    print(f"  Sheets                             : {list(frames.keys())}")
    print(f"  Total rows                         : {total_rows}")
    print(f"  Split  → train={n_train}  val={n_val}  test={n_test}\n")

    # ── fixed test + val (sampled once, reused) ───────────────────────────────
    print("Sampling fixed test / val sets …")
    rng_fixed = np.random.default_rng(seed)
    used: dict[str, set] = {nt: set() for nt in frames}

    # Split each topic's cap directly: train+val+test == cap per topic exactly.
    # Calling scale_caps three times independently causes rounding loss per topic
    # (e.g. disaster cap=36 → 5+5+25=35 instead of 36).
    test_caps, val_caps, train_caps = {}, {}, {}
    for topic, cap in caps.items():
        tr, vl, te        = split_counts(cap)
        train_caps[topic] = tr
        val_caps[topic]   = vl
        test_caps[topic]  = te

    test_parts, val_parts = [], []

    for news_type in frames:
        # test
        part = undersample_sheet(
            frames[news_type], test_caps, rng_fixed,
            exclude_idx=used[news_type], label=f" [{news_type}/test]"
        )
        part["split"] = "test"
        used[news_type].update(part["_idx"].tolist())
        test_parts.append(part)

        # val
        part = undersample_sheet(
            frames[news_type], val_caps, rng_fixed,
            exclude_idx=used[news_type], label=f" [{news_type}/val]"
        )
        part["split"] = "val"
        used[news_type].update(part["_idx"].tolist())
        val_parts.append(part)

    test_df = pd.concat(test_parts, ignore_index=True)
    val_df  = pd.concat(val_parts,  ignore_index=True)
    print(f"  test rows : {len(test_df)}")
    print(f"  val rows  : {len(val_df)}\n")

    # ── training set ──────────────────────────────────────────────────────────
    print("Sampling training set …")
    rng_train = np.random.default_rng(seed + 999)
    train_parts = []

    for news_type in frames:
        part = undersample_sheet(
            frames[news_type], train_caps, rng_train,
            exclude_idx=used[news_type], label=f" [{news_type}/train]"
        )
        part["split"] = "train"
        train_parts.append(part)

    train_df = pd.concat(train_parts, ignore_index=True)
    train_df = train_df.sample(frac=1, random_state=seed).reset_index(drop=True)
    print(f"  train rows: {len(train_df)}\n")

    # ── combine + write ───────────────────────────────────────────────────────
    combined = pd.concat([train_df, val_df, test_df], ignore_index=True)
    combined.drop(columns=["_idx"], inplace=True)

    print("Split summary:")
    print(combined["split"].value_counts().to_string())
    print("\nNews-type distribution:")
    print(combined["news_type"].value_counts().to_string())
    print("\nTopic distribution (train only):")
    print(
        train_df.drop(columns=["_idx"], errors="ignore")["topic"]
        .value_counts()
        .to_string()
    )

    wb = Workbook()
    wb.remove(wb.active)

    ws_data = wb.create_sheet(title="HR-HF-Undersampled")
    write_data_sheet(ws_data, combined)

    ws_sum = wb.create_sheet(title="Summary")
    write_summary_sheet(ws_sum, combined, frames, caps)

    out_path = pathlib.Path(output_file)
    wb.save(str(out_path))
    print(f"\n✅  Saved → {out_path}  (sheets: 'HR-HF-Undersampled', 'Summary')")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python undersample.py config_undersample.json")
        sys.exit(1)
    main(sys.argv[1])
