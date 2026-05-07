"""
stratify.py
-----------
Produces a single Excel workbook where each sheet corresponds to one
news-type-ratio configuration, plus a Summary sheet.

Input — two modes (mutually exclusive):
  A) input_file   : path to a single Excel workbook whose sheets are already
                    merged (i.e. the output of merge_sheets.py).
  B) merge_sources: list of {file, sheet, rename?} entries — the same format
                    as merge_sheets.py's config — so you can skip the merge
                    step entirely and run just this script.

Split logic (applied to every sheet):
  - test  : 15% of total_samples, always equal across news types — FIXED
  - val   : 15% of total_samples, always equal across news types — FIXED
  - train : 70% of total_samples, at the sheet's configured news-type ratio

Test and val rows are sampled once and reused identically in every sheet.
Training rows are sampled fresh per configuration from the remaining pool.

Usage:
    python stratify.py config.json

Config keys:
    merge_sources   (option A) list of {file, sheet, rename?} — skips merge step
    input_file      (option B) path to already-merged Excel file
    sheet_names     (option B only, optional) map of news_type → sheet name
                    default: {"HR":"HR","HF":"HF","AI-F":"AI-F","AI-R":"AI-R"}
    output_file     (optional) default: stratified_dataset.xlsx
    seed            (optional) default: 42
    total_samples   (required)
    topic_ratios    (required) must sum to 1.0
    news_type_ratios (required) list of ratio dicts, each must sum to 1.0
    truncation      (optional) enable article truncation after sampling.
                    Sub-keys:
                      model  (required) HuggingFace tokenizer name, e.g.
                             "jcblaise/bert-tagalog-base-cased"
                      cap    (optional) explicit integer token cap.
                             If omitted, the cap is auto-computed as the mean
                             token count of the news-type with the lowest mean
                             across the WHOLE dataset (all rows, before any
                             split), and is shared identically by every sheet.
                    When enabled, train articles are truncated to the cap and
                    the original text is preserved in an `original_article`
                    column.  Val/test articles are left untouched.
                    Token counts per split/news_type are added to the Summary.
"""

import json
import sys
import math
import pathlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ── tokenizer (lazy-loaded only when truncation is requested) ─────────────────

_tokenizer = None

def _get_tokenizer(model_name: str):
    global _tokenizer
    if _tokenizer is None:
        from transformers import AutoTokenizer
        print(f"  Loading tokenizer: {model_name} …")
        _tokenizer = AutoTokenizer.from_pretrained(model_name)
        print("  Tokenizer ready.")
    return _tokenizer


def _tokenize_texts(texts: list[str], model_name: str,
                    batch_size: int = 128) -> list[list[int]]:
    tok = _get_tokenizer(model_name)
    all_ids: list[list[int]] = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i : i + batch_size]
        enc   = tok(batch, add_special_tokens=True, truncation=False,
                    padding=False, return_attention_mask=False)
        all_ids.extend(enc["input_ids"])
        if i > 0 and i % 2048 == 0:
            print(f"    tokenized {i}/{len(texts)} rows …")
    return all_ids


def _decode_ids(ids: list[int], model_name: str) -> str:
    return _get_tokenizer(model_name).decode(ids, skip_special_tokens=True)


def _truncate_ids(ids: list[int], cap: int) -> list[int]:
    """Preserve [CLS] … [SEP] structure while capping at `cap` tokens."""
    if len(ids) <= cap:
        return ids
    return ids[:cap - 1] + [ids[-1]]


def compute_global_cap(tokenized_frames: dict[str, pd.DataFrame]) -> tuple[int, str]:
    """
    Compute the truncation cap **once** from the entire original corpus
    (all rows, all news types — before any train/val/test split).

    Cap = mean token count of the news_type with the lowest mean across
    the whole dataset, so the value is identical for every condition sheet.
    """
    means: dict[str, float] = {}
    for nt, df in tokenized_frames.items():
        vals = df["token_count"].dropna().values
        if len(vals):
            means[nt] = float(np.mean(vals))
    if not means:
        raise ValueError("No token counts found — cannot compute global truncation cap.")
    shortest_nt = min(means, key=means.get)
    cap  = int(means[shortest_nt])
    desc = (f"mean of {shortest_nt} (whole dataset) "
            f"[{', '.join(f'{k}={v:.0f}' for k, v in sorted(means.items()))}]")
    print(f"  Global truncation cap = {cap}  ({desc})")
    return cap, desc


def apply_truncation(combined: pd.DataFrame,
                     trunc_cfg: dict,
                     cap_override: int | None = None,
                     cap_desc_override: str | None = None,
                     ) -> tuple[pd.DataFrame, int, str]:
    """
    Tokenize the `article` column, truncate TRAIN rows to the cap, and
    store the original text in `original_article`.
    """
    if not isinstance(trunc_cfg, dict):
         raise TypeError("Config error: 'truncation' must be a dictionary with a 'model' key.")

    model_name = trunc_cfg["model"]
    df = combined.copy()

    # Tokenize every row
    texts   = df["article"].fillna("").astype(str).tolist()
    all_ids = _tokenize_texts(texts, model_name)
    df["_ids"]        = all_ids
    df["token_count"] = [len(ids) for ids in all_ids]

    # Determine cap
    if cap_override is not None:
        cap      = cap_override
        cap_desc = cap_desc_override or f"global cap={cap}"
    elif "cap" in trunc_cfg:
        cap      = int(trunc_cfg["cap"])
        cap_desc = f"fixed cap={cap}"
    else:
        raise ValueError(
            "No truncation cap available.  Either supply 'cap' in the config "
            "or call compute_global_cap() before the condition loop and pass "
            "the result as cap_override."
        )

    # Preserve original text then overwrite train articles
    df["original_article"] = df["article"]

    is_train = df["split"] == "train"
    trunc_ids = df.loc[is_train, "_ids"].apply(lambda ids: _truncate_ids(ids, cap))
    df.loc[is_train, "article"]      = trunc_ids.apply(lambda ids: _decode_ids(ids, model_name))
    df.loc[is_train, "token_count"]  = trunc_ids.apply(len)
    df.loc[is_train, "_ids"]         = trunc_ids

    pct_hit = (df.loc[is_train, "token_count"] == cap).mean()
    print(f"    cap={cap}  |  {pct_hit:.1%} of train rows truncated")

    df.drop(columns=["_ids"], inplace=True)
    return df, cap, cap_desc


def tokenize_frames(frames: dict[str, pd.DataFrame], model_name: str) -> dict[str, pd.DataFrame]:
    """
    Tokenize every raw frame's `article` column and attach token_count
    and _ids columns. Used for original and truncated statistics.
    """
    result = {}
    for nt, df in frames.items():
        print(f"  Tokenizing original '{nt}' ({len(df)} rows) …")
        texts   = df["article"].fillna("").astype(str).tolist()
        all_ids = _tokenize_texts(texts, model_name)
        df2 = df.copy()
        df2["token_count"] = [len(ids) for ids in all_ids]
        df2["_ids"]        = all_ids
        result[nt] = df2
    return result


def compute_original_token_stats(frames: dict[str, pd.DataFrame]) -> dict:
    """
    Given frames that already have a `token_count` column, return a nested
    dict:  { news_type: {"count": int, "avg": float, "min": int, "max": int,
                         "median": float, "std": float} }
    plus an "overall" key for the concatenated corpus.
    """
    stats: dict[str, dict] = {}
    all_counts: list[int] = []

    for nt, df in frames.items():
        tc = df["token_count"].dropna().astype(int)
        all_counts.extend(tc.tolist())
        stats[nt] = {
            "count":  int(len(tc)),
            "avg":    float(tc.mean()),
            "median": float(tc.median()),
            "std":    float(tc.std()),
            "min":    int(tc.min()),
            "max":    int(tc.max()),
        }

    tc_all = pd.Series(all_counts)
    stats["overall"] = {
        "count":  int(len(tc_all)),
        "avg":    float(tc_all.mean()),
        "median": float(tc_all.median()),
        "std":    float(tc_all.std()),
        "min":    int(tc_all.min()),
        "max":    int(tc_all.max()),
    }
    return stats


def compute_truncated_token_stats(frames: dict[str, pd.DataFrame], cap: int) -> dict:
    """
    Apply truncation to all rows using the specified cap (via _ids column)
    and return token count statistics for the truncated whole dataset.
    """
    stats: dict[str, dict] = {}
    all_counts: list[int] = []

    for nt, df in frames.items():
        ids_series = df["_ids"]
        truncated_counts = ids_series.apply(lambda ids: len(_truncate_ids(ids, cap)))
        tc = truncated_counts.astype(int)
        all_counts.extend(tc.tolist())
        stats[nt] = {
            "count":  int(len(tc)),
            "avg":    float(tc.mean()),
            "median": float(tc.median()),
            "std":    float(tc.std()),
            "min":    int(tc.min()),
            "max":    int(tc.max()),
        }

    tc_all = pd.Series(all_counts)
    stats["overall"] = {
        "count":  int(len(tc_all)),
        "avg":    float(tc_all.mean()),
        "median": float(tc_all.median()),
        "std":    float(tc_all.std()),
        "min":    int(tc_all.min()),
        "max":    int(tc_all.max()),
    }
    return stats


# ── config loading ────────────────────────────────────────────────────────────

def load_config(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        cfg = json.load(f)

    if not cfg.get("merge_sources") and not cfg.get("input_file"):
        raise ValueError("Config must have either 'merge_sources' or 'input_file'.")
    if not cfg.get("total_samples"):
        raise ValueError("'total_samples' is required.")

    total = sum(cfg.get("topic_ratios", {}).values())
    if not math.isclose(total, 1.0, abs_tol=1e-6):
        raise ValueError(f"topic_ratios must sum to 1.0 (currently {total:.6f}).")
    return cfg


# ── data loading ──────────────────────────────────────────────────────────────

def _prep_df(df: pd.DataFrame, news_type: str, source_label: str) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower()

    missing = {"label", "article", "topic"} - set(df.columns)
    if missing:
        raise ValueError(f"'{source_label}' missing columns: {missing}")

    df["topic"]     = df["topic"].str.strip().str.lower()
    df["news_type"] = news_type
    df["_idx"]      = range(len(df))
    return df.reset_index(drop=True)


def load_from_sources(sources: list[dict]) -> dict[str, pd.DataFrame]:
    frames = {}
    for src in sources:
        file_path  = pathlib.Path(src["file"])
        sheet_name = src["sheet"]
        news_type  = src.get("rename", sheet_name)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path.resolve()}")

        xl  = pd.read_excel(file_path, sheet_name=sheet_name)
        df  = _prep_df(xl, news_type, f"{file_path.name}[{sheet_name}]")
        frames[news_type] = df
        print(f"  ✓  {file_path.name}[{sheet_name}]  →  '{news_type}'  ({len(df)} rows)")
    return frames


def load_from_file(input_file: str, sheet_map: dict) -> dict[str, pd.DataFrame]:
    xl = pd.read_excel(input_file, sheet_name=None)
    frames = {}
    for news_type, sheet_name in sheet_map.items():
        if sheet_name not in xl:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {list(xl.keys())}")
        df = _prep_df(xl[sheet_name], news_type, f"{input_file}[{sheet_name}]")
        frames[news_type] = df
        print(f"  ✓  {input_file}[{sheet_name}]  →  '{news_type}'  ({len(df)} rows)")
    return frames


def ratios_to_counts(ratios: dict[str, float], total: int) -> dict[str, int]:
    raw    = {k: v * total for k, v in ratios.items()}
    floors = {k: int(v) for k, v in raw.items()}
    deficit = total - sum(floors.values())
    for k in sorted(raw, key=lambda k: raw[k] - floors[k], reverse=True)[:deficit]:
        floors[k] += 1
    return floors


def equal_ratio(news_types: list[str]) -> dict[str, float]:
    w = 1.0 / len(news_types)
    return {nt: w for nt in news_types}


def sample_stratified(df: pd.DataFrame, n_total: int, topic_ratios: dict[str, float], rng: np.random.Generator, exclude_idx: set | None = None) -> pd.DataFrame:
    pool = df if not exclude_idx else df[~df["_idx"].isin(exclude_idx)]
    topic_counts = ratios_to_counts(topic_ratios, n_total)

    parts = []
    for topic, count in topic_counts.items():
        if count == 0: continue
        topic_pool = pool[pool["topic"] == topic]
        if topic_pool.empty: continue
        n_draw = min(count, len(topic_pool))
        parts.append(topic_pool.sample(n=n_draw, random_state=int(rng.integers(1 << 31))))

    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=df.columns)


def sheet_label(type_ratios: dict[str, float]) -> str:
    hr, air = type_ratios.get("HR", 0), type_ratios.get("AI-R", 0)
    hf, aif = type_ratios.get("HF", 0), type_ratios.get("AI-F", 0)
    real_total, fake_total = hr + air or 1, hf + aif or 1
    return f"HR{int(round(hr/real_total*100))}-AIR{int(round(air/real_total*100))}-HF{int(round(hf/fake_total*100))}-AIF{int(round(aif/fake_total*100))}"


# ── Formatting & Excel Writing ────────────────────────────────────────────────

COL_ORDER  = ["label", "article", "original_article", "token_count", "topic", "news_type", "split"]
COL_WIDTHS = {"label": 8, "article": 80, "original_article": 80, "token_count": 12, "topic": 20, "news_type": 12, "split": 10}

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(name="Arial", size=10)
SECTION_FONT = Font(name="Arial", bold=True, size=11)
SPLIT_COLORS = {"train": "E2EFDA", "val": "FFF2CC", "test": "FCE4D6"}

SUMMARY_HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
SUMMARY_SECTION_FILL = PatternFill("solid", start_color="D6E4F0")
SUMMARY_ALT_FILL     = PatternFill("solid", start_color="F2F7FB")
SUMMARY_TOTAL_FILL   = PatternFill("solid", start_color="BDD7EE")
NEWS_TYPE_COLORS     = {"HR": "C6EFCE", "AI-R": "FFEB9C", "HF": "FFC7CE", "AI-F": "E2AFFF"}
ORIGINAL_ROW_FILL    = PatternFill("solid", start_color="F0F0F0")
OVERALL_ROW_FILL     = PatternFill("solid", start_color="D9D9D9")


def write_sheet(ws, df: pd.DataFrame) -> None:
    cols = [c for c in COL_ORDER if c in df.columns]
    df   = df[cols]
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ri, row in enumerate(df.itertuples(index=False), 2):
        split_val = getattr(row, "split", "train")
        row_fill  = PatternFill("solid", start_color=SPLIT_COLORS.get(split_val, "FFFFFF"))
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font, cell.fill = DATA_FONT, row_fill
            cell.alignment = Alignment(wrap_text=(cols[ci - 1] in {"article", "original_article"}), vertical="top")

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = COL_WIDTHS.get(col, 15)
    ws.freeze_panes = "A2"


def _summary_cell(ws, row, col, value, bold=False, fill=None, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fill: cell.fill = fill


def write_summary_sheet(
    ws,
    summary_records,
    news_types,
    topic_ratios,
    total_samples,
    n_train,
    n_val,
    n_test,
    original_token_stats: dict | None = None,
    truncated_token_stats: dict | None = None,
    truncation_cap: int | None = None,
    truncation_desc: str | None = None,
):
    splits = ["train", "val", "test"]
    topics = sorted(topic_ratios.keys())
    ws.title, current_row = "Summary", 1
    ws.column_dimensions["A"].width, ws.column_dimensions["B"].width = 28, 16

    # §1 GLOBAL CONFIG
    _summary_cell(ws, current_row, 1, "GLOBAL CONFIGURATION", bold=True, fill=SUMMARY_HEADER_FILL)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    for label, val in [("Total samples", total_samples), ("Train (target)", n_train), ("Val (target)", n_val), ("Test (target)", n_test)]:
        _summary_cell(ws, current_row, 1, label, bold=True, align="left", fill=SUMMARY_ALT_FILL)
        _summary_cell(ws, current_row, 2, val, align="left", fill=SUMMARY_ALT_FILL)
        current_row += 1
    current_row += 1

    # §2 PER-SHEET OVERVIEW
    _summary_cell(ws, current_row, 1, "PER-SHEET OVERVIEW", bold=True, fill=SUMMARY_HEADER_FILL)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    current_row += 1
    for ci, h in enumerate(["Sheet", "Train", "Val", "Test", "Total"], 1):
        _summary_cell(ws, current_row, ci, h, bold=True, fill=SUMMARY_SECTION_FILL)
    current_row += 1
    for rec in summary_records:
        _summary_cell(ws, current_row, 1, rec["sheet"], bold=True, align="left")
        _summary_cell(ws, current_row, 2, rec["n_train"])
        _summary_cell(ws, current_row, 3, rec["n_val"])
        _summary_cell(ws, current_row, 4, rec["n_test"])
        _summary_cell(ws, current_row, 5, rec["n_train"]+rec["n_val"]+rec["n_test"], bold=True, fill=SUMMARY_TOTAL_FILL)
        current_row += 1
    current_row += 1

    # §3 NEWS-TYPE BREAKDOWN
    _summary_cell(ws, current_row, 1, "NEWS-TYPE BREAKDOWN", bold=True, fill=SUMMARY_HEADER_FILL)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(news_types)+3)
    current_row += 1
    for ci, h in enumerate(["Sheet", "Split"] + news_types + ["Total"], 1):
        _summary_cell(ws, current_row, ci, h, bold=True, fill=SUMMARY_SECTION_FILL)
    current_row += 1
    for rec in summary_records:
        for split in splits:
            _summary_cell(ws, current_row, 1, rec["sheet"] if split=="train" else "", align="left")
            _summary_cell(ws, current_row, 2, split, fill=PatternFill("solid", start_color=SPLIT_COLORS.get(split)))
            nt_counts = rec["news_type_counts"].get(split, {})
            for ci, nt in enumerate(news_types, 3):
                _summary_cell(ws, current_row, ci, nt_counts.get(nt, 0), fill=PatternFill("solid", start_color=NEWS_TYPE_COLORS.get(nt)))
            _summary_cell(ws, current_row, len(news_types)+3, sum(nt_counts.values()), bold=True, fill=SUMMARY_TOTAL_FILL)
            current_row += 1
    current_row += 1

    # §4 TOPIC BREAKDOWN
    _summary_cell(ws, current_row, 1, "TOPIC BREAKDOWN", bold=True, fill=SUMMARY_HEADER_FILL)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(topics)+3)
    current_row += 1
    for ci, h in enumerate(["Sheet", "Split"] + topics + ["Total"], 1):
        _summary_cell(ws, current_row, ci, h, bold=True, fill=SUMMARY_SECTION_FILL)
    current_row += 1
    for rec in summary_records:
        for split in splits:
            _summary_cell(ws, current_row, 1, rec["sheet"] if split=="train" else "", align="left")
            _summary_cell(ws, current_row, 2, split, fill=PatternFill("solid", start_color=SPLIT_COLORS.get(split)))
            tp_counts = rec["topic_counts"].get(split, {})
            for ci, topic in enumerate(topics, 3):
                _summary_cell(ws, current_row, ci, tp_counts.get(topic, 0))
            _summary_cell(ws, current_row, len(topics)+3, sum(tp_counts.values()), bold=True, fill=SUMMARY_TOTAL_FILL)
            current_row += 1
    current_row += 1

    # §5 TOKEN STATISTICS — ORIGINAL DATASET
    if original_token_stats:
        stat_news_types = [nt for nt in news_types if nt in original_token_stats]
        n_stat_cols = 6  # count | avg | median | std | min | max
        total_cols  = 1 + n_stat_cols

        _summary_cell(ws, current_row, 1, "TOKEN STATISTICS — ORIGINAL DATASET (before sampling)",
                      bold=True, fill=SUMMARY_HEADER_FILL)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=total_cols)
        current_row += 1

        stat_headers = ["News Type", "Count", "Avg Tokens", "Median", "Std Dev", "Min", "Max"]
        for ci, h in enumerate(stat_headers, 1):
            _summary_cell(ws, current_row, ci, h, bold=True, fill=SUMMARY_SECTION_FILL)
        ws.column_dimensions[ws.cell(row=current_row, column=1).column_letter].width = 16
        current_row += 1

        for nt in stat_news_types:
            s = original_token_stats[nt]
            nt_fill = PatternFill("solid", start_color=NEWS_TYPE_COLORS.get(nt, "FFFFFF"))
            _summary_cell(ws, current_row, 1, nt, bold=True, fill=nt_fill, align="left")
            _summary_cell(ws, current_row, 2, s["count"],           fill=nt_fill)
            _summary_cell(ws, current_row, 3, f"{s['avg']:.1f}",    fill=nt_fill)
            _summary_cell(ws, current_row, 4, f"{s['median']:.1f}", fill=nt_fill)
            _summary_cell(ws, current_row, 5, f"{s['std']:.1f}",    fill=nt_fill)
            _summary_cell(ws, current_row, 6, s["min"],             fill=nt_fill)
            _summary_cell(ws, current_row, 7, s["max"],             fill=nt_fill)
            current_row += 1

        if "overall" in original_token_stats:
            s = original_token_stats["overall"]
            _summary_cell(ws, current_row, 1, "OVERALL", bold=True, fill=OVERALL_ROW_FILL, align="left")
            _summary_cell(ws, current_row, 2, s["count"],           fill=OVERALL_ROW_FILL, bold=True)
            _summary_cell(ws, current_row, 3, f"{s['avg']:.1f}",    fill=OVERALL_ROW_FILL, bold=True)
            _summary_cell(ws, current_row, 4, f"{s['median']:.1f}", fill=OVERALL_ROW_FILL, bold=True)
            _summary_cell(ws, current_row, 5, f"{s['std']:.1f}",    fill=OVERALL_ROW_FILL, bold=True)
            _summary_cell(ws, current_row, 6, s["min"],             fill=OVERALL_ROW_FILL, bold=True)
            _summary_cell(ws, current_row, 7, s["max"],             fill=OVERALL_ROW_FILL, bold=True)
            current_row += 1

        current_row += 1

    # §6 TOKEN STATISTICS — AFTER TRUNCATION (whole dataset)
    if truncated_token_stats:
        stat_news_types = [nt for nt in news_types if nt in truncated_token_stats]
        n_stat_cols = 6
        total_cols  = 1 + n_stat_cols

        cap_str = f"{truncation_cap}" if truncation_cap is not None else "N/A"
        desc_str = f" ({truncation_desc})" if truncation_desc else ""
        _summary_cell(ws, current_row, 1,
                      f"TOKEN STATISTICS — AFTER TRUNCATION (whole dataset, cap = {cap_str}{desc_str})",
                      bold=True, fill=SUMMARY_HEADER_FILL)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=total_cols)
        current_row += 1

        stat_headers = ["News Type", "Count", "Avg Tokens", "Median", "Std Dev", "Min", "Max"]
        for ci, h in enumerate(stat_headers, 1):
            _summary_cell(ws, current_row, ci, h, bold=True, fill=SUMMARY_SECTION_FILL)
        current_row += 1

        for nt in stat_news_types:
            s = truncated_token_stats[nt]
            nt_fill = PatternFill("solid", start_color=NEWS_TYPE_COLORS.get(nt, "FFFFFF"))
            _summary_cell(ws, current_row, 1, nt, bold=True, fill=nt_fill, align="left")
            _summary_cell(ws, current_row, 2, s["count"],           fill=nt_fill)
            _summary_cell(ws, current_row, 3, f"{s['avg']:.1f}",    fill=nt_fill)
            _summary_cell(ws, current_row, 4, f"{s['median']:.1f}", fill=nt_fill)
            _summary_cell(ws, current_row, 5, f"{s['std']:.1f}",    fill=nt_fill)
            _summary_cell(ws, current_row, 6, s["min"],             fill=nt_fill)
            _summary_cell(ws, current_row, 7, s["max"],             fill=nt_fill)
            current_row += 1

        if "overall" in truncated_token_stats:
            s = truncated_token_stats["overall"]
            _summary_cell(ws, current_row, 1, "OVERALL", bold=True, fill=OVERALL_ROW_FILL, align="left")
            _summary_cell(ws, current_row, 2, s["count"],           fill=OVERALL_ROW_FILL, bold=True)
            _summary_cell(ws, current_row, 3, f"{s['avg']:.1f}",    fill=OVERALL_ROW_FILL, bold=True)
            _summary_cell(ws, current_row, 4, f"{s['median']:.1f}", fill=OVERALL_ROW_FILL, bold=True)
            _summary_cell(ws, current_row, 5, f"{s['std']:.1f}",    fill=OVERALL_ROW_FILL, bold=True)
            _summary_cell(ws, current_row, 6, s["min"],             fill=OVERALL_ROW_FILL, bold=True)
            _summary_cell(ws, current_row, 7, s["max"],             fill=OVERALL_ROW_FILL, bold=True)
            current_row += 1

        current_row += 1

    # §7 TOKEN STATISTICS — PER CONDITION (MEAN ± per news_type)
    nt_cols = news_types
    total_header_cols = 2 + len(nt_cols) + 3

    _summary_cell(ws, current_row, 1, "TOKEN STATISTICS — PER CONDITION",
                  bold=True, fill=SUMMARY_HEADER_FILL)
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=total_header_cols)
    current_row += 1

    cond_headers = ["Sheet", "Split"] + [f"{nt} Avg" for nt in nt_cols] + ["Overall Avg", "Truncation Cap", "Truncation Method"]
    for ci, h in enumerate(cond_headers, 1):
        _summary_cell(ws, current_row, ci, h, bold=True, fill=SUMMARY_SECTION_FILL)
    current_row += 1

    for rec in summary_records:
        has_token = "token_stats" in rec and rec["token_stats"]
        for split in splits:
            _summary_cell(ws, current_row, 1, rec["sheet"] if split == "train" else "", align="left")
            _summary_cell(ws, current_row, 2, split,
                          fill=PatternFill("solid", start_color=SPLIT_COLORS.get(split)))

            nt_avgs_this_split = rec.get("token_stats_by_nt", {}).get(split, {})
            overall_avg = rec["token_stats"].get(split, {}).get("avg", "-") if has_token else "-"

            for ci, nt in enumerate(nt_cols, 3):
                nt_fill = PatternFill("solid", start_color=NEWS_TYPE_COLORS.get(nt, "FFFFFF"))
                val = nt_avgs_this_split.get(nt)
                display = f"{val:.1f}" if val is not None else "-"
                _summary_cell(ws, current_row, ci, display, fill=nt_fill)

            overall_display = f"{overall_avg:.1f}" if isinstance(overall_avg, float) else "-"
            cap_col    = 3 + len(nt_cols)
            method_col = cap_col + 1
            _summary_cell(ws, current_row, cap_col - 1, overall_display)
            _summary_cell(ws, current_row, cap_col,     rec.get("cap", "-") if split == "train" else "-")
            _summary_cell(ws, current_row, method_col,  rec.get("cap_desc", "-") if split == "train" else "-",
                          align="left")
            current_row += 1

    ws.freeze_panes = "A2"


# ── MAIN LOOP ─────────────────────────────────────────────────────────────────

def main(config_path):
    cfg = load_config(config_path)
    seed, output_file = cfg.get("seed", 42), cfg.get("output_file", "stratified_dataset.xlsx")
    total_samples, topic_ratios = cfg["total_samples"], cfg["topic_ratios"]
    ratio_configs = cfg["news_type_ratios"]

    sc = ratios_to_counts({"test": 0.15, "val": 0.15, "train": 0.70}, total_samples)
    n_tv, n_vv, n_tr = sc["test"], sc["val"], sc["train"]

    if "merge_sources" in cfg:
        frames = load_from_sources(cfg["merge_sources"])
    else:
        frames = load_from_file(cfg["input_file"], cfg.get("sheet_names", {"HR":"HR","HF":"HF","AI-F":"AI-F","AI-R":"AI-R"}))

    news_types, eq_ratio = sorted(frames.keys()), equal_ratio(list(frames.keys()))
    rng_fixed = np.random.default_rng(seed)
    used = {nt: set() for nt in frames}

    # ── Tokenize original frames once (when truncation is requested) ──────────
    original_token_stats: dict | None = None
    truncated_token_stats: dict | None = None
    tokenized_frames: dict | None     = None
    global_cap: int | None       = None
    global_cap_desc: str | None  = None

    if "truncation" in cfg:
        print("Tokenizing original dataset for baseline token statistics…")
        tokenized_frames     = tokenize_frames(frames, cfg["truncation"]["model"])
        original_token_stats = compute_original_token_stats(tokenized_frames)
        print("  Original token statistics computed.")

        if "cap" in cfg["truncation"]:
            global_cap      = int(cfg["truncation"]["cap"])
            global_cap_desc = f"fixed cap={global_cap}"
            print(f"  Using fixed truncation cap = {global_cap}")
        else:
            global_cap, global_cap_desc = compute_global_cap(tokenized_frames)

        # Compute truncated stats for the whole dataset (all rows, all types)
        truncated_token_stats = compute_truncated_token_stats(tokenized_frames, global_cap)
        print("  Truncated (whole dataset) token statistics computed.")

        # Drop _ids column to free memory
        for df in tokenized_frames.values():
            df.drop(columns=["_ids"], inplace=True)

    print("Sampling fixed test/val sets...")
    test_parts = [sample_stratified(frames[nt], c, topic_ratios, rng_fixed, used[nt]).assign(split="test") for nt, c in ratios_to_counts(eq_ratio, n_tv).items()]
    for p in test_parts: used[p["news_type"].iloc[0]].update(p["_idx"].tolist())
    val_parts = [sample_stratified(frames[nt], c, topic_ratios, rng_fixed, used[nt]).assign(split="val") for nt, c in ratios_to_counts(eq_ratio, n_vv).items()]
    for p in val_parts: used[p["news_type"].iloc[0]].update(p["_idx"].tolist())

    test_df, val_df = pd.concat(test_parts, ignore_index=True), pd.concat(val_parts, ignore_index=True)

    wb = Workbook()
    wb.remove(wb.active)
    summary_records = []

    for idx, type_ratios in enumerate(ratio_configs, 1):
        name = sheet_label(type_ratios)
        print(f"Generating sheet: {name}")
        train_counts = ratios_to_counts(type_ratios, n_tr)
        train_parts = [sample_stratified(frames[nt], c, topic_ratios, np.random.default_rng(seed+idx*1000), used[nt]).assign(split="train") for nt, c in train_counts.items()]
        train_df = pd.concat(train_parts, ignore_index=True).sample(frac=1, random_state=seed)

        combined = pd.concat([train_df, val_df, test_df], ignore_index=True).drop(columns=["_idx"])

        cap, cap_desc, token_stats, token_stats_by_nt = None, None, {}, {}
        if "truncation" in cfg:
            combined, cap, cap_desc = apply_truncation(
                combined, cfg["truncation"],
                cap_override=global_cap,
                cap_desc_override=global_cap_desc,
            )

            # Overall avg per split
            for split in ["train", "val", "test"]:
                split_df = combined[combined["split"] == split]
                token_stats[split] = {"avg": float(split_df["token_count"].mean())}

            # Per-news-type avg per split
            for split in ["train", "val", "test"]:
                split_df = combined[combined["split"] == split]
                nt_avgs  = {}
                for nt in news_types:
                    nt_rows = split_df[split_df["news_type"] == nt]["token_count"]
                    if not nt_rows.empty:
                        nt_avgs[nt] = float(nt_rows.mean())
                token_stats_by_nt[split] = nt_avgs

        summary_records.append({
            "sheet": name, "n_train": len(train_df), "n_val": len(val_df), "n_test": len(test_df),
            "news_type_counts": {s: combined[combined["split"]==s]["news_type"].value_counts().to_dict() for s in ["train","val","test"]},
            "topic_counts":     {s: combined[combined["split"]==s]["topic"].value_counts().to_dict()     for s in ["train","val","test"]},
            "token_stats":      token_stats,
            "token_stats_by_nt": token_stats_by_nt,
            "cap": cap, "cap_desc": cap_desc,
        })
        write_sheet(wb.create_sheet(title=name), combined)

    write_summary_sheet(
        wb.create_sheet(title="Summary"),
        summary_records, news_types, topic_ratios,
        total_samples, n_tr, n_vv, n_tv,
        original_token_stats=original_token_stats,
        truncated_token_stats=truncated_token_stats,
        truncation_cap=global_cap,
        truncation_desc=global_cap_desc,
    )
    wb.save(output_file)
    print(f"✅ Success! Saved to {output_file}")

if __name__ == "__main__":
    if len(sys.argv) == 2: main(sys.argv[1])
    else: print("Usage: python stratify.py config.json")