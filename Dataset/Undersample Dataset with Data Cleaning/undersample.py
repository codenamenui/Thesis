"""
undersample.py
--------------
Produces an Excel workbook with two data sheets — one for HR and one for HF —
each containing their balanced, undersampled rows, plus a Summary sheet with
counts and percentages for every breakdown.

Undersampling logic:
  For each topic, find the minimum row count between HR and HF sheets.
  Sample exactly that many rows from EACH sheet for that topic.
  This ensures:
    - Topics are balanced within each news type (HR and HF have equal counts per topic)
    - HR and HF are balanced against each other

Output sheets:
  - HR-Undersampled : balanced HR rows
  - HF-Undersampled : balanced HF rows
  - Summary         : counts and percentages for every breakdown

Optional cleaning (requires clean.py in the same directory):
  Set "clean": true in the config to run the full clean.py pipeline on
  the 'article' column before undersampling. This eliminates the need for
  a separate cleaning step.

Usage:
    python undersample.py config_undersample.json

Config keys:
    input_file   (required) path to the source Excel file
    output_file  (optional) default: undersampled_dataset.xlsx
    seed         (optional) default: 42
    sheet_names  (optional) default: {"HR": "HR", "HF": "HF"}
    clean        (optional) default: false — set true to apply cleaning pipeline
"""

import json
import sys
import pathlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from clean import clean_article as _clean_article
    _CLEAN_AVAILABLE = True
except ImportError:
    _CLEAN_AVAILABLE = False


# ── config loading ────────────────────────────────────────────────────────────

def load_config(path: str) -> dict:
    with open(path) as f:
        cfg = json.load(f)
    if not cfg.get("input_file"):
        raise ValueError("'input_file' is required in the config.")
    return cfg


# ── data loading ──────────────────────────────────────────────────────────────

def load_sheets(
    input_file: str,
    sheet_map: dict,
    apply_cleaning: bool = False,
) -> dict[str, pd.DataFrame]:
    if apply_cleaning and not _CLEAN_AVAILABLE:
        raise ImportError(
            "'clean': true is set in config but clean.py could not be imported. "
            "Ensure clean.py is in the same directory as undersample.py."
        )

    xl = pd.read_excel(input_file, sheet_name=None)
    frames = {}

    for news_type, sheet_name in sheet_map.items():
        if sheet_name not in xl:
            raise KeyError(
                f"Sheet '{sheet_name}' (news type '{news_type}') not found. "
                f"Available: {list(xl.keys())}"
            )
        df = xl[sheet_name].copy()
        df.columns = df.columns.str.lower()

        missing = {"label", "article", "topic"} - set(df.columns)
        if missing:
            raise ValueError(f"Sheet '{sheet_name}' missing columns: {missing}")

        if apply_cleaning:
            original      = df["article"].astype(str)
            df["article"] = original.apply(_clean_article)
            n_changed     = (original != df["article"]).sum()
            chars_before  = original.str.len().sum()
            chars_after   = df["article"].str.len().sum()
            print(
                f"  [{news_type}] cleaning: {n_changed}/{len(df)} rows changed, "
                f"{chars_before - chars_after:,} chars removed "
                f"({(chars_before - chars_after) / max(chars_before, 1):.1%})"
            )

        df["topic"] = df["topic"].str.strip().str.lower()
        df["_idx"]  = range(len(df))
        frames[news_type] = df.reset_index(drop=True)

    return frames


# ── undersampling ─────────────────────────────────────────────────────────────

def compute_topic_caps(frames: dict[str, pd.DataFrame]) -> dict[str, int]:
    topic_counts: dict[str, dict[str, int]] = {}

    for news_type, df in frames.items():
        for topic, group in df.groupby("topic"):
            topic_counts.setdefault(topic, {})[news_type] = len(group)

    caps = {}
    for topic, counts in topic_counts.items():
        min_count = min(counts.values())
        caps[topic] = min_count
        count_str = "  ".join(f"{nt}={n}" for nt, n in sorted(counts.items()))
        print(f"  topic '{topic}': {count_str}  → cap={min_count}")

    return caps


def undersample_sheet(
    df: pd.DataFrame,
    caps: dict[str, int],
    rng: np.random.Generator,
    label: str = "",
) -> pd.DataFrame:
    parts = []

    for topic, cap in caps.items():
        if cap == 0:
            continue
        topic_pool = df[df["topic"] == topic]
        if topic_pool.empty:
            print(f"  [WARN]{label} topic '{topic}' not found — skipped.")
            continue
        n_draw = min(cap, len(topic_pool))
        if n_draw < cap:
            print(
                f"  [WARN]{label} topic '{topic}': "
                f"wanted {cap}, only {n_draw} available."
            )
        parts.append(
            topic_pool.sample(n=n_draw, random_state=int(rng.integers(1 << 31)))
        )

    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=df.columns)


# ── Excel styles ──────────────────────────────────────────────────────────────

COL_ORDER  = ["label", "article", "topic"]
COL_WIDTHS = {"label": 8, "article": 80, "topic": 28}

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)

SECTION_FILL = PatternFill("solid", start_color="1F4E79")
SECTION_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)

SUBHDR_FILL  = PatternFill("solid", start_color="2E75B6")
SUBHDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)

TOTAL_FILL   = PatternFill("solid", start_color="D6DCE4")
EVEN_FILL    = PatternFill("solid", start_color="EEF3FB")
ODD_FILL     = PatternFill("solid", start_color="FFFFFF")
DATA_FONT    = Font(name="Arial", size=10)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


# ── data sheet writer ─────────────────────────────────────────────────────────

def write_data_sheet(ws, df: pd.DataFrame) -> None:
    cols = [c for c in COL_ORDER if c in df.columns]
    df   = df[cols]

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    for ri, row in enumerate(df.itertuples(index=False), 2):
        row_fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = DATA_FONT
            cell.fill      = row_fill
            cell.alignment = Alignment(
                wrap_text=(cols[ci - 1] == "article"), vertical="top"
            )

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── summary sheet writer ──────────────────────────────────────────────────────

def _cell(ws, row, col, value, font=None, fill=None, alignment=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:          c.font          = font
    if fill:          c.fill          = fill
    if alignment:     c.alignment     = alignment
    if number_format: c.number_format = number_format
    return c


def _section_title(ws, row, col, title, n_cols):
    _cell(ws, row, col, title, font=SECTION_FONT, fill=SECTION_FILL, alignment=LEFT)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + n_cols - 1)
    ws.row_dimensions[row].height = 20
    return row + 1


def _table_headers(ws, row, col, headers):
    for i, h in enumerate(headers):
        _cell(ws, row, col + i, h, font=SUBHDR_FONT, fill=SUBHDR_FILL, alignment=CENTER)
    ws.row_dimensions[row].height = 18
    return row + 1


def _data_row(ws, row, col, values, even=True, bold=False, fill_override=None):
    fill = fill_override or (EVEN_FILL if even else ODD_FILL)
    fnt  = Font(name="Arial", bold=bold, size=10)
    for i, v in enumerate(values):
        align = LEFT if i == 0 else CENTER
        fmt   = "0.0%" if isinstance(v, float) and 0 <= v <= 1 else None
        _cell(ws, row, col + i, v, font=fnt, fill=fill, alignment=align, number_format=fmt)
    return row + 1


def write_summary_sheet(
    ws,
    combined: pd.DataFrame,
    frames_raw: dict[str, pd.DataFrame],
    caps: dict[str, int],
) -> None:
    for col_letter, width in zip("ABCDEFGHI", [32, 14, 14, 14, 14, 14, 14, 12, 12]):
        ws.column_dimensions[col_letter].width = width

    total  = len(combined)
    topics = sorted(combined["topic"].unique())

    ROW = 1

    # ── 1. OVERVIEW ───────────────────────────────────────────────────────────
    ROW = _section_title(ws, ROW, 1, "1 · Dataset Overview", 4)
    ROW = _table_headers(ws, ROW, 1, ["Metric", "Value"])

    for i, (k, v) in enumerate([
        ("Total rows (final)", total),
        ("Unique topics",      len(topics)),
    ]):
        ROW = _data_row(ws, ROW, 1, [k, v], even=(i % 2 == 0))

    ROW += 1

    # ── 2. TOPIC BREAKDOWN ────────────────────────────────────────────────────
    ROW = _section_title(ws, ROW, 1, "2 · Topic Breakdown", 3)
    ROW = _table_headers(ws, ROW, 1, ["Topic", "Count", "% of Total"])

    for i, topic in enumerate(topics):
        n = len(combined[combined["topic"] == topic])
        ROW = _data_row(ws, ROW, 1, [topic, n, n / total], even=(i % 2 == 0))
    ROW = _data_row(ws, ROW, 1, ["TOTAL", total, 1.0], bold=True, fill_override=TOTAL_FILL)
    ROW += 1

    # ── 3. PRE vs POST UNDERSAMPLING ──────────────────────────────────────────
    all_news_types = sorted(frames_raw.keys())
    hdr = ["Topic"]
    for nt in all_news_types:
        hdr += [f"{nt} Before", f"{nt} After", f"{nt} Dropped"]
    hdr += ["Cap (per type)"]

    ROW = _section_title(ws, ROW, 1,
        "3 · Pre- vs Post-Undersampling Topic Counts (per news type)", len(hdr))
    ROW = _table_headers(ws, ROW, 1, hdr)

    for ci in range(2, len(hdr) + 2):
        letter = get_column_letter(ci)
        if ws.column_dimensions[letter].width < 14:
            ws.column_dimensions[letter].width = 14

    for i, topic in enumerate(topics):
        cap = caps.get(topic, 0)
        row_vals = [topic]
        for nt in all_news_types:
            before = len(frames_raw[nt][frames_raw[nt]["topic"] == topic])
            row_vals += [before, cap, before - cap]
        row_vals.append(cap)
        ROW = _data_row(ws, ROW, 1, row_vals, even=(i % 2 == 0))

    grand_vals = ["TOTAL"]
    for nt in all_news_types:
        before_total = sum(len(frames_raw[nt][frames_raw[nt]["topic"] == t]) for t in topics)
        after_total  = sum(caps.get(t, 0) for t in topics)
        grand_vals  += [before_total, after_total, before_total - after_total]
    grand_vals.append(sum(caps.get(t, 0) for t in topics))
    _data_row(ws, ROW, 1, grand_vals, bold=True, fill_override=TOTAL_FILL)

    ws.freeze_panes = "B1"


# ── main ──────────────────────────────────────────────────────────────────────

def main(config_path: str) -> None:
    cfg            = load_config(config_path)
    seed           = cfg.get("seed", 42)
    input_file     = cfg["input_file"]
    output_file    = cfg.get("output_file", "undersampled_dataset.xlsx")
    sheet_map      = cfg.get("sheet_names", {"HR": "HR", "HF": "HF"})
    apply_cleaning = cfg.get("clean", False)

    print(f"\n📂  Input  : {input_file}")
    print(f"📄  Output : {output_file}")
    print(f"🌱  Seed   : {seed}")
    print(f"🧹  Clean  : {'yes (clean.py)' if apply_cleaning else 'no'}\n")

    if apply_cleaning:
        print("Cleaning articles …")
    frames = load_sheets(input_file, sheet_map, apply_cleaning=apply_cleaning)
    if apply_cleaning:
        print()

    print("Computing per-topic row caps (min across HR and HF) …")
    caps = compute_topic_caps(frames)
    total_per_sheet = sum(caps.values())
    total_rows      = total_per_sheet * len(frames)

    print(f"\n  Rows per sheet after undersampling : {total_per_sheet}")
    print(f"  Sheets                             : {list(frames.keys())}")
    print(f"  Total rows                         : {total_rows}\n")

    # ── undersample each sheet ────────────────────────────────────────────────
    rng = np.random.default_rng(seed)
    sampled_frames = {}

    for news_type in frames:
        print(f"Sampling [{news_type}] …")
        sampled = undersample_sheet(
            frames[news_type], caps, rng, label=f" [{news_type}]"
        )
        sampled.drop(columns=["_idx"], inplace=True)
        sampled_frames[news_type] = sampled
        print(f"  → {len(sampled)} rows\n")

    combined = pd.concat(sampled_frames.values(), ignore_index=True)

    # ── write workbook: one sheet per news type + Summary ─────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    for news_type in sorted(sampled_frames.keys()):
        sheet_df = sampled_frames[news_type].reset_index(drop=True)
        ws = wb.create_sheet(title=f"{news_type}-Undersampled")
        write_data_sheet(ws, sheet_df)
        print(f"  Sheet '{news_type}-Undersampled': {len(sheet_df)} rows")

    ws_sum = wb.create_sheet(title="Summary")
    write_summary_sheet(ws_sum, combined, frames, caps)

    out_path = pathlib.Path(output_file)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    sheets = [f"{nt}-Undersampled" for nt in sorted(sampled_frames.keys())] + ["Summary"]
    print(f"\n✅  Saved → {out_path}  (sheets: {', '.join(repr(s) for s in sheets)})")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python undersample.py config_undersample.json")
        sys.exit(1)
    main(sys.argv[1])