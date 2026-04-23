"""
clean.py
--------
Cleans the 'article' column of an Excel or CSV file containing news data.

Cleaning pipeline (applied in order):

  1. HTML entity decoding
       ↳ Decodes HTML-encoded characters (e.g. &amp; → &, &nbsp; → space).
         Removes encoding artifacts common in scraped Philippine news content
         (Cruz et al., 2020).

  2. HTML artifact removal
       ↳ Removes <script> and <style> blocks (content + tags), then strips
         all remaining HTML markup. Addresses structural noise introduced
         during web scraping (Reyes et al., 2025).

  3. Emoji removal
       ↳ Strips all Unicode emoji and pictograph characters, which carry no
         linguistic content in formal news text.

  4. Non-content element removal
       ↳ Removes boilerplate that is not part of the news body. Covered
         under Reyes et al. (2025):
           - Source attributions  : "Source: CNN Philippines", etc.
           - Author bylines       : "By John Doe", "Reporter: Name", etc.

  5. URL replacement
       ↳ Hyperlinks and raw URLs are replaced with the placeholder token
         [LINK]. Consistent with Cruz et al. (2020); prevents meaningless
         subword fragmentation without discarding the positional signal that
         a link was present.

  6. Email address removal
       ↳ Regex-based removal of email addresses, which carry no news content.

  7. Whitespace normalization
       ↳ Unicode whitespace variants (non-breaking spaces, zero-width chars,
         etc.) are normalised to plain spaces or dropped. Consecutive spaces
         and tabs are collapsed; leading/trailing whitespace is stripped.
         Resolves formatting inconsistencies from web scraping (Reyes et al.,
         2025).

  8. UTF-8 verification
       ↳ Text is NFC-normalised. Characters from the Latin Extended and
         General Punctuation Unicode blocks (common in Filipino news) are
         preserved.

Steps deliberately EXCLUDED:
  - Stopword removal — kept for full BERT context (Cruz et al., 2020).
  - Stemming         — skipped; BERT's subword tokenizer handles morphology.

Usage:
    python clean.py <input.xlsx|csv> <output.xlsx>
    python clean.py <input.xlsx|csv>               # writes <input>_cleaned.xlsx
"""

import re
import sys
import html
import pathlib
import unicodedata

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── compiled regex patterns ───────────────────────────────────────────────────

# 2. HTML artifact removal
_RE_SCRIPT    = re.compile(r"<script[^>]*>.*?</script>", re.IGNORECASE | re.DOTALL)
_RE_STYLE     = re.compile(r"<style[^>]*>.*?</style>",  re.IGNORECASE | re.DOTALL)
_RE_HTML_TAGS = re.compile(r"<[^>]+>")

# 3. Emoji — all Unicode emoji / pictograph ranges
_RE_EMOJI = re.compile(
    "["
    "\U0001F600-\U0001F64F"   # emoticons
    "\U0001F300-\U0001F5FF"   # symbols & pictographs
    "\U0001F680-\U0001F6FF"   # transport & map
    "\U0001F1E0-\U0001F1FF"   # flags
    "\U00002700-\U000027BF"   # dingbats
    "\U0001F900-\U0001F9FF"   # supplemental symbols
    "\U00002600-\U000026FF"   # miscellaneous symbols
    "\U0001FA00-\U0001FA6F"   # chess, etc.
    "\U0001FA70-\U0001FAFF"   # food, objects
    "]+",
    flags=re.UNICODE,
)

# 4a. Source attribution  ("Source: CNN Philippines", "Source: 1 | 2 | 3", etc.)
#     Always trailing — strip from keyword to end of string.
_RE_SOURCE = re.compile(r"\s*Sources?\s*:.*$", re.IGNORECASE | re.DOTALL)

# 4b. Author bylines and photo credits
#     "By John Doe", "Photo by Jane Smith", "Reporter: Name", "Written by Name"
_RE_BYLINE = re.compile(
    r"(?:(?:Photo|Image|Video|Written|Reported|Story)?\s*[Bb]y\s+[A-Z][a-zA-Z\s\-\.]{2,40})"
    r"|(?:(?:Reporter|Photographer|Author|Writer)\s*:\s*[A-Z][a-zA-Z\s\-\.]{2,40})",
    re.MULTILINE,
)

# 5. URLs → replaced with [LINK] placeholder (Cruz et al., 2020)
_RE_URL = re.compile(r"https?://\S+|www\.\S+", re.IGNORECASE)

# 6. Email addresses
_RE_EMAIL = re.compile(r"\S+@\S+\.\S+")

# 7. Unicode whitespace variants → plain space / drop
_UNICODE_SPACES = str.maketrans({
    "\u00a0": " ",   # non-breaking space
    "\u200b": "",    # zero-width space
    "\u200c": "",    # zero-width non-joiner
    "\u200d": "",    # zero-width joiner
    "\u2009": " ",   # thin space
    "\u202f": " ",   # narrow no-break space
    "\u3000": " ",   # ideographic space
    "\ufeff": "",    # BOM
})

# Whitespace collapsing helpers
_RE_MULTI_SPACE = re.compile(r"[ \t]+")
_RE_MULTI_NL    = re.compile(r"\n{3,}")


# ── core cleaning function ────────────────────────────────────────────────────

def clean_article(text: str) -> str:
    """
    Apply the preprocessing pipeline to a single article string.
    Returns the cleaned string, or the original value if it is not a string.
    """
    if not isinstance(text, str):
        return text

    # 1. HTML entity decoding  (e.g. &amp; → &, &nbsp; → space)
    text = html.unescape(text)

    # 2. HTML artifact removal
    text = _RE_SCRIPT.sub(" ", text)
    text = _RE_STYLE.sub(" ", text)
    text = _RE_HTML_TAGS.sub(" ", text)

    # 3. Emoji removal
    text = _RE_EMOJI.sub(" ", text)

    # 4. Non-content element removal
    text = _RE_SOURCE.sub("", text)          # 4a. source attributions (trailing)
    text = _RE_BYLINE.sub(" ", text)         # 4b. author bylines / photo credits

    # 5. URL replacement with [LINK] placeholder  (Cruz et al., 2020)
    text = _RE_URL.sub("[LINK]", text)

    # 6. Email address removal
    text = _RE_EMAIL.sub(" ", text)

    # 7 & 8. Whitespace normalisation and UTF-8 / NFC verification
    text = text.translate(_UNICODE_SPACES)
    text = unicodedata.normalize("NFC", text)
    text = _RE_MULTI_SPACE.sub(" ", text)
    text = "\n".join(line.strip() for line in text.splitlines())
    text = _RE_MULTI_NL.sub("\n\n", text)
    text = text.strip()

    return text


# ── statistics ────────────────────────────────────────────────────────────────

def diff_stats(original: pd.Series, cleaned: pd.Series) -> dict:
    changed    = (original != cleaned).sum()
    total      = len(original)
    orig_chars = original.str.len().sum()
    cln_chars  = cleaned.str.len().sum()
    return {
        "total_rows":        total,
        "rows_changed":      int(changed),
        "pct_changed":       changed / total if total else 0,
        "chars_before":      int(orig_chars),
        "chars_after":       int(cln_chars),
        "chars_removed":     int(orig_chars - cln_chars),
        "pct_chars_removed": (orig_chars - cln_chars) / orig_chars if orig_chars else 0,
    }


# ── Excel writing ─────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(name="Arial", size=10)
SPLIT_COLORS = {"train": "E2EFDA", "val": "FFF2CC", "test": "FCE4D6"}
COL_ORDER    = ["label", "article", "topic", "news_type", "split"]
COL_WIDTHS   = {"label": 8, "article": 80, "topic": 28, "news_type": 12, "split": 10}


def write_sheet(ws, df: pd.DataFrame) -> None:
    cols = [c for c in COL_ORDER if c in df.columns]
    cols += [c for c in df.columns if c not in cols]
    df = df[cols]

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    has_split = "split" in cols
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = (
            PatternFill("solid", start_color=SPLIT_COLORS.get(getattr(row, "split", ""), "FFFFFF"))
            if has_split
            else PatternFill("solid", start_color="EEF3FB" if ri % 2 == 0 else "FFFFFF")
        )
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.alignment = Alignment(
                wrap_text=(cols[ci - 1] == "article"), vertical="top"
            )

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── main ──────────────────────────────────────────────────────────────────────

def main(input_path: str, output_path: str) -> None:
    print(f"\n📂  Input : {input_path}")
    print(f"📄  Output: {output_path}\n")

    ext = pathlib.Path(input_path).suffix.lower()
    raw_sheets = (
        {"Sheet1": pd.read_csv(input_path)}
        if ext == ".csv"
        else pd.read_excel(input_path, sheet_name=None)
    )

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in raw_sheets.items():
        df = df.copy()
        df.columns = df.columns.str.lower()

        if "article" not in df.columns:
            print(f"['{sheet_name}']  no 'article' column — copied unchanged.")
            ws = wb.create_sheet(title=sheet_name)
            write_sheet(ws, df)
            continue

        original      = df["article"].astype(str)
        df["article"] = original.apply(clean_article)

        s = diff_stats(original, df["article"])
        print(f"['{sheet_name}']  {s['total_rows']} rows")
        print(f"  Rows changed      : {s['rows_changed']}  ({s['pct_changed']:.1%})")
        print(f"  Characters removed: {s['chars_removed']:,}  ({s['pct_chars_removed']:.1%})\n")

        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, df)

    wb.save(output_path)
    print(f"✅  Saved → {output_path}")


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print("Usage: python clean.py <input.xlsx|csv> [output.xlsx]")
        sys.exit(1)

    inp = sys.argv[1]
    out = (
        sys.argv[2]
        if len(sys.argv) == 3
        else str(
            pathlib.Path(inp)
            .with_stem(pathlib.Path(inp).stem + "_cleaned")
            .with_suffix(".xlsx")
        )
    )
    main(inp, out)
