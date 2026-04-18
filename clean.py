"""
clean.py
--------
Removes the trailing "Source: ..." attribution from the 'article' column
and applies basic whitespace cleanup.

The source line always appears at the end of the article text, e.g.:
    "... ayon sa post ni Abay noong 2017. Source: 1 | 2 | 3 | 4"

Usage:
    python clean.py <input.xlsx> <output.xlsx>
    python clean.py <input.xlsx>               # writes <input>_cleaned.xlsx
"""

import re
import sys
import pathlib
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── cleaning ──────────────────────────────────────────────────────────────────

# Matches "Source:" (or "Sources:") and everything after it until end of string.
# Handles variations like "Source: 1 | 2 | 3", "Source: CNN Philippines", etc.
_RE_SOURCE = re.compile(r"\s*Sources?\s*:.*$", re.IGNORECASE | re.DOTALL)


def clean_article(text: str) -> str:
    if not isinstance(text, str):
        return text
    text = _RE_SOURCE.sub("", text)
    return text.strip()


# ── Excel helpers ─────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(name="Arial", size=10)
SPLIT_COLORS = {"train": "E2EFDA", "val": "FFF2CC", "test": "FCE4D6"}
COL_ORDER    = ["label", "article", "topic", "news_type", "split"]
COL_WIDTHS   = {"label": 8, "article": 80, "topic": 28, "news_type": 12, "split": 10}


def write_sheet(ws, df: pd.DataFrame) -> None:
    cols = [c for c in COL_ORDER if c in df.columns]
    cols += [c for c in df.columns if c not in cols]
    df = df[cols]

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    has_split = "split" in cols
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = (PatternFill("solid",
                    start_color=SPLIT_COLORS.get(getattr(row, "split", ""), "FFFFFF"))
                if has_split
                else PatternFill("solid",
                    start_color="EEF3FB" if ri % 2 == 0 else "FFFFFF"))
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.alignment = Alignment(
                wrap_text=(cols[ci - 1] == "article"), vertical="top"
            )

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ── main ──────────────────────────────────────────────────────────────────────

def main(input_path: str, output_path: str) -> None:
    print(f"\n📂  Input : {input_path}")
    print(f"📄  Output: {output_path}\n")

    xl = pd.read_excel(input_path, sheet_name=None)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in xl.items():
        df.columns = df.columns.str.lower()

        if "article" not in df.columns:
            print(f"['{sheet_name}']  no 'article' column — copied unchanged.")
        else:
            original      = df["article"].astype(str)
            df["article"] = original.apply(clean_article)
            changed       = (original != df["article"]).sum()
            print(f"['{sheet_name}']  {len(df)} rows  |  {changed} sources removed")

        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, df)

    wb.save(output_path)
    print(f"\n✅  Saved → {output_path}")


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print("Usage: python clean.py <input.xlsx> [output.xlsx]")
        sys.exit(1)

    inp = sys.argv[1]
    out = (sys.argv[2] if len(sys.argv) == 3
           else str(pathlib.Path(inp).with_stem(pathlib.Path(inp).stem + "_cleaned")))
    main(inp, out)
